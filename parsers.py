from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Any


TXT_EXPECTED_COLUMNS = [
    "ARTICULO",
    "CORTE",
    "FECHA",
    "programa",
    "PROCESO",
    "BODEGA",
    "SALDO",
    "CORTE",
    "TALLER",
    "T.EXTERNO",
    "LIMPIADO",
    "LAVANDERIA",
    "TERMINACION",
    "MUESTRA",
    "SEGUNDA",
    "TALLER",
]


def parse_uploaded_file(file_storage) -> dict[str, Any]:
    filename = (file_storage.filename or "").lower()
    content = file_storage.read()

    if filename.endswith(".txt") or filename.endswith(".csv"):
        kind = detect_txt_kind(content, filename)
        if kind == "corte_etapas":
            return {"kind": "corte_etapas", "rows": parse_corte_etapas_txt(content)}
        if kind == "pedidos_talla_todas":
            return {"kind": "pedidos_talla_todas", "rows": parse_pedidos_talla_todas_txt(content)}
        if kind == "pedidos_talla":
            return {"kind": "pedidos_talla", "rows": parse_pedidos_talla_txt(content)}
        if kind == "comparativo_clientes":
            return {"kind": "comparativo_clientes", "rows": parse_comparativo_clientes_txt(content)}
        return {"kind": "saldos", "rows": parse_saldos_txt(content)}
    if filename.endswith(".xlsx"):
        kind = detect_xlsx_kind(content, filename)
        if kind == "exs_map":
            return {"kind": "exs_map", "rows": parse_exs_xlsx(content)}
        return {"kind": "saldos", "rows": parse_saldos_xlsx(content)}
    raise ValueError("Formato no soportado. Usa .txt, .csv o .xlsx")


def detect_xlsx_kind(content: bytes, filename: str) -> str:
    if "exs" in filename:
        return "exs_map"
    try:
        from openpyxl import load_workbook
    except ImportError:
        return "saldos"

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    header = None
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        cells = ["" if c is None else str(c).strip().lower() for c in row]
        if any(cells):
            header = cells
            break
    if not header:
        return "saldos"

    first = header[0] if len(header) > 0 else ""
    second = header[1] if len(header) > 1 else ""
    if ("actual" in first or "familia actual" in first) and ("ex" in second):
        return "exs_map"
    return "saldos"


def detect_txt_kind(content: bytes, filename: str) -> str:
    if "pedidosxtallatodas" in filename:
        return "pedidos_talla_todas"
    if "pedidosxtalla" in filename:
        return "pedidos_talla"
    text = _decode_bytes(content)
    first_line = ""
    for line in text.splitlines():
        if line.strip():
            first_line = line.strip()
            break
    if not first_line:
        return "saldos"
    normalized_header = (
        first_line.replace("\ufeff", "").replace('"', "").replace(" ", "").upper()
    )
    if normalized_header.startswith("O.CORTE;FECHA;ARTICULO"):
        return "corte_etapas"
    if normalized_header.startswith("RUT;RAZONSOCIAL;CODVEN;VENDEDOR;CIUDAD;CANTIDADT:01"):
        return "comparativo_clientes"
    if normalized_header.startswith("ARTICULO;CORTE;FECHA"):
        return "saldos"
    if ";Ventas;" in first_line or ";Despacho;" in first_line or ";saldo;" in first_line:
        if "todas" in filename:
            return "pedidos_talla_todas"
        return "pedidos_talla"
    return "saldos"


def parse_saldos_txt(content: bytes) -> list[dict]:
    text = _decode_bytes(content)
    reader = csv.reader(io.StringIO(text), delimiter=";")
    rows = list(reader)
    if not rows:
        return []

    data_rows = rows[1:]
    parsed: list[dict] = []
    for raw in data_rows:
        if not raw or not any(cell.strip() for cell in raw):
            continue
        parsed_row = _map_txt_row(raw)
        if parsed_row:
            parsed.append(parsed_row)
    return parsed


def parse_comparativo_clientes_txt(content: bytes) -> list[dict]:
    text = _decode_bytes(content)
    reader = csv.reader(io.StringIO(text), delimiter=";")
    rows = list(reader)
    if len(rows) <= 1:
        return []

    header_raw = [str(h or "").strip().upper() for h in rows[0]]

    def _idx(col_name: str, default: int) -> int:
        normalized = col_name.replace(" ", "")
        for i, h in enumerate(header_raw):
            if h.replace(" ", "") == normalized:
                return i
        return default

    idx_rut = _idx("RUT", 0)
    idx_razon = _idx("RAZON SOCIAL", 1)
    idx_cod = _idx("COD VEN", 2)
    idx_vendedor = _idx("VENDEDOR", 3)
    idx_ciudad = _idx("CIUDAD", 4)
    idx_c_t01 = _idx("CANTIDAD T:01", 5)
    idx_v_t01 = _idx("VALOR T:01", 6)
    idx_f_t01 = _idx("FACTURADO T:01", 7)
    idx_vf_t01 = _idx("VALOR FACT. T:01", 8)
    idx_c_t02 = _idx("CANTIDAD T:02", 9)
    idx_v_t02 = _idx("VALOR T:02", 10)
    idx_f_t02 = _idx("FACTURADO T:02", 11)
    idx_vf_t02 = _idx("VALOR FACT. T:02", 12)
    idx_c_t03 = _idx("CANTIDAD T:03", 13)
    idx_v_t03 = _idx("VALOR T:03", 14)
    idx_f_t03 = _idx("FACTURADO T:03", 15)
    idx_vf_t03 = _idx("VALOR FACT. T:03", 16)

    parsed: list[dict] = []
    for raw in rows[1:]:
        if not raw or not any(str(c).strip() for c in raw):
            continue
        cells = [str(c).strip() for c in raw]
        rut = cells[idx_rut] if idx_rut < len(cells) else ""
        if not rut:
            continue
        parsed.append(
            {
                "rut": rut,
                "razon_social": cells[idx_razon] if idx_razon < len(cells) else "",
                "cod_vendedor": cells[idx_cod] if idx_cod < len(cells) else "",
                "vendedor": cells[idx_vendedor] if idx_vendedor < len(cells) else "",
                "ciudad": cells[idx_ciudad] if idx_ciudad < len(cells) else "",
                "cantidad_t01": _to_int(cells[idx_c_t01] if idx_c_t01 < len(cells) else "0"),
                "valor_t01": _to_int(cells[idx_v_t01] if idx_v_t01 < len(cells) else "0"),
                "facturado_t01": _to_int(cells[idx_f_t01] if idx_f_t01 < len(cells) else "0"),
                "valor_fact_t01": _to_int(cells[idx_vf_t01] if idx_vf_t01 < len(cells) else "0"),
                "cantidad_t02": _to_int(cells[idx_c_t02] if idx_c_t02 < len(cells) else "0"),
                "valor_t02": _to_int(cells[idx_v_t02] if idx_v_t02 < len(cells) else "0"),
                "facturado_t02": _to_int(cells[idx_f_t02] if idx_f_t02 < len(cells) else "0"),
                "valor_fact_t02": _to_int(cells[idx_vf_t02] if idx_vf_t02 < len(cells) else "0"),
                "cantidad_t03": _to_int(cells[idx_c_t03] if idx_c_t03 < len(cells) else "0"),
                "valor_t03": _to_int(cells[idx_v_t03] if idx_v_t03 < len(cells) else "0"),
                "facturado_t03": _to_int(cells[idx_f_t03] if idx_f_t03 < len(cells) else "0"),
                "valor_fact_t03": _to_int(cells[idx_vf_t03] if idx_vf_t03 < len(cells) else "0"),
            }
        )
    return parsed


def parse_pedidos_talla_txt(content: bytes) -> list[dict]:
    text = _decode_bytes(content)
    reader = csv.reader(io.StringIO(text), delimiter=";")
    parsed: list[dict] = []

    for raw in reader:
        if not raw:
            continue
        cells = [str(c).strip() for c in raw]
        if not any(cells):
            continue
        if len(cells) < 6:
            continue
        if not cells[0]:
            continue

        articulo = _clean_code(cells[0])
        descripcion = cells[1] if len(cells) > 1 else ""
        tipo = (cells[2] if len(cells) > 2 else "").strip().lower()
        if not tipo:
            continue

        qty_start = 4 if len(cells) > 4 else 3
        qty_cells = [c for c in cells[qty_start:] if c != ""]
        if not qty_cells:
            continue
        total = _to_int(qty_cells[-1])
        tallas = [_to_int(c) for c in qty_cells[:-1]]

        parsed.append(
            {
                "articulo": articulo,
                "descripcion": descripcion,
                "tipo": tipo,
                "tallas": tallas,
                "total": total,
            }
        )

    return parsed


def parse_pedidos_talla_todas_txt(content: bytes) -> list[dict]:
    text = _decode_bytes(content)
    reader = csv.reader(io.StringIO(text), delimiter=";")
    parsed: list[dict] = []

    for raw in reader:
        if not raw:
            continue
        cells = [str(c).strip() for c in raw]
        if not any(cells):
            continue
        if len(cells) < 6:
            continue
        if not cells[0]:
            continue

        articulo = _clean_code(cells[0])
        descripcion = cells[1] if len(cells) > 1 else ""
        tipo = (cells[2] if len(cells) > 2 else "").strip().lower()
        if not tipo:
            continue

        qty_start = 4 if len(cells) > 4 else 3
        qty_cells = [c for c in cells[qty_start:] if c != ""]
        if not qty_cells:
            continue
        total = _to_int_signed(qty_cells[-1])
        tallas = [_to_int_signed(c) for c in qty_cells[:-1]]

        parsed.append(
            {
                "articulo": articulo,
                "descripcion": descripcion,
                "tipo": tipo,
                "tallas": tallas,
                "total": total,
            }
        )

    return parsed


def parse_corte_etapas_txt(content: bytes) -> list[dict]:
    text = _decode_bytes(content)
    reader = csv.reader(io.StringIO(text), delimiter=";")
    rows = list(reader)
    if not rows:
        return []

    parsed: list[dict] = []
    for raw in rows[1:]:
        if not raw or not any(str(c).strip() for c in raw):
            continue
        cells = [str(c).strip() for c in raw]
        if len(cells) < 35:
            continue

        corte = _clean_code(cells[0])
        if not corte:
            continue

        parsed.append(
            {
                "corte": corte,
                "fecha_orden_iso": _parse_date(cells[1] if len(cells) > 1 else ""),
                "articulo": _clean_code(cells[2]) if len(cells) > 2 else "",
                "programado": _to_int(cells[3] if len(cells) > 3 else "0"),
                "cortado": _to_int(cells[4] if len(cells) > 4 else "0"),
                "entrega": _to_int(cells[5] if len(cells) > 5 else "0"),
                "saldo": _to_int(cells[6] if len(cells) > 6 else "0"),
                "corte_inicio_iso": _parse_date(cells[7] if len(cells) > 7 else ""),
                "corte_fin_iso": _parse_date(cells[8] if len(cells) > 8 else ""),
                "taller_inicio_iso": _parse_date(cells[11] if len(cells) > 11 else ""),
                "taller_fin_iso": _parse_date(cells[12] if len(cells) > 12 else ""),
                "t_externo_inicio_iso": _parse_date(cells[15] if len(cells) > 15 else ""),
                "t_externo_fin_iso": _parse_date(cells[16] if len(cells) > 16 else ""),
                "limpiado_inicio_iso": _parse_date(cells[19] if len(cells) > 19 else ""),
                "limpiado_fin_iso": _parse_date(cells[20] if len(cells) > 20 else ""),
                "lavanderia_inicio_iso": _parse_date(cells[23] if len(cells) > 23 else ""),
                "lavanderia_fin_iso": _parse_date(cells[24] if len(cells) > 24 else ""),
                "terminacion_inicio_iso": _parse_date(cells[27] if len(cells) > 27 else ""),
                "terminacion_fin_iso": _parse_date(cells[28] if len(cells) > 28 else ""),
                "muestra_inicio_iso": _parse_date(cells[31] if len(cells) > 31 else ""),
                "muestra_fin_iso": _parse_date(cells[32] if len(cells) > 32 else ""),
            }
        )
    return parsed


def parse_exs_xlsx(content: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Para importar Excel instala dependencias: pip install -r requirements.txt") from exc

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    parsed: list[dict] = []
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        if not row:
            continue
        actual_raw = "" if row[0] is None else str(row[0]).strip()
        ex_raw = "" if len(row) < 2 or row[1] is None else str(row[1]).strip()
        if not actual_raw and not ex_raw:
            continue
        actual_digits = "".join(ch for ch in actual_raw if ch.isdigit())
        ex_digits = "".join(ch for ch in ex_raw if ch.isdigit())
        if not actual_digits:
            continue
        actual = actual_digits[-4:] if len(actual_digits) >= 4 else actual_digits
        # EX puede venir como 416901 / 416900 / 4169-01.
        # Se preserva visualmente el valor original y el calculo usa familia (4 digitos).
        ex = ex_raw if ex_raw else ex_digits
        parsed.append({"actual": actual, "ex": ex})

    return parsed


def parse_saldos_xlsx(content: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Para importar Excel instala dependencias: pip install -r requirements.txt") from exc

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    parsed: list[dict] = []
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        if not row or not any(cell is not None and str(cell).strip() for cell in row):
            continue
        raw = ["" if cell is None else str(cell) for cell in row]
        parsed_row = _map_txt_row(raw)
        if parsed_row:
            parsed.append(parsed_row)
    return parsed


def parse_lavanderia_productividad_xlsx(content: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Para importar Excel instala dependencias: pip install -r requirements.txt") from exc

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    if not wb.worksheets:
        return []
    ws = wb.worksheets[0]

    header_idx = None
    header: list[str] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = ["" if c is None else str(c).strip() for c in row]
        first = cells[0].strip().lower() if cells else ""
        second = cells[1].strip().lower() if len(cells) > 1 else ""
        if first == "articulo" and second in {"o.corte", "o corte", "ocorte"}:
            header_idx = idx
            header = cells
            break
    if not header_idx:
        return []

    ignore_headers = {
        "",
        "ingreso",
        "inicio",
        "hr inicio",
        "hr termino",
        "salida",
        "cantidad",
        "tiempo",
        "proceso",
        "lavandería",
        "lavanderia",
        "terminación",
        "terminacion",
    }

    stage_columns: list[int] = []
    for i, name in enumerate(header):
        n = str(name or "").strip().lower()
        if n in ignore_headers:
            continue
        if i < 3:
            continue
        if i + 1 < len(header) and str(header[i + 1] or "").strip().lower() == "cantidad":
            stage_columns.append(i)

    parsed: list[dict] = []
    for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
        vals = [None if v is None else v for v in row]
        if not vals:
            continue
        articulo = _clean_code(vals[0] if len(vals) > 0 else "")
        corte = _clean_code(vals[1] if len(vals) > 1 else "")
        bota = str(vals[2]).strip() if len(vals) > 2 and vals[2] is not None else ""
        if not articulo and not corte:
            continue

        base_fecha = _parse_excel_date(vals[4] if len(vals) > 4 else None)
        for c in stage_columns:
            etapa = str(header[c] or "").strip()
            if not etapa:
                continue
            empleado_raw = vals[c] if c < len(vals) else None
            empleado = str(empleado_raw).strip() if empleado_raw is not None else ""
            if not empleado:
                continue

            cantidad = _to_int(vals[c + 1] if c + 1 < len(vals) else 0)
            minutos = _parse_minutes(vals[c + 2] if c + 2 < len(vals) else None)
            hora_inicio = _parse_excel_time(vals[c - 1] if c - 1 >= 0 and c - 1 < len(vals) else None)
            hora_fin = _parse_excel_time(vals[c + 3] if c + 3 < len(vals) else None)
            fecha_fin = _parse_excel_date(vals[c + 4] if c + 4 < len(vals) else None) or base_fecha

            if cantidad <= 0 and minutos <= 0 and not hora_inicio and not hora_fin:
                continue

            parsed.append(
                {
                    "articulo": articulo,
                    "corte": corte,
                    "bota": bota,
                    "etapa": etapa,
                    "empleado": empleado,
                    "cantidad": max(cantidad, 0),
                    "minutos": max(minutos, 0),
                    "fecha_inicio_iso": base_fecha,
                    "hora_inicio": hora_inicio,
                    "fecha_fin_iso": fecha_fin,
                    "hora_fin": hora_fin,
                }
            )

    return parsed


def parse_lavanderia_botas_maestros_xlsx(content: bytes) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Para importar Excel instala dependencias: pip install -r requirements.txt") from exc

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = None
    for sheet in wb.worksheets:
        if str(sheet.title or "").strip().lower().startswith("maestro"):
            ws = sheet
            break
    if ws is None:
        return []

    botas: list[str] = []
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        value = "" if row[0] is None else str(row[0]).strip()
        if not value:
            continue
        if value.lower() == "bota":
            continue
        botas.append(value)

    # Mantener orden y eliminar repetidos.
    seen: set[str] = set()
    unique: list[str] = []
    for bota in botas:
        key = bota.lower()
        if key in seen:
            continue
        seen.add(key)
        unique.append(bota)
    return unique


def parse_lavanderia_etapas_gestion_xlsx(content: bytes) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Para importar Excel instala dependencias: pip install -r requirements.txt") from exc

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    if not wb.worksheets:
        return []
    ws = wb.worksheets[0]

    header: list[str] = []
    for row in ws.iter_rows(values_only=True):
        cells = ["" if c is None else str(c).strip() for c in row]
        first = cells[0].strip().lower() if cells else ""
        second = cells[1].strip().lower() if len(cells) > 1 else ""
        if first == "articulo" and second in {"o.corte", "o corte", "ocorte"}:
            header = cells
            break
    if not header:
        return []

    ignore_headers = {
        "",
        "articulo",
        "o.corte",
        "o corte",
        "ocorte",
        "bota",
        "cantidad",
        "ingreso",
        "inicio",
        "hr inicio",
        "hr termino",
        "salida",
        "tiempo",
        "proceso",
        "lavandería",
        "lavanderia",
        "terminación",
        "terminacion",
    }

    stages: list[str] = []
    for i, name in enumerate(header):
        n = str(name or "").strip()
        n_lower = n.lower()
        if n_lower in ignore_headers:
            continue
        if i + 1 < len(header) and str(header[i + 1] or "").strip().lower() == "cantidad":
            stages.append(n)

    seen: set[str] = set()
    unique: list[str] = []
    for stage in stages:
        key = stage.lower()
        if key in seen:
            continue
        seen.add(key)
        unique.append(stage)
    return unique


def _decode_bytes(content: bytes) -> str:
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("latin-1", errors="replace")


def _map_txt_row(raw: list[str]) -> dict | None:
    # Algunos archivos pueden traer columna final vacia por ';' de cierre.
    cells = [str(c).strip() for c in raw]
    if len(cells) < 16:
        return None

    # Mapeo por posicion basado en el archivo de ejemplo.
    articulo = _clean_code(cells[0])
    corte = _clean_code(cells[1])
    fecha_iso = _parse_date(cells[2])

    return {
        "articulo": articulo,
        "corte": corte,
        "fecha_iso": fecha_iso,
        "programa": _to_int(cells[3] if len(cells) > 3 else "0"),
        "proceso": _to_int(cells[4] if len(cells) > 4 else "0"),
        "bodega": _to_int(cells[5] if len(cells) > 5 else "0"),
        "saldo": _to_int(cells[6] if len(cells) > 6 else "0"),
        "corte_1": _to_int(cells[7] if len(cells) > 7 else "0"),
        "taller": _to_int(cells[8] if len(cells) > 8 else "0"),
        "t_externo": _to_int(cells[9] if len(cells) > 9 else "0"),
        "limpiado": _to_int(cells[10] if len(cells) > 10 else "0"),
        "lavanderia": _to_int(cells[11] if len(cells) > 11 else "0"),
        "terminacion": _to_int(cells[12] if len(cells) > 12 else "0"),
        "muestra": _to_int(cells[13] if len(cells) > 13 else "0"),
        "segunda": _to_int(cells[14] if len(cells) > 14 else "0"),
        "taller_nombre": cells[15].strip() if len(cells) > 15 else "",
    }


def _clean_code(value: str) -> str:
    return str(value).strip()


def _to_int(value: str) -> int:
    s = str(value).strip().replace(".", "").replace(",", "")
    if not s:
        return 0
    try:
        return int(s)
    except ValueError:
        digits = "".join(ch for ch in s if ch.isdigit())
        return int(digits) if digits else 0


def _parse_excel_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _parse_excel_time(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%H:%M:%S")
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(s, fmt).strftime("%H:%M:%S")
        except ValueError:
            continue
    if s.count(".") == 1 and s.replace(".", "").isdigit():
        parts = s.split(".")
        hh = int(parts[0])
        mm = int(parts[1][:2] or "0")
        if 0 <= hh <= 23 and 0 <= mm <= 59:
            return f"{hh:02d}:{mm:02d}:00"
    return None


def _parse_minutes(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, datetime):
        return value.hour * 60 + value.minute + (1 if value.second > 0 else 0)
    s = str(value).strip()
    if not s:
        return 0
    if ":" in s:
        parts = s.split(":")
        if len(parts) >= 2:
            try:
                hh = int(parts[0])
                mm = int(parts[1])
                ss = int(parts[2]) if len(parts) > 2 else 0
                return max(hh * 60 + mm + (1 if ss > 0 else 0), 0)
            except ValueError:
                return 0
    try:
        # Excel puede guardar horas como decimal (0.2 = 12 minutos aprox) o 1.5 (1h30).
        as_float = float(s.replace(",", "."))
    except ValueError:
        return 0
    if as_float <= 0:
        return 0
    return int(round(as_float * 60)) if as_float <= 24 else int(round(as_float))


def _to_int_signed(value: str) -> int:
    s = str(value).strip().replace(".", "").replace(",", "")
    if not s:
        return 0
    sign = -1 if s.startswith("-") else 1
    digits = "".join(ch for ch in s if ch.isdigit())
    return sign * int(digits) if digits else 0


def _parse_date(value: str) -> str | None:
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None
