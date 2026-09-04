"""
Actualiza los 4 archivos JSON con datos reales desde Z:\BI
Ejecutar: python actualizar_datos.py
"""
import csv, json, os
from datetime import datetime, date

TODAY = date.today()
BI = r"Z:\BI"
OUT      = os.path.dirname(os.path.abspath(__file__))
DOCS_OUT = os.path.join(os.path.dirname(OUT), 'docs')

TEMP_MIN, TEMP_MAX = 27, 99  # Todas las temporadas disponibles
def temp_valida(t):
    try: return TEMP_MIN <= int(t) <= TEMP_MAX
    except: return False

def parse_date(s):
    for fmt in ['%d/%m/%Y','%d-%m-%Y','%Y-%m-%d']:
        try: return datetime.strptime(str(s).strip(), fmt).date()
        except: pass
    return None

def to_iso(s): d = parse_date(s); return d.strftime('%Y-%m-%d') if d else ''
def fmt_date(s): d = parse_date(s); return d.strftime('%d/%m/%Y') if d else ''
def clean_int(s):
    try: return int(str(s).strip().replace(' ',''))
    except: return 0
def has_date(s):
    s = str(s).strip()
    return 1 if (s and '    ' not in s and parse_date(s)) else 0
def dias_desde(s):
    d = parse_date(s)
    return (TODAY - d).days if d else 0
def stage_dias(ini, fin):
    d1, d2 = parse_date(ini), parse_date(fin)
    if d1 and d2: return (d2 - d1).days or 1
    return 1 if d1 else 0

# ── 1. TRAZABILIDAD ────────────────────────────────────────────
print("Leyendo TRAZABILIDAD2.CSV...")
with open(f"{BI}/TRAZABILIDAD2.CSV", encoding="latin-1") as f:
    traza_rows = list(csv.DictReader(f, delimiter=';'))

full_table = []
traza_oc   = []
for r in traza_rows:
    art  = r['Articulo'].strip()
    temp = art[2:4] if len(art) >= 4 else ''
    if not temp_valida(temp):
        continue
    tipo_r = r['Tipo'].strip().upper()
    tipo = 'Muestras' if 'MUESTRA' in tipo_r else ('Set' if 'SET' in tipo_r else 'Producción')

    full_table.append({
        "articulo": art, "corte": r['O.Corte'].strip(),
        "fecha": fmt_date(r['Fecha']), "fecha_iso": to_iso(r['Fecha']),
        "temporada": temp, "tipo": tipo, "m": r['Muestra'].strip(),
        "programa": clean_int(r['Programado']), "proceso": clean_int(r['Cortado']),
        "bodega": clean_int(r['Entrega']), "saldo": clean_int(r['Saldo']),
        "corte_u": has_date(r['Corte']), "taller": has_date(r['Taller']),
        "texterno": has_date(r['Taller Ext']), "limpiado": has_date(r['Limpiado']),
        "lavanderia": has_date(r['Lavander']), "terminacion": has_date(r['Terminacion']),
        "muestra": 1 if r['Muestras'].strip() else 0, "segunda": 0
    })

    stage_cols = [("Corte","Corte"),("Taller","Taller"),("Taller Ext","Taller Ext"),
                  ("Limpiado","Limpiado"),("Lavander","Lavander"),("Terminacion","Terminacion")]
    stages = []
    prev_fin = ''
    for name, col in stage_cols:
        ini_raw = str(r.get(col,'')).strip()
        ini_raw = '' if '    ' in ini_raw else ini_raw
        fin_raw = ''
        if name == "Corte":      fin_raw = ini_raw
        elif name == "Lavander": fin_raw = str(r.get('Terminacion','')).strip()
        fin_raw = '' if '    ' in fin_raw else fin_raw
        stages.append({"name": name,
                        "ini": fmt_date(ini_raw) if ini_raw else "",
                        "fin": fmt_date(fin_raw) if fin_raw else "",
                        "dias": stage_dias(ini_raw, fin_raw)})

    traza_oc.append({
        "oc": r['O.Corte'].strip(), "tipo": r['Tipo'].strip(),
        "m": r['Muestra'].strip(), "fecha": fmt_date(r['Fecha']),
        "articulo": art, "prog": clean_int(r['Programado']),
        "cort": clean_int(r['Cortado']), "ent": clean_int(r['Entrega']),
        "saldo": clean_int(r['Saldo']), "stages": stages,
        "totDias": clean_int(r['Tot.dias']), "inc": r.get('Incidencias','').strip()
    })

# ── 2. PEDIDOS ─────────────────────────────────────────────────
print("Leyendo PEDIDOS.CSV...")
with open(f"{BI}/PEDIDOS.CSV", encoding="latin-1") as f:
    ped_rows = list(csv.DictReader(f, delimiter=';'))

pedidos_dict = {}

for r in ped_rows:
    pid  = r['PEDIDO'].strip()
    art  = r.get('ARTICULO','').strip()
    temp = art[2:4] if len(art) >= 4 else r.get('TEMPORADA','').strip()[:2]
    if not temp_valida(temp):
        continue

    if pid not in pedidos_dict:
        pedidos_dict[pid] = {
            "pedido": pid, "fecha": fmt_date(r['FECHA']), "fecha_iso": to_iso(r['FECHA']),
            "rut": r['RUT'].strip(), "nombre": r['CLIENTE'].strip(),
            "ciudad": r['CIUDAD'].strip(), "vendedor": r.get('VENDEDOR','').strip(),
            "unidades": 0, "despacho": 0, "saldo": 0,
            "valor": 0, "valor_desp": 0, "valor_sal": 0,
            "dias": dias_desde(r['FECHA']), "temps": [], "u_temp": {}, "u_desp": {}, "u_sal": {},
            "_arts": {}
        }
    p = pedidos_dict[pid]
    if temp and temp not in p['temps']:
        p['temps'].append(temp)
    sol  = clean_int(r.get('SOLICITADO', 0))
    desp = clean_int(r.get('DESPACHADO', 0))
    sal  = clean_int(r.get('saldo', 0))
    try:
        # DCTO columna contiene códigos internos, no porcentajes → ignorar
        precio = float(str(r.get('PRECIO','0')).strip() or 0)
        p['valor']      += int(sol  * precio)
        p['valor_desp'] += int(desp * precio)
        p['valor_sal']  += int(sal  * precio)
    except: pass
    p['unidades'] += sol
    p['despacho'] += desp
    p['saldo']    += sal
    # Unidades desglosadas por temporada para que el filtro por temp sea exacto
    p['u_temp'][temp] = p['u_temp'].get(temp, 0) + sol
    p['u_desp'][temp] = p['u_desp'].get(temp, 0) + desp
    p['u_sal'][temp]  = p['u_sal'].get(temp, 0)  + sal
    # Detalle por artículo (art8 = sin talla)
    art8 = art[:8]
    if art8 and len(art8) == 8:
        if art8 not in p['_arts']:
            p['_arts'][art8] = {'sol': 0, 'desp': 0, 'sal': 0}
        p['_arts'][art8]['sol']  += sol
        p['_arts'][art8]['desp'] += desp
        p['_arts'][art8]['sal']  += sal

# Convertir _arts en lineas y limpiar clave interna
for p in pedidos_dict.values():
    arts = p.pop('_arts', {})
    p['lineas'] = sorted(
        [{'art': k, 'temp': k[2:4], 'modelo': k[2:6], 'color': k[6:8], **v}
         for k, v in arts.items() if v['sol'] > 0],
        key=lambda x: -(x['sal'])
    )

pedidos = list(pedidos_dict.values())

# Mapeo art8 → tipo de bota (desde SubCateg de PEDIDOS.CSV)
def norm_bota(sc):
    s = sc.strip().upper()
    if 'PITILLO' in s: return 'Pitillo'
    if 'FLARE'   in s: return 'Flare'
    if 'BOOTCUT' in s: return 'Bootcut'
    if 'WIDE LEG' in s: return 'Wide Leg'
    if 'OXFORD'  in s: return 'Oxford'
    if 'PALAZZO' in s: return 'Palazzo'
    if 'RECTO'   in s: return 'Recto'
    if 'BALLOON' in s: return 'Balloon'
    if 'BERMUDA' in s: return 'Bermuda'
    if 'CALZA'   in s: return 'Calza'
    return s.title() if s else ''

mod_bota = {}
for r in ped_rows:
    art8 = r.get('ARTICULO','').strip()[:8]
    sc   = r.get('SubCateg','').strip()
    if art8 and sc:
        mod_bota[art8] = norm_bota(sc)

# ── 2b. ARTÍCULOS POR TALLA (ARCHIVO_TALLAS.CSV) ───────────────────────────
# Fuente correcta para unidades pedidas por artículo/modelo.
# PEDIDOS.CSV suma SOLICITADO que incluye líneas no confirmadas → cifra mayor.
# ARCHIVO_TALLAS.CSV "Ventas" refleja las unidades reales de la temporada.
print("Leyendo ARCHIVO_TALLAS.CSV...")
art_dict = {}  # {temp: {base: {mod: qty}}}
tallas_fallback = False
try:
    with open(f"{BI}/ARCHIVO_TALLAS.CSV", encoding="latin-1") as f:
        for line in f:
            cells = [c.strip() for c in line.strip().split(';')]
            if len(cells) < 6 or not cells[0] or not cells[2]:
                continue
            if cells[3].lower() != 'ventas':
                continue
            art  = cells[0]
            temp = art[2:4] if len(art) >= 4 else ''
            if not temp_valida(temp):
                continue
            base = art[2:6] if len(art) >= 6 else ''
            mod  = art[6:8] if len(art) >= 8 else ''
            if not base:
                continue
            # El último campo no vacío (desde posición 4) es el total de la fila
            non_empty = [c for c in cells[4:] if c]
            if not non_empty:
                continue
            qty = clean_int(non_empty[-1])
            if temp not in art_dict:
                art_dict[temp] = {}
            if base not in art_dict[temp]:
                art_dict[temp][base] = {}
            art_dict[temp][base][mod] = art_dict[temp][base].get(mod, 0) + qty
except FileNotFoundError:
    print("  ARCHIVO_TALLAS.CSV no encontrado, usando PEDIDOS.CSV para artículos")
    tallas_fallback = True
    for r in ped_rows:
        art  = r.get('ARTICULO','').strip()
        temp = art[2:4] if len(art) >= 4 else r.get('TEMPORADA','').strip()[:2]
        if not temp_valida(temp):
            continue
        base = art[2:6] if len(art) >= 6 else ''
        mod  = art[6:8] if len(art) >= 8 else ''
        sol  = clean_int(r.get('SOLICITADO', 0))
        if base:
            if temp not in art_dict:
                art_dict[temp] = {}
            if base not in art_dict[temp]:
                art_dict[temp][base] = {}
            art_dict[temp][base][mod] = art_dict[temp][base].get(mod, 0) + sol

# Construir pedidos_art: lista de {temp, base, total, modelos:[{mod,qty}]}
pedidos_art = []
for temp, bases in art_dict.items():
    for base, mods in bases.items():
        total = sum(mods.values())
        pedidos_art.append({
            "temp": temp, "base": base, "total": total,
            "modelos": [{"mod": m, "qty": q} for m, q in sorted(mods.items())]
        })

# ── 3. VENTAS ──────────────────────────────────────────────────
# Total C/descto = neto con descuentos aplicados (pre-IVA)
# Bruto = neto × 1.19 para tipos afectos (Factura, Boleta)
EXENTOS = {'02', '34', '56'}
def calc_bruto(neto, tipo):
    return neto if tipo in EXENTOS else int(round(neto * 1.19))

print("Leyendo VENTAS-TOD-2026.CSV...")
with open(f"{BI}/VENTAS-TOD-2026.CSV", encoding="latin-1") as f:
    reader = csv.DictReader(f, delimiter=';')
    venta_rows = list(reader)

docs_dict = {}
for r in venta_rows:
    tipo  = str(r.get('Tipo') or '').strip()
    num   = str(r.get('Numero') or '').strip()
    bod   = str(r.get('Bod') or '').strip().zfill(2)
    # Todas las bodegas: 04 San Gerardo (mayorista) + 00 Central / 12 Outlet (retail boletas)
    key = (tipo, num)
    if key not in docs_dict:
        fecha = str(r.get('fecha') or '').strip()
        docs_dict[key] = {
            "tipo": tipo, "dcto": num, "bod": bod,
            "fecha": fmt_date(fecha), "fecha_iso": to_iso(fecha),
            "rut": str(r.get('Rut') or '').strip(),
            "razon": str(r.get('cliente') or '').strip(),
            "vendedor": str(r.get('Vendedor') or '').strip(),
            "prendas": 0, "neto": 0, "bruto": 0,
            "fpago": str(r.get('Fpago') or '').strip()
        }
    d = docs_dict[key]
    d['prendas'] += clean_int(r.get('Cant', 0))
    # Usar Total C/descto si existe, sino Total como fallback
    desc_raw = str(r.get('Total C/descto') or '').strip()
    tot_raw  = str(r.get('Total') or '0').strip()
    neto_val = clean_int(desc_raw) if desc_raw else clean_int(tot_raw)
    d['neto'] += neto_val

for d in docs_dict.values():
    d['bruto'] = calc_bruto(d['neto'], d['tipo'])
    d['total'] = d['neto']  # compatibilidad con vistas existentes

docs_venta = list(docs_dict.values())

# ── 4. SALDOS POR LOCAL (SALDOSXLOCAL.CSV desde Z:\BI) ────────
SUCURSALES_PRENDAS = {'01','02','04','05','10','12','33'}

print("Leyendo SALDOSXLOCAL.CSV...")
saldo_map = {}  # {art8: {'sucs': {suc: qty}, 'tallas': {talla: qty}}}
saldo_file = os.path.join(BI, 'SALDOSXLOCAL.CSV')
try:
    with open(saldo_file, encoding='latin-1') as f:
        reader = csv.DictReader(f, delimiter=';')
        for r in reader:
            code = str(r.get('Articulo', '')).strip()
            suc  = str(r.get('Bodega', '')).strip().zfill(2)
            try:
                qty = float(str(r.get('SaldoFisico', '0')).strip() or 0)
            except:
                continue
            if qty <= 0:
                continue
            if not code.startswith('01') or len(code) < 10:
                continue
            try:
                t_num = int(code[2:4])
            except:
                continue
            if not (40 <= t_num <= 44):
                continue
            art8  = code[:8]
            talla = code[8:10].lstrip('0') or code[8:10]
            try:
                cajas = float(str(r.get('Cajas', '0')).strip() or 0)
            except:
                cajas = 0
            if art8 not in saldo_map:
                saldo_map[art8] = {'sucs': {}, 'tallas': {}, 'cajas': {}, 'cajas_talla': {}}
            saldo_map[art8]['sucs'][suc] = saldo_map[art8]['sucs'].get(suc, 0) + qty
            saldo_map[art8]['cajas'][suc] = saldo_map[art8]['cajas'].get(suc, 0) + cajas
            if talla:
                saldo_map[art8]['tallas'][talla] = saldo_map[art8]['tallas'].get(talla, 0) + qty
                saldo_map[art8]['cajas_talla'][talla] = saldo_map[art8]['cajas_talla'].get(talla, 0) + cajas
except FileNotFoundError:
    print("  SALDOSXLOCAL.CSV no encontrado en Z:\\BI")

saldos_bodega = []
for art8, data in sorted(saldo_map.items()):
    sucs   = data['sucs']
    tallas = data['tallas']
    t      = art8[2:4]
    modelo = art8[2:6]
    color  = art8[6:8]
    total_prendas = sum(v for k, v in sucs.items() if k in SUCURSALES_PRENDAS)
    total_all = sum(sucs.values())
    if total_all <= 0:
        continue
    def _tsort(k):
        try: return int(k)
        except: return 999
    cajas_d = data.get('cajas', {})
    cajas_talla = data.get('cajas_talla', {})
    total_cajas = int(sum(v for k, v in cajas_d.items() if k in SUCURSALES_PRENDAS))
    tallas_sorted = {k: int(v) for k, v in sorted(tallas.items(), key=lambda x: _tsort(x[0]))}
    saldo_talla = {k: max(0, int(v) - int(cajas_talla.get(k, 0))) for k, v in tallas_sorted.items()}
    saldos_bodega.append({
        "art": art8, "temp": t, "modelo": modelo, "color": color,
        "suc": {k: int(v) for k, v in sucs.items()},
        "cajas": {k: int(v) for k, v in cajas_d.items()},
        "cajas_total": total_cajas,
        "saldo": max(0, int(total_prendas) - total_cajas),
        "tallas": tallas_sorted,
        "saldo_talla": saldo_talla,
        "prendas": int(total_prendas), "total": int(total_all),
        "bota": mod_bota.get(art8, '')
    })
saldos_bodega.sort(key=lambda x: -x['prendas'])

# ── 5. PVC EX MAPPING (COLE44_ORIGEN.xlsx + TRAZABILIDAD T40-T43) ─
import re as _re

# Saldo EX = Cortado - Entrega en T40-T43
corte_hist = {}; entrega_hist = {}
for r in traza_rows:
    art_r = r['Articulo'].strip()
    t_str = art_r[2:4] if len(art_r) >= 4 else ''
    try:
        t_num = int(t_str)
    except:
        continue
    if not (40 <= t_num <= 43):
        continue
    if 'PRODUCCION' not in r['Tipo'].strip().upper():
        continue
    key = art_r[2:6]
    corte_hist[key]   = corte_hist.get(key, 0)   + clean_int(r['Cortado'])
    entrega_hist[key] = entrega_hist.get(key, 0) + clean_int(r['Entrega'])

seed_dir = os.path.join(os.path.dirname(OUT), 'seed')
origen_xlsx = os.path.join(seed_dir, 'COLE44_ORIGEN.xlsx')
pvc_ex = {}
try:
    from openpyxl import load_workbook
    wb = load_workbook(origen_xlsx, read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=6, values_only=True):
        art_raw    = row[2] if len(row) > 2 else None
        origen_raw = row[3] if len(row) > 3 else None
        if not art_raw:
            continue
        parts  = str(art_raw).strip().split('-')
        modelo = parts[0].strip()
        if len(modelo) != 4:
            continue
        origen_str = str(origen_raw).strip() if origen_raw else ''
        m = _re.match(r'(?i)^EX\s*(\d{4})', origen_str)
        if not m:
            continue
        ex_base  = m.group(1)
        ex_saldo = max(0, corte_hist.get(ex_base, 0) - entrega_hist.get(ex_base, 0))
        if modelo not in pvc_ex:
            pvc_ex[modelo] = {"ex_base": ex_base, "ex_saldo": ex_saldo}
except Exception as e:
    print(f"  COLE44_ORIGEN.xlsx: {e}")

# ── Guardar ────────────────────────────────────────────────────
from datetime import datetime
NOW = datetime.now()
meta = {
    "updated_iso": NOW.strftime('%Y-%m-%dT%H:%M'),
    "updated_str": NOW.strftime('%d/%m/%Y %H:%M'),
    "updated_date": NOW.strftime('%d/%m'),
    "updated_time": NOW.strftime('%H:%M'),
}

DATASETS = [("full_table", full_table), ("traza_oc", traza_oc),
            ("pedidos", pedidos), ("docs_venta", docs_venta),
            ("pedidos_art", pedidos_art),
            ("saldos_bodega", saldos_bodega), ("pvc_ex", pvc_ex),
            ("meta", meta)]

for name, data in DATASETS:
    for dest in [OUT, DOCS_OUT]:
        os.makedirs(dest, exist_ok=True)
        path = os.path.join(dest, f"{name}.json")
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False)
    n = len(data) if isinstance(data, list) else 1
    print(f"  {name}.json -> {n} registros")

print("Datos actualizados. Recarga el navegador.")
