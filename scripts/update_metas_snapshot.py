from __future__ import annotations

import argparse
import json
import re
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


def _norm_text(value: object) -> str:
    txt = str(value or "").strip().lower()
    txt = re.sub(r"\s+", " ", txt)
    return txt


def _to_float(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    probe = str(value).strip().replace(".", "").replace(",", ".")
    try:
        return float(probe)
    except Exception:
        return 0.0


def build_snapshot(excel_path: Path) -> dict[str, object]:
    wb = load_workbook(excel_path, data_only=True, read_only=True)

    month_aliases = {
        1: "enero",
        2: "febrero",
        3: "marzo",
        4: "abril",
        5: "mayo",
        6: "junio",
        7: "julio",
        8: "agosto",
        9: "septiembre",
        10: "octubre",
        11: "noviembre",
        12: "diciembre",
    }
    current_label = f"{month_aliases.get(date.today().month, '')} {date.today().year}".strip()
    sheet_name = None
    for name in wb.sheetnames:
        if _norm_text(name) == _norm_text(current_label):
            sheet_name = name
            break

    if not sheet_name:
        candidates: list[tuple[int, int, str]] = []
        for name in wb.sheetnames:
            norm = _norm_text(name)
            if any(m in norm for m in month_aliases.values()):
                year_match = re.search(r"(20\d{2})", norm)
                year = int(year_match.group(1)) if year_match else 0
                month_idx = next((k for k, v in month_aliases.items() if v in norm), 0)
                candidates.append((year, month_idx, name))
        if candidates:
            candidates.sort()
            sheet_name = candidates[-1][2]

    if not sheet_name:
        raise ValueError("No se encontro una hoja mensual valida.")

    ws = wb[sheet_name]
    row2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), ())
    days_month = int(_to_float(row2[13] if len(row2) > 13 else 0) or 0)
    days_elapsed = int(_to_float(row2[16] if len(row2) > 16 else 0) or 0)
    days_remaining = max(days_month - days_elapsed, 0)

    section_key_map = {
        "corte": "corte",
        "urrutia": "urrutia",
        "tierra del fuego": "tierra_fuego",
        "lavanderia": "lavanderia",
        "terminacion": "terminacion",
    }
    sections: dict[str, dict[str, float]] = {}
    for r in range(5, 10):
        row = next(ws.iter_rows(min_row=r, max_row=r, values_only=True), ())
        raw_name = str(row[15] or "").strip() if len(row) > 15 else ""
        key = section_key_map.get(_norm_text(raw_name), "")
        if not key:
            continue
        sections[key] = {
            "name": raw_name,
            "meta_day": float(_to_float(row[13] if len(row) > 13 else 0)),
            "meta_month": float(_to_float(row[14] if len(row) > 14 else 0)),
            "projected": float(_to_float(row[16] if len(row) > 16 else 0)),
            "actual": float(_to_float(row[17] if len(row) > 17 else 0)),
            "daily_avg": float(_to_float(row[20] if len(row) > 20 else 0)),
        }

    week_starts = [4, 22, 40, 58, 76]
    week_area_rows = [
        ("corte", 0, "Corte"),
        ("urrutia", 3, "Urrutia"),
        ("sur", 6, "Sur"),
        ("lavanderia", 9, "Lavanderia"),
        ("terminacion", 12, "Terminacion"),
    ]
    weeks: list[dict[str, object]] = []
    for idx, start in enumerate(week_starts, start=1):
        hab_row = next(
            ws.iter_rows(min_row=max(start - 3, 1), max_row=max(start - 3, 1), values_only=True),
            (),
        )
        day_row = next(
            ws.iter_rows(min_row=max(start - 2, 1), max_row=max(start - 2, 1), values_only=True),
            (),
        )
        day_candidates = [
            (i, int(_to_float(v) or 0))
            for i, v in enumerate(day_row)
            if int(_to_float(v) or 0) > 0 and int(_to_float(v) or 0) <= 31
        ]
        best_run: list[tuple[int, int]] = []
        for i in range(len(day_candidates)):
            run = [day_candidates[i]]
            for j in range(i + 1, len(day_candidates)):
                prev_col, prev_val = run[-1]
                col, val = day_candidates[j]
                if col > prev_col and val == prev_val + 1:
                    run.append((col, val))
                elif val <= prev_val:
                    break
            if len(run) > len(best_run):
                best_run = run

        selected = best_run if len(best_run) >= 2 else day_candidates
        day_col_indices = [i for i, _ in selected]
        habiles = []
        fechas = []
        for i_col in day_col_indices:
            h_raw = hab_row[i_col] if i_col < len(hab_row) else None
            d_raw = day_row[i_col] if i_col < len(day_row) else None
            h_val = int(_to_float(h_raw) or 0)
            d_val = int(_to_float(d_raw) or 0)
            habiles.append(h_val if h_val > 0 else None)
            fechas.append(d_val if d_val > 0 else None)

        week_rows = []
        for area_key, offset, label in week_area_rows:
            row = next(ws.iter_rows(min_row=start + offset, max_row=start + offset, values_only=True), ())
            values = []
            for i_col in day_col_indices:
                raw = row[i_col] if i_col < len(row) else None
                num = int(_to_float(raw) or 0)
                values.append(num if num > 0 else None)
            total_cell = int(_to_float(row[10] if len(row) > 10 else 0) or 0)
            total = total_cell if total_cell > 0 else sum(int(v or 0) for v in values)
            week_rows.append({"key": area_key, "name": label, "values": values, "total": total})

        weeks.append(
            {
                "label": f"Semana {idx}",
                "habiles": habiles,
                "fechas": fechas,
                "rows": week_rows,
            }
        )

    return {
        "sheet_name": sheet_name,
        "days_month": days_month,
        "days_elapsed": days_elapsed,
        "days_remaining": days_remaining,
        "sections": sections,
        "weeks": weeks,
    }


def main() -> None:
    repo_root = Path(__file__).resolve().parent.parent
    default_excel = repo_root / "seed" / "1_PROGRAMAS DE PRODUCCION MHC .xlsx"
    default_output = repo_root / "seed" / "PROGRAMAS_MHC_SNAPSHOT.json"

    parser = argparse.ArgumentParser(description="Actualiza snapshot de metas desde Excel PROGRAMAS MHC.")
    parser.add_argument("--excel", type=Path, default=default_excel, help="Ruta al Excel fuente.")
    parser.add_argument("--output", type=Path, default=default_output, help="Ruta del JSON de salida.")
    args = parser.parse_args()

    snapshot = build_snapshot(args.excel)
    args.output.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Snapshot actualizado: {args.output}")
    print(f"Hoja usada: {snapshot['sheet_name']}")


if __name__ == "__main__":
    main()
