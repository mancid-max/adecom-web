from __future__ import annotations

import csv
import hashlib
import hmac
import io
import json
import os
import re
import threading
import time
import unicodedata
from datetime import date, datetime
from pathlib import Path
from difflib import SequenceMatcher
from urllib import error as url_error
from urllib import request as url_request

from flask import Flask, Response, flash, jsonify, redirect, render_template, request, session, url_for
from werkzeug.exceptions import RequestEntityTooLarge

from adecom_db import (
    add_lavanderia_registro,
    delete_inventory_stock_row,
    delete_lavanderia_registro,
    get_conn,
    import_lavanderia_rows,
    import_lavanderia_botas_maestro,
    import_lavanderia_etapas_maestro,
    import_corte_etapas_rows,
    import_comparativo_clientes_rows,
    import_deuda_clientes_rows,
    import_exs_map_rows,
    import_pedidos_talla_todas_rows,
    init_db,
    import_pedidos_talla_rows,
    import_rows,
    query_inventory_stock_rows,
    query_lavanderia_productividad,
    query_lavanderia_catalogos,
    query_assistant_rules,
    query_exs_balance_summary,
    query_comparativo_clientes,
    query_pedidos_talla_sections,
    query_rows,
    replace_inventory_stock_rows,
    save_inventory_stock_row,
)
from parsers import (
    parse_comparativo_clientes_txt,
    parse_deudas_vencidas_csv,
    parse_lavanderia_botas_maestros_xlsx,
    parse_lavanderia_etapas_gestion_xlsx,
    parse_lavanderia_productividad_xlsx,
    parse_pedidos_talla_txt,
    parse_saldos_txt,
    parse_uploaded_file,
)
from parsers import parse_corte_etapas_txt


BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
DB_PATH = os.environ.get("DATABASE_URL") or os.environ.get(
    "ADECOM_DB_PATH", str(BASE_DIR / "data" / "adecom.db")
)
SEED_DIR = BASE_DIR / "seed"
SEED_SALDOS = SEED_DIR / "SALDOS-SECCI.TXT"
SEED_PEDIDOS = SEED_DIR / "PEDIDOSXTALLA.TXT"
SEED_COMPARATIVO = SEED_DIR / "COMPARATIVO.Txt"
SEED_DEUDAS = SEED_DIR / "Deudas_Vencidas.CSV"
SEED_VENTAS_DOCS = SEED_DIR / "VENTAS-TOD-2026.CSV"
SEED_CORTES_4200_XLSX = SEED_DIR / "Cortes 4200.xlsx"
PROGRAMAS_MHC_PATH = Path(
    os.environ.get(
        "ADECOM_PROGRAMAS_MHC_XLSX",
        str(SEED_DIR / "1_PROGRAMAS DE PRODUCCION MHC .xlsx"),
    )
)
PROGRAMAS_MHC_SNAPSHOT_PATH = Path(
    os.environ.get(
        "ADECOM_PROGRAMAS_MHC_SNAPSHOT",
        str(SEED_DIR / "PROGRAMAS_MHC_SNAPSHOT.json"),
    )
)
INVENTORY_BOOK_PATH = Path(
    os.environ.get(
        "ADECOM_INVENTARIO_XLSX",
        r"C:\Users\manuh\OneDrive - Mohicano Jeans\INVENTARIO 01-04 COMPLETO.xlsx",
    )
)
AUTOLOAD_DIR = Path(
    os.environ.get(
        "ADECOM_AUTOLOAD_DIR",
        r"C:\Users\manuh\Desktop\APIS\Documentos a cargar ADECOM WEB",
    )
)
AUTOLOAD_DIR_FALLBACK = Path(os.environ.get("ADECOM_AUTOLOAD_DIR_FALLBACK", r"Z:\\"))
AUTOLOAD_SALDOS_SOURCE = os.environ.get("ADECOM_AUTOLOAD_SALDOS_SOURCE", "").strip()
AUTOLOAD_PEDIDOS_SOURCE = os.environ.get("ADECOM_AUTOLOAD_PEDIDOS_SOURCE", "").strip()
AUTOLOAD_ETAPAS_SOURCE = os.environ.get("ADECOM_AUTOLOAD_ETAPAS_SOURCE", "").strip()
AUTOLOAD_COMPARATIVO_SOURCE = os.environ.get("ADECOM_AUTOLOAD_COMPARATIVO_SOURCE", "").strip()
AUTO_REFRESH_WEB_ON_START = os.environ.get("ADECOM_AUTO_REFRESH_WEB_ON_START", "1").strip() == "1"
AUTO_REFRESH_WEB_POLL_SECONDS = max(int(os.environ.get("ADECOM_AUTO_REFRESH_WEB_POLL_SECONDS", "60").strip() or "60"), 0)
AUTO_REFRESH_WEB_BACKGROUND = os.environ.get("ADECOM_AUTO_REFRESH_WEB_BACKGROUND", "1").strip() == "1"
AUTO_REFRESH_WEB_DAILY_TIME = os.environ.get("ADECOM_AUTO_REFRESH_WEB_DAILY_TIME", "").strip()
AUTO_REFRESH_WEB_ONLY_DAILY = os.environ.get("ADECOM_AUTO_REFRESH_WEB_ONLY_DAILY", "0").strip() == "1"
ASSISTANT_ENABLED = os.environ.get("ADECOM_ASSISTANT_ENABLED", "0").strip() == "1"
NEW_SECTION_ENABLED = os.environ.get("ADECOM_ENABLE_NEW_SECTION", "0").strip() == "1"
OTHER_SECTION_ENABLED = os.environ.get("ADECOM_ENABLE_OTHER_SECTION", "1").strip() == "1"
ENABLE_SEED = os.environ.get("ADECOM_ENABLE_SEED", "1").strip() == "1"
PROYECCION_STATE_PATH = BASE_DIR / "data" / "proyeccion_personas.json"
AREA_WEIGHTS = {
    "CORTE": 600,
    "TALLER": 400,
    "TALLER EXTERNO": 200,
    "LIMPIADO": 600,
    "LAVANDERIA": 600,
    "TERMINACION": 600,
}
DEFAULT_LAVANDERIA_BOTAS = [
    "Oxford",
    "Flared",
    "Pitillo",
    "Recto",
    "Tobillero",
    "Cargo",
    "Palazzo",
    "Wide Legs",
    "Falda",
    "Falda/Short",
]

UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
if not str(DB_PATH).startswith(("postgres://", "postgresql://")):
    Path(DB_PATH).resolve().parent.mkdir(parents=True, exist_ok=True)

app = Flask(__name__)
app.secret_key = os.environ.get("ADECOM_SECRET_KEY", "dev-secret-change-me")
app.config["MAX_CONTENT_LENGTH"] = int(os.environ.get("ADECOM_MAX_UPLOAD_MB", "25")) * 1024 * 1024
PROYECCION_STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
_refresh_lock = threading.Lock()
_refresh_thread_started = False
_last_sources_signature = ""
_last_daily_refresh_date = ""
_last_refresh_mode = ""


@app.after_request
def ensure_utf8_charset(response):
    mimetype = (response.mimetype or "").lower()
    if mimetype in {
        "text/html",
        "text/plain",
        "text/css",
        "application/javascript",
        "application/json",
    }:
        response.headers["Content-Type"] = f"{mimetype}; charset=utf-8"
    return response


def _admin_key() -> str:
    return os.environ.get("ADECOM_ADMIN_KEY", "Mohicano1079@").strip()


def _access_key_web() -> str:
    return os.environ.get("ADECOM_ACCESS_KEY_WEB", "adecom-web").strip()


def _access_key_new() -> str:
    return os.environ.get("ADECOM_ACCESS_KEY_NEW", "adecom-nueva").strip()


def _access_key_other() -> str:
    return os.environ.get("ADECOM_ACCESS_KEY_OTHER", "adecom-landing").strip()


def _access_key_web_aliases() -> list[str]:
    raw = os.environ.get("ADECOM_ACCESS_KEY_WEB_ALIASES", "adecom-web,adecom,web")
    return [x.strip() for x in raw.split(",") if x.strip()]


def _match_any_key(entered: str, keys: list[str]) -> bool:
    if not entered:
        return False
    for key in keys:
        if key and hmac.compare_digest(entered, key):
            return True
    return False


def _match_any_key_ci(entered: str, keys: list[str]) -> bool:
    normalized = str(entered or "").strip().casefold()
    if not normalized:
        return False
    for key in keys:
        candidate = str(key or "").strip().casefold()
        if candidate and hmac.compare_digest(normalized, candidate):
            return True
    return False


def _portal_section() -> str:
    return str(session.get("portal_section") or "").strip().lower()


def _is_authenticated() -> bool:
    return _portal_section() in {"web", "new", "other"}


@app.before_request
def _guard_portal_routes():
    endpoint = request.endpoint or ""
    public_endpoints = {"login", "login_post", "static"}
    if endpoint in public_endpoints:
        return

    section = _portal_section()
    if not section:
        return redirect(url_for("login"))

    if section == "new":
        allowed_new = {"new_section", "logout", "static", "login"}
        if endpoint not in allowed_new:
            return redirect(url_for("new_section"))

    if section == "other":
        allowed_other = {
            "other_section",
            "other_import_excel",
            "other_add_registro",
            "other_delete_registro",
            "logout",
            "static",
            "login",
        }
        if endpoint not in allowed_other:
            return redirect(url_for("other_section"))

    if section == "web" and endpoint in {"new_section", "other_section"}:
        return redirect(url_for("index"))


def _can_upload() -> bool:
    key = _admin_key()
    if not key:
        return False
    return bool(session.get("can_upload"))


@app.get("/login")
def login():
    if _is_authenticated():
        if _portal_section() == "new":
            return redirect(url_for("new_section"))
        if _portal_section() == "other":
            return redirect(url_for("other_section"))
        return redirect(url_for("index"))
    return render_template("login.html")


@app.post("/login")
def login_post():
    entered_key = str(request.form.get("access_key") or "").strip()
    web_keys = [_access_key_web(), *_access_key_web_aliases()]
    key_new = _access_key_new()
    key_other = _access_key_other()
    if OTHER_SECTION_ENABLED and entered_key and hmac.compare_digest(entered_key, key_other):
        session["portal_section"] = "other"
        session.permanent = True
        return redirect(url_for("other_section"))

    if _match_any_key_ci(entered_key, web_keys):
        session["portal_section"] = "web"
        session.permanent = True
        return redirect(url_for("index"))
    if NEW_SECTION_ENABLED and entered_key and hmac.compare_digest(entered_key, key_new):
        session["portal_section"] = "new"
        session.permanent = True
        return redirect(url_for("new_section"))
    if not NEW_SECTION_ENABLED and entered_key and hmac.compare_digest(entered_key, key_new):
        flash("El acceso para la nueva seccion aun no esta habilitado.", "error")
        return redirect(url_for("login"))
    flash("Clave incorrecta.", "error")
    return redirect(url_for("login"))


@app.post("/logout")
def logout():
    session.pop("portal_section", None)
    session.pop("can_upload", None)
    flash("Sesion cerrada.", "success")
    return redirect(url_for("login"))


@app.get("/nueva-seccion")
def new_section():
    if not NEW_SECTION_ENABLED:
        return redirect(url_for("index"))
    return render_template("new_section.html", dashboard=_build_new_section_dashboard())


@app.get("/otra-landing")
def other_section():
    if not OTHER_SECTION_ENABLED:
        return redirect(url_for("index"))
    fecha = str(request.args.get("fecha") or "").strip()
    empleado = str(request.args.get("empleado") or "").strip()
    data = query_lavanderia_productividad(DB_PATH, fecha=fecha, empleado=empleado, limit_rows=400)
    catalogos = query_lavanderia_catalogos(DB_PATH)
    if not (catalogos.get("botas") or []):
        catalogos["botas"] = DEFAULT_LAVANDERIA_BOTAS[:]
    etapa_defaults = {
        str(item.get("etapa") or "").strip(): float(item.get("min_por_prenda") or 0)
        for item in (data.get("top_etapas") or [])
        if str(item.get("etapa") or "").strip()
    }
    return render_template(
        "other_section.html",
        data=data,
        filters={"fecha": fecha, "empleado": empleado},
        catalogos=catalogos,
        etapa_defaults=etapa_defaults,
        today_iso=date.today().isoformat(),
    )


@app.post("/otra-landing/import-excel")
def other_import_excel():
    if _portal_section() != "other":
        return redirect(url_for("login"))
    try:
        file = request.files.get("file")
        if not file or not file.filename:
            raise ValueError("Debes seleccionar un archivo Excel.")
        content = file.read()
        rows = parse_lavanderia_productividad_xlsx(content)
        botas = parse_lavanderia_botas_maestros_xlsx(content)
        etapas = parse_lavanderia_etapas_gestion_xlsx(content)
        stats = {"read": 0, "inserted": 0, "updated": 0}
        if rows:
            stats = import_lavanderia_rows(DB_PATH, rows, replace_all=True, source="excel")
        botas_stats = import_lavanderia_botas_maestro(DB_PATH, botas, replace_all=True)
        etapas_stats = import_lavanderia_etapas_maestro(DB_PATH, etapas, replace_all=True)
        if not rows and botas_stats.get("inserted", 0) == 0 and etapas_stats.get("inserted", 0) == 0:
            raise ValueError("No se detectaron datos validos en Gestion/Maestros.")
        flash(
            "Excel importado. "
            f"Registros: {stats.get('inserted', 0)} | "
            f"Botas maestro: {botas_stats.get('inserted', 0)} | "
            f"Etapas maestro: {etapas_stats.get('inserted', 0)}.",
            "success",
        )
    except Exception as exc:
        flash(f"No se pudo importar el Excel: {exc}", "error")
    return redirect(url_for("other_section"))


@app.post("/otra-landing/add")
def other_add_registro():
    if _portal_section() != "other":
        return redirect(url_for("login"))
    try:
        payload = {
            "articulo": str(request.form.get("articulo") or "").strip(),
            "corte": str(request.form.get("corte") or "").strip(),
            "bota": str(request.form.get("bota") or "").strip(),
            "etapa": str(request.form.get("etapa") or "").strip(),
            "empleado": str(request.form.get("empleado") or "").strip(),
            "cantidad": int(str(request.form.get("cantidad") or "0").strip() or "0"),
            "minutos": int(str(request.form.get("minutos") or "0").strip() or "0"),
            "fecha_inicio_iso": str(request.form.get("fecha_inicio_iso") or "").strip() or None,
            "hora_inicio": str(request.form.get("hora_inicio") or "").strip() or None,
            "fecha_fin_iso": str(request.form.get("fecha_fin_iso") or "").strip() or None,
            "hora_fin": str(request.form.get("hora_fin") or "").strip() or None,
            "source": "web",
        }
        if not payload["etapa"] or not payload["empleado"]:
            raise ValueError("Etapa y empleado son obligatorios.")
        add_lavanderia_registro(DB_PATH, payload)
        flash("Registro agregado.", "success")
    except Exception as exc:
        flash(f"No se pudo agregar el registro: {exc}", "error")
    return redirect(url_for("other_section"))


@app.post("/otra-landing/delete/<int:row_id>")
def other_delete_registro(row_id: int):
    if _portal_section() != "other":
        return redirect(url_for("login"))
    if delete_lavanderia_registro(DB_PATH, row_id):
        flash("Registro eliminado.", "success")
    else:
        flash("No se encontro el registro.", "error")
    return redirect(url_for("other_section"))


def _norm_text(value: str) -> str:
    raw = unicodedata.normalize("NFKD", str(value or ""))
    no_accents = "".join(ch for ch in raw if not unicodedata.combining(ch))
    cleaned = re.sub(r"[^a-zA-Z0-9 ]+", " ", no_accents).lower()
    return re.sub(r"\s+", " ", cleaned).strip()


def _has_keyword(text: str, options: list[str]) -> bool:
    words = _norm_text(text).split()
    normalized = _norm_text(text)
    for opt in options:
        opt_n = _norm_text(opt)
        if not opt_n:
            continue
        if opt_n in normalized:
            return True
        for w in words:
            if not w:
                continue
            if SequenceMatcher(None, w, opt_n).ratio() >= 0.82:
                return True
    return False


def _extract_rank(text: str) -> int:
    tn = _norm_text(text)
    m = re.search(r"\btop\s+(\d+)\b", tn)
    if m:
        return max(int(m.group(1)), 1)
    m2 = re.search(r"\b(\d+)(?:er|do|to|ro)?\b", tn)
    if m2 and int(m2.group(1)) <= 10:
        return max(int(m2.group(1)), 1)
    if _has_keyword(tn, ["primero", "primer"]):
        return 1
    if _has_keyword(tn, ["segundo", "segunda", "segun", "2do"]):
        return 2
    if _has_keyword(tn, ["tercero", "tercera", "3ro"]):
        return 3
    if _has_keyword(tn, ["cuarto", "cuarta", "4to"]):
        return 4
    if _has_keyword(tn, ["quinto", "quinta", "5to"]):
        return 5
    return 1


def _extract_query_code(question: str) -> str:
    digits = re.findall(r"\d+", question or "")
    if not digits:
        return ""
    # Preferimos codigos de al menos 4 digitos para familia/articulo.
    candidates = [d for d in digits if len(d) >= 4]
    return candidates[0] if candidates else digits[0]


def _extract_family_code(code: str) -> str:
    digits = "".join(ch for ch in str(code or "") if ch.isdigit())
    if not digits:
        return ""
    if len(digits) == 4:
        return digits
    candidates: list[str] = []
    if len(digits) >= 8:
        candidates.append(digits[2:6])
    if len(digits) >= 6:
        candidates.append(digits[:4])
        candidates.append(digits[-4:])
    if len(digits) >= 4:
        candidates.append(digits[:4])
    for c in candidates:
        if len(c) == 4:
            return c
    return digits[:4]


def _resolve_ex_details(code: str) -> dict | None:
    family = _extract_family_code(code)
    if not family:
        return None
    ex_summary = query_exs_balance_summary(DB_PATH, "")
    rows = ex_summary.get("rows") or []
    for item in rows:
        if str(item.get("actual") or "").strip() == family:
            ex_raw = str(item.get("ex") or "").strip()
            ex_family = _extract_family_code(ex_raw)
            return {
                "family_actual": family,
                "family_ex": ex_family,
                "ex_raw": ex_raw,
                "saldo_actual": int(item.get("saldo_actual") or 0),
                "saldo_ex": int(item.get("saldo_ex") or 0),
            }
    return None


def _build_assistant_context(question: str) -> str:
    rows, _, summary = query_rows(DB_PATH, {"q": "", "fecha": ""})
    pedidos_sections = query_pedidos_talla_sections(DB_PATH, "")
    exs_summary = query_exs_balance_summary(DB_PATH, "")
    assistant_rules = query_assistant_rules(DB_PATH, limit=60)
    ventas_rows = pedidos_sections.get("ventas", [])

    ventas_por_articulo: dict[str, int] = {}
    ventas_por_familia: dict[str, int] = {}
    ventas_por_talla: dict[int, int] = {}
    for r in ventas_rows:
        articulo = str(r.get("articulo") or "").strip()
        total = int(r.get("total") or 0)
        if not articulo:
            continue
        ventas_por_articulo[articulo] = ventas_por_articulo.get(articulo, 0) + total
        familia = articulo[2:6] if len(articulo) >= 6 else articulo
        ventas_por_familia[familia] = ventas_por_familia.get(familia, 0) + total
        for item in r.get("tallas_items") or []:
            talla = int(item.get("talla") or 0)
            cantidad = int(item.get("cantidad") or 0)
            if talla > 0:
                ventas_por_talla[talla] = ventas_por_talla.get(talla, 0) + cantidad

    top_articulos = [
        {"articulo": a, "total": t}
        for a, t in sorted(ventas_por_articulo.items(), key=lambda x: x[1], reverse=True)[:12]
    ]
    top_familias = [
        {"familia": f, "total": t}
        for f, t in sorted(ventas_por_familia.items(), key=lambda x: x[1], reverse=True)[:12]
    ]
    curva_tallas = [
        {"talla": talla, "total": total}
        for talla, total in sorted(ventas_por_talla.items(), key=lambda x: x[0])
    ]

    etapas: dict[str, int] = {}
    bodega_por_articulo: dict[str, int] = {}
    fechas_conteo: dict[str, int] = {}
    for r in rows:
        stage = str(r.get("proceso_actual") or "Sin movimiento")
        proceso = int(r.get("proceso") or 0)
        etapas[stage] = etapas.get(stage, 0) + proceso
        articulo = str(r.get("articulo") or "").strip()
        if articulo:
            bodega_por_articulo[articulo] = bodega_por_articulo.get(articulo, 0) + int(r.get("bodega") or 0)
        fecha_iso = str(r.get("fecha_iso") or "").strip()
        if fecha_iso:
            fechas_conteo[fecha_iso] = fechas_conteo.get(fecha_iso, 0) + 1

    top_etapas = [
        {"etapa": e, "total": t}
        for e, t in sorted(etapas.items(), key=lambda x: x[1], reverse=True)[:10]
    ]
    top_bodega_articulo = [
        {"articulo": a, "bodega": t}
        for a, t in sorted(bodega_por_articulo.items(), key=lambda x: x[1], reverse=True)[:10]
        if t > 0
    ]
    today_iso = date.today().isoformat()
    fechas_ordenadas = sorted(fechas_conteo.items(), key=lambda x: x[0], reverse=True)
    ultimas_fechas = [{"fecha_iso": f, "registros": c} for f, c in fechas_ordenadas[:10]]

    q_code = _extract_query_code(question or "")
    detalle_codigo = {}
    if q_code:
        code_rows, _, _ = query_rows(DB_PATH, {"q": q_code, "fecha": ""})
        code_pedidos = query_pedidos_talla_sections(DB_PATH, q_code).get("ventas", [])
        code_ex = _resolve_ex_details(q_code)
        saldos_detalle_limpio = []
        total_bodega_code = 0
        total_restante_code = 0
        total_proceso_code = 0
        for r in code_rows[:30]:
            bodega = int(r.get("bodega") or 0)
            restante = int(r.get("pendiente_en_trazabilidad") or 0)
            proceso = int(r.get("proceso") or 0)
            total_bodega_code += bodega
            total_restante_code += restante
            total_proceso_code += proceso
            saldos_detalle_limpio.append(
                {
                    "articulo": str(r.get("articulo") or ""),
                    "orden_corte": str(r.get("corte") or ""),
                    "proceso_total": proceso,
                    "bodega": bodega,
                    "restante": restante,
                    "proceso_actual": str(r.get("proceso_actual") or ""),
                    "restante_detalle": str(r.get("restante_detalle") or ""),
                }
            )
        detalle_codigo = {
            "codigo_consultado": q_code,
            "resumen_saldos": {
                "total_proceso": total_proceso_code,
                "total_bodega": total_bodega_code,
                "total_restante": total_restante_code,
            },
            "saldos": saldos_detalle_limpio,
            "ventas": code_pedidos[:20],
            "ex": code_ex or {},
        }

    context_data = {
        "reglas_interpretacion": {
            "bodega": "columna bodega (unidades actualmente en bodega)",
            "restante": "columna restante o pendiente_en_trazabilidad (NO es bodega)",
            "proceso_total": "columna proceso/total de la orden",
            "nota": "nunca confundir restante con bodega",
        },
        "reglas_negocio": [
            {
                "key": str(item.get("rule_key") or ""),
                "text": str(item.get("rule_text") or ""),
                "priority": int(item.get("priority") or 0),
            }
            for item in assistant_rules
        ],
        "fechas": {
            "campo_principal": "fecha_iso",
            "descripcion": "fecha del registro importado (no fecha de ingreso al sistema web)",
            "hoy_iso_servidor": today_iso,
            "registros_hoy": int(fechas_conteo.get(today_iso, 0)),
            "ultimas_fechas": ultimas_fechas,
        },
        "resumen_global": {
            "total_registros_saldos": len(rows),
            "ordenes_en_bodega": int(summary.get("ordenes_en_bodega", 0)),
            "cantidad_en_bodega": int(summary.get("cantidad_en_bodega", 0)),
            "pendiente_trazabilidad_bodega": int(summary.get("pendiente_en_trazabilidad_bodega", 0)),
            "total_ventas": sum(int(r.get("total") or 0) for r in ventas_rows),
        },
        "exs": {
            "vinculados": int(exs_summary.get("count", 0)),
            "saldo_actual_total": int(exs_summary.get("total_actual", 0)),
            "saldo_ex_total": int(exs_summary.get("total_ex", 0)),
            "muestras": (exs_summary.get("rows") or [])[:30],
        },
        "ventas": {
            "top_articulos": top_articulos,
            "top_familias": top_familias,
            "curva_tallas": curva_tallas,
        },
        "saldos": {
            "top_etapas": top_etapas,
            "top_bodega_articulo": top_bodega_articulo,
        },
        "detalle_codigo": detalle_codigo,
    }
    return json.dumps(context_data, ensure_ascii=False, default=str)


def _answer_precise_metrics(question: str) -> str | None:
    qn = _norm_text(question or "")
    if not qn:
        return None

    rows, _, summary = query_rows(DB_PATH, {"q": "", "fecha": ""})
    today_iso = date.today().isoformat()

    asks_orders = _has_keyword(qn, ["orden", "ordenes"])
    asks_cut_order = _has_keyword(qn, ["orden de corte", "ordenes de corte"])
    asks_bodega = _has_keyword(qn, ["bodega", "almacen"])
    asks_today = _has_keyword(qn, ["hoy", "dia de hoy", "hoy dia"])

    # "Orden de corte" se interpreta como identificador de orden, no como etapa corte_1.
    if asks_orders and asks_bodega and not asks_today:
        return (
            f"Actualmente hay {int(summary.get('ordenes_en_bodega', 0))} ordenes en bodega, "
            f"con {int(summary.get('cantidad_en_bodega', 0))} prendas en bodega y "
            f"{int(summary.get('pendiente_en_trazabilidad_bodega', 0))} restantes."
        )

    if asks_today and (asks_cut_order or asks_orders):
        rows_today = [r for r in rows if str(r.get("fecha_iso") or "") == today_iso]
        if asks_bodega:
            rows_today_bodega = [r for r in rows_today if int(r.get("bodega") or 0) > 0]
            prendas_bodega = sum(int(r.get("bodega") or 0) for r in rows_today_bodega)
            return (
                f"Hoy ({today_iso}) hay {len(rows_today_bodega)} ordenes en bodega, "
                f"con {prendas_bodega} prendas en bodega."
            )
        return f"Hoy ({today_iso}) hay {len(rows_today)} ordenes de corte registradas."

    return None


def _answer_assistant(question: str) -> str:
    q = (question or "").strip()
    if not q:
        return "Escribe una pregunta. Ejemplo: En que parte se encuentra 4210."

    ql = q.lower()
    qn = _norm_text(q)
    rows, _, summary = query_rows(DB_PATH, {"q": "", "fecha": ""})
    pedidos_sections = query_pedidos_talla_sections(DB_PATH, "")
    exs_summary = query_exs_balance_summary(DB_PATH, "")
    asks_location = _has_keyword(qn, ["donde", "parte", "encuentra", "ubicacion", "ubica"])

    if _has_keyword(
        qn,
        [
            "puedes leer txt",
            "puedes leer excel",
            "leer archivos",
            "archivos que cargue",
            "archivos cargados",
            "puedes ver mis archivos",
        ],
    ):
        return (
            "Si. Trabajo con la data que cargaste en ADECOM WEB (TXT/Excel) una vez importada a la base de datos. "
            "Puedo responder por articulo, familia, tallas, bodega, pedidos y cruces entre esas tablas."
        )

    if _has_keyword(qn, ["ayuda", "que puedes", "que sabes", "como funcionas"]):
        return (
            "Puedo responder sobre: ordenes en bodega, tabla completa, "
            "pedidos totales, familia con mas pedidos, top articulos y ubicacion por articulo/familia "
            "(ej: 4210 o 01420100)."
        )

    code = _extract_query_code(q)
    if code and asks_location:
        code_rows, _, _ = query_rows(DB_PATH, {"q": code, "fecha": ""})
        if not code_rows:
            return f"No encontre datos para {code}."

        bodega_rows = [r for r in code_rows if int(r.get("bodega") or 0) > 0]
        prendas_bodega = sum(int(r.get("bodega") or 0) for r in code_rows)
        prendas_proceso = sum(int(r.get("proceso") or 0) for r in code_rows)
        pendientes = sum(int(r.get("pendiente_en_trazabilidad") or 0) for r in code_rows)

        if bodega_rows:
            return (
                f"{code}: se encuentra en bodega en {len(bodega_rows)} orden(es), "
                f"con {prendas_bodega} prendas en bodega. "
                f"Total en proceso: {prendas_proceso}. Pendiente en trazabilidad: {pendientes}."
            )

        top_stage: dict[str, int] = {}
        for row in code_rows:
            stage = str(row.get("proceso_actual") or "Sin movimiento")
            top_stage[stage] = top_stage.get(stage, 0) + int(row.get("proceso") or 0)
        stage_name, stage_total = max(top_stage.items(), key=lambda x: x[1])
        return (
            f"{code}: no tiene prendas en bodega actualmente. "
            f"La mayor cantidad esta en {stage_name} con {stage_total} prendas. "
            f"Total en proceso: {prendas_proceso}."
        )

    if _has_keyword(qn, ["bodega", "almacen"]):
        return (
            f"Ordenes en bodega: {summary.get('ordenes_en_bodega', 0)}. "
            f"Cantidad en bodega: {summary.get('cantidad_en_bodega', 0)}. "
            f"Pendiente en trazabilidad: {summary.get('pendiente_en_trazabilidad_bodega', 0)}."
        )

    if _has_keyword(qn, ["muestra", "muestras"]):
        muestras_rows = [
            row for row in rows if str(row.get("corte", "")).lstrip("0").startswith("96")
        ]
        muestras_total = sum(int(row.get("proceso") or 0) for row in muestras_rows)
        muestras_bodega = sum(int(row.get("bodega") or 0) for row in muestras_rows)
        return (
            f"Total muestras: {len(muestras_rows)} orden(es). "
            f"Prendas en muestras: {muestras_total}. En bodega: {muestras_bodega}."
        )

    if _has_keyword(qn, ["exs", "ex"]):
        if code:
            ex_data = _resolve_ex_details(code)
            if ex_data:
                return (
                    f"EX para {code}: familia actual {ex_data['family_actual']}, "
                    f"EX {ex_data['ex_raw']} (familia {ex_data['family_ex']}). "
                    f"Saldo actual: {ex_data['saldo_actual']}. Saldo EX: {ex_data['saldo_ex']}."
                )
            return f"No encontre mapeo EX para {code}."
        return (
            f"EXS vinculados: {exs_summary.get('count', 0)}. "
            f"Total saldo actual: {exs_summary.get('total_actual', 0)}. "
            f"Total saldo ex: {exs_summary.get('total_ex', 0)}."
        )

    if _has_keyword(qn, ["venta", "ventas", "pedido", "pedidos", "vendido", "vendida", "vender", "mas vendido", "mas pedidos"]):
        ventas_rows = pedidos_sections.get("ventas", [])
        ventas_total = sum(int(r.get("total") or 0) for r in ventas_rows)
        ventas_por_familia: dict[str, int] = {}
        ventas_por_articulo: dict[str, int] = {}
        for r in ventas_rows:
            articulo = str(r.get("articulo") or "").strip()
            total = int(r.get("total") or 0)
            if not articulo:
                continue
            familia = articulo[2:6] if len(articulo) >= 6 else articulo
            ventas_por_familia[familia] = ventas_por_familia.get(familia, 0) + total
            ventas_por_articulo[articulo] = ventas_por_articulo.get(articulo, 0) + total

        rank = _extract_rank(qn)
        familias_sorted = sorted(ventas_por_familia.items(), key=lambda x: x[1], reverse=True)
        articulos_sorted = sorted(ventas_por_articulo.items(), key=lambda x: x[1], reverse=True)
        asks_familia = _has_keyword(qn, ["familia"])
        asks_articulo = _has_keyword(qn, ["articulo", "modelo", "referencia"])
        asks_rank = rank > 1 or _has_keyword(qn, ["top", "mas vendido", "mas vendida", "mas pedidos"])

        if asks_familia:
            if not familias_sorted:
                return "No hay datos de pedidos para calcular familias."
            if rank > len(familias_sorted):
                return f"No hay suficientes familias para obtener el puesto {rank}."
            fam, total = familias_sorted[rank - 1]
            return f"Familia #{rank} en pedidos: {fam}, con {total} unidades."

        if asks_articulo:
            if not articulos_sorted:
                return "No hay datos de pedidos para calcular articulos."
            if rank > len(articulos_sorted):
                return f"No hay suficientes articulos para obtener el puesto {rank}."
            art, total = articulos_sorted[rank - 1]
            return f"Articulo #{rank} en pedidos: {art}, con {total} unidades."

        if asks_rank:
            if not articulos_sorted or not familias_sorted:
                return "No hay datos de pedidos para calcular ranking."
            if rank > len(articulos_sorted) or rank > len(familias_sorted):
                return f"No hay suficientes datos para obtener el puesto {rank}."
            art, art_total = articulos_sorted[rank - 1]
            fam, fam_total = familias_sorted[rank - 1]
            return (
                f"Puesto #{rank} en pedidos: articulo {art} ({art_total}) y familia {fam} ({fam_total}). "
                f"Si quieres uno especifico, pregunta por 'articulo' o 'familia'."
            )

        return f"Total pedidos actual: {ventas_total} unidades."

    if _has_keyword(qn, ["tabla completa", "total ordenes", "cuantas ordenes", "cuantos registros"]):
        return f"Tabla completa: {len(rows)} orden(es) registradas."

    if code and _has_keyword(qn, ["toda la informacion", "todo sobre", "detalle completo", "toda la info"]):
        code_rows, _, _ = query_rows(DB_PATH, {"q": code, "fecha": ""})
        if not code_rows:
            ex_data = _resolve_ex_details(code)
            if ex_data:
                return (
                    f"{code}: no tiene registros en saldos actuales, pero su mapeo EX es "
                    f"{ex_data['ex_raw']} (familia {ex_data['family_ex']}). "
                    f"Saldo actual: {ex_data['saldo_actual']}; saldo EX: {ex_data['saldo_ex']}."
                )
            return f"No encontre datos para {code}."
        ordenes = len(code_rows)
        prendas_bodega = sum(int(r.get("bodega") or 0) for r in code_rows)
        prendas_proceso = sum(int(r.get("proceso") or 0) for r in code_rows)
        pendientes = sum(int(r.get("pendiente_en_trazabilidad") or 0) for r in code_rows)
        by_stage: dict[str, int] = {}
        for r in code_rows:
            stage = str(r.get("proceso_actual") or "Sin movimiento")
            by_stage[stage] = by_stage.get(stage, 0) + int(r.get("proceso") or 0)
        stage_txt = ", ".join(f"{k}:{v}" for k, v in sorted(by_stage.items(), key=lambda x: x[1], reverse=True)[:4])

        ventas_related = query_pedidos_talla_sections(DB_PATH, code).get("ventas", [])
        ventas_total = sum(int(r.get("total") or 0) for r in ventas_related)
        ex_data = _resolve_ex_details(code)
        ex_txt = (
            f"EX {ex_data['ex_raw']} (familia {ex_data['family_ex']}), saldo ex {ex_data['saldo_ex']}"
            if ex_data
            else "sin mapeo EX"
        )
        return (
            f"{code}: ordenes {ordenes}; en bodega {prendas_bodega}; total proceso {prendas_proceso}; "
            f"pendiente {pendientes}; etapas {stage_txt}; pedidos relacionados {ventas_total}; {ex_txt}."
        )

    if code:
        code_rows, _, _ = query_rows(DB_PATH, {"q": code, "fecha": ""})
        if not code_rows:
            return f"No encontre datos para {code}."
        return (
            f"Encontre {len(code_rows)} registro(s) para {code}. "
            "Si quieres ubicacion exacta pregunta: 'en que parte se encuentra ...'."
        )

    return (
        "Puedo ayudarte con bodega, pedidos o por codigo de articulo/familia. "
        "Ejemplos: 'Ordenes en bodega', 'Familia con mas pedidos', 'En que parte se encuentra 4210'."
    )


def _answer_with_gemini(question: str) -> str:
    api_key = os.environ.get("GEMINI_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("GEMINI_API_KEY no configurada.")

    precise_answer = _answer_precise_metrics(question)
    if precise_answer:
        return precise_answer

    env_model = os.environ.get("GEMINI_MODEL", "").strip()
    env_api_version = os.environ.get("GEMINI_API_VERSION", "").strip()
    api_versions: list[str] = []
    for v in [env_api_version, "v1", "v1beta"]:
        vv = str(v or "").strip()
        if vv and vv not in api_versions:
            api_versions.append(vv)

    discovered_models: list[str] = []
    list_errors: list[str] = []
    for api_version in api_versions:
        try:
            list_endpoint = f"https://generativelanguage.googleapis.com/{api_version}/models"
            list_req = url_request.Request(
                list_endpoint,
                headers={"x-goog-api-key": api_key},
                method="GET",
            )
            with url_request.urlopen(list_req, timeout=12) as resp:
                raw = resp.read().decode("utf-8")
            data = json.loads(raw or "{}")
            for item in data.get("models") or []:
                methods = item.get("supportedGenerationMethods") or []
                if "generateContent" not in methods:
                    continue
                name = str(item.get("name") or "").strip()
                if name.startswith("models/"):
                    name = name.split("/", 1)[1]
                if name and name not in discovered_models:
                    discovered_models.append(name)
            if discovered_models:
                break
        except Exception as exc:
            list_errors.append(f"{api_version}:{exc}")
            continue

    fallback_models = [
        "gemini-3-flash-preview",
        "gemini-2.5-flash",
        "gemini-2.0-flash",
        "gemini-2.0-flash-lite",
        "gemini-1.5-flash",
    ]
    model_candidates: list[str] = []
    for m in [env_model, *discovered_models, *fallback_models]:
        m = str(m or "").strip()
        if m and m not in model_candidates:
            model_candidates.append(m)

    context = _build_assistant_context(question)
    instruction = (
        "Responde en espanol con tono cercano y natural, como un asistente humano. "
        "Prioriza frases cortas y claras; evita sonar mecanico. "
        "No uses listas largas salvo que te pidan detalle. "
        "Usa SOLO el contexto entregado (proviene de todos los archivos cargados e importados en ADECOM WEB). "
        "Aplica siempre 'reglas_negocio' y 'reglas_interpretacion' antes de responder. "
        "Interpretacion obligatoria: BODEGA es solo columna bodega; RESTANTE (pendiente_en_trazabilidad) es distinto y no debe reportarse como bodega. "
        "Cuando pregunten por fecha o por 'hoy', usa el bloque fechas (campo fecha_iso). "
        "No digas que no hay fecha si existe fechas.ultimas_fechas o fechas.registros_hoy. "
        "Si preguntan por un codigo/familia, prioriza detalle_codigo.resumen_saldos para los totales. "
        "Si te preguntan si puedes leer archivos, aclara que si puedes analizarlos una vez cargados al sistema. "
        "Si falta dato, dilo explicitamente sin inventar."
    )
    payload = {
        "contents": [
            {
                "role": "user",
                "parts": [{"text": f"{instruction}\n{context}\nPregunta: {question}"}],
            }
        ]
    }

    # 1) Intento principal con SDK oficial de Gemini.
    sdk_errors: list[str] = []
    try:
        from google import genai  # type: ignore

        client = genai.Client(api_key=api_key)
        for model in model_candidates:
            try:
                response = client.models.generate_content(
                    model=model,
                    contents=f"{instruction}\n{context}\nPregunta: {question}",
                )
                text = str(getattr(response, "text", "") or "").strip()
                if text:
                    return text
                sdk_errors.append(f"{model}:sin_texto")
            except Exception as exc:
                sdk_errors.append(f"{model}:{exc}")
                continue
    except Exception as exc:
        sdk_errors.append(f"sdk_import_or_client:{exc}")

    # 2) Fallback REST si el SDK falla.
    last_error = None
    tried: list[str] = []
    for api_version in api_versions:
        for model in model_candidates:
            tried.append(f"{api_version}:{model}")
            endpoint = (
                f"https://generativelanguage.googleapis.com/{api_version}/models/{model}:generateContent"
            )
            req = url_request.Request(
                endpoint,
                data=json.dumps(payload).encode("utf-8"),
                headers={
                    "Content-Type": "application/json",
                    "x-goog-api-key": api_key,
                },
                method="POST",
            )
            try:
                with url_request.urlopen(req, timeout=18) as resp:
                    raw = resp.read().decode("utf-8")
                data = json.loads(raw or "{}")
                candidates = data.get("candidates") or []
                if not candidates:
                    last_error = RuntimeError(f"Gemini ({model}) no retorno candidatos.")
                    continue
                parts = (((candidates[0] or {}).get("content") or {}).get("parts") or [])
                text = " ".join(str(p.get("text") or "").strip() for p in parts).strip()
                if not text:
                    last_error = RuntimeError(f"Gemini ({model}) no retorno texto.")
                    continue
                return text
            except url_error.HTTPError as exc:
                # 404: modelo no disponible. 403/401: key/permisos.
                last_error = RuntimeError(f"Gemini ({model}) HTTP {exc.code}")
                continue
            except Exception as exc:
                last_error = exc
                continue

    raise RuntimeError(
        "Gemini no disponible. "
        f"SDK: {'; '.join(sdk_errors[:8]) if sdk_errors else 'sin_intentos_sdk'}. "
        f"Modelos intentados: {', '.join(tried)}. "
        f"Listado modelos: {'; '.join(list_errors) if list_errors else 'ok'}. "
        f"Detalle final: {last_error}"
    )


def _answer_assistant_router(question: str) -> dict:
    provider = os.environ.get("ADECOM_ASSISTANT_PROVIDER", "local").strip().lower()
    if provider in {"gemini", "google"}:
        return {
            "answer": _answer_with_gemini(question),
            "provider": "gemini",
            "fallback": False,
            "detail": "",
        }
    try:
        return {
            "answer": _answer_assistant(question),
            "provider": "local",
            "fallback": False,
            "detail": "",
        }
    except Exception as exc:
        app.logger.exception("Fallo en asistente local", exc_info=exc)
        return {
            "answer": "No fue posible responder en este momento. Intenta nuevamente.",
            "provider": "local",
            "fallback": True,
            "detail": str(exc),
        }


def _table_count(table_name: str) -> int:
    conn = get_conn(DB_PATH)
    try:
        row = conn.execute(f"SELECT COUNT(*) AS n FROM {table_name}").fetchone()
        return int(row["n"] if row else 0)
    except Exception as exc:
        app.logger.warning("No se pudo contar tabla %s: %s", table_name, exc)
        return 0
    finally:
        conn.close()


def _norm_file_key(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(name or "").lower())


def _find_autoload_file(folder: Path, token: str, *, exclude_token: str = "") -> Path | None:
    if not folder.exists() or not folder.is_dir():
        return None
    token_n = _norm_file_key(token)
    exclude_n = _norm_file_key(exclude_token) if exclude_token else ""
    candidates: list[Path] = []
    for p in folder.iterdir():
        if not p.is_file():
            continue
        if p.suffix.lower() not in {".txt", ".csv"}:
            continue
        key = _norm_file_key(p.name)
        if token_n not in key:
            continue
        if exclude_n and exclude_n in key:
            continue
        candidates.append(p)
    if not candidates:
        return None
    candidates.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    return candidates[0]


def _read_source_bytes(source: str) -> bytes:
    src = str(source or "").strip()
    if not src:
        raise ValueError("Fuente vacia.")
    if src.startswith(("http://", "https://")):
        req = url_request.Request(
            src,
            headers={
                "User-Agent": "ADECOM-WEB/1.0",
                "Accept": "text/plain,application/octet-stream,*/*",
            },
        )
        with url_request.urlopen(req, timeout=35) as resp:
            return resp.read()
    path = Path(src)
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"No existe archivo fuente: {src}")
    return path.read_bytes()


def _refresh_web_data() -> dict:
    saldos_sources = [
        src.strip()
        for src in re.split(r"[,\n;]+", str(AUTOLOAD_SALDOS_SOURCE or ""))
        if src.strip()
    ]
    saldos_rows: list[dict] = []
    for source in saldos_sources:
        saldos_rows.extend(parse_saldos_txt(_read_source_bytes(source)))
    pedidos_rows = parse_pedidos_talla_txt(_read_source_bytes(AUTOLOAD_PEDIDOS_SOURCE))
    etapas_rows = parse_corte_etapas_txt(_read_source_bytes(AUTOLOAD_ETAPAS_SOURCE))
    if not saldos_rows or not pedidos_rows or not etapas_rows:
        raise ValueError(
            f"Lectura vacia: saldos={len(saldos_rows)}, pedidos={len(pedidos_rows)}, etapas={len(etapas_rows)}"
        )
    stats_saldos = import_rows(DB_PATH, saldos_rows, replace_all=True, accumulate_on_conflict=True)
    stats_pedidos = import_pedidos_talla_rows(DB_PATH, pedidos_rows)
    stats_etapas = import_corte_etapas_rows(DB_PATH, etapas_rows)
    stats_comparativo = {"read": 0, "inserted": 0, "updated": 0}
    if AUTOLOAD_COMPARATIVO_SOURCE:
        comparativo_rows = parse_comparativo_clientes_txt(_read_source_bytes(AUTOLOAD_COMPARATIVO_SOURCE))
        if comparativo_rows:
            stats_comparativo = import_comparativo_clientes_rows(DB_PATH, comparativo_rows)
    return {
        "saldos": stats_saldos,
        "pedidos": stats_pedidos,
        "etapas": stats_etapas,
        "comparativo": stats_comparativo,
    }


def _autoload_watch_dirs() -> list[Path]:
    dirs: list[Path] = []
    seen: set[str] = set()
    for candidate in (AUTOLOAD_DIR, AUTOLOAD_DIR_FALLBACK):
        key = str(candidate).strip().lower()
        if not key or key in seen:
            continue
        seen.add(key)
        if candidate.exists() and candidate.is_dir():
            dirs.append(candidate)
    return dirs


def _find_latest_autoload_file(*patterns: str, exclude_terms: tuple[str, ...] = ()) -> Path | None:
    candidates: list[Path] = []
    excludes = tuple(term.lower() for term in exclude_terms)
    for base_dir in _autoload_watch_dirs():
        for pattern in patterns:
            try:
                candidates.extend(base_dir.glob(pattern))
            except OSError:
                continue
    files = [
        path
        for path in candidates
        if path.is_file() and not any(term in path.name.lower() for term in excludes)
    ]
    if not files:
        return None
    return max(files, key=lambda p: p.stat().st_mtime)


def _find_all_autoload_files(*patterns: str, exclude_terms: tuple[str, ...] = ()) -> list[Path]:
    candidates: list[Path] = []
    excludes = tuple(term.lower() for term in exclude_terms)
    seen: set[str] = set()
    for base_dir in _autoload_watch_dirs():
        for pattern in patterns:
            try:
                matches = sorted(base_dir.glob(pattern), key=lambda p: p.stat().st_mtime)
            except OSError:
                continue
            for path in matches:
                if not path.is_file():
                    continue
                if any(term in path.name.lower() for term in excludes):
                    continue
                key = str(path.resolve()).lower()
                if key in seen:
                    continue
                seen.add(key)
                candidates.append(path)
    return candidates


def _directory_autoload_sources() -> dict[str, Path]:
    sources: dict[str, Path] = {}
    latest_saldos = _find_latest_autoload_file("*SALDOS-SECCI*.TXT", "*SALDOS-SECCI*.txt")
    latest_pedidos = _find_latest_autoload_file(
        "*PEDIDOSXTALLA*.TXT",
        "*PEDIDOSXTALLA*.txt",
        exclude_terms=("todas",),
    )
    latest_etapas = _find_latest_autoload_file("*Grande-Adecom*.TXT", "*Grande-Adecom*.txt")
    latest_comparativo = _find_latest_autoload_file("*COMPARATIVO*.TXT", "*COMPARATIVO*.txt")
    latest_deudas = _find_latest_autoload_file("*Deudas_Vencidas*.CSV", "*Deudas_Vencidas*.csv")
    if latest_saldos:
        sources["saldos"] = latest_saldos
    if latest_pedidos:
        sources["pedidos"] = latest_pedidos
    if latest_etapas:
        sources["etapas"] = latest_etapas
    if latest_comparativo:
        sources["comparativo"] = latest_comparativo
    if latest_deudas:
        sources["deudas"] = latest_deudas
    return sources


def _refresh_directory_data() -> dict:
    sources = _directory_autoload_sources()
    saldos_files = _find_all_autoload_files("*SALDOS-SECCI*.TXT", "*SALDOS-SECCI*.txt")
    pedidos_path = sources.get("pedidos")
    etapas_path = sources.get("etapas")
    if not saldos_files or not pedidos_path or not etapas_path:
        raise FileNotFoundError("Faltan archivos requeridos en carpeta autoload.")

    saldos_rows: list[dict] = []
    for saldos_path in saldos_files:
        saldos_rows.extend(parse_saldos_txt(saldos_path.read_bytes()))
    pedidos_rows = parse_pedidos_talla_txt(pedidos_path.read_bytes())
    etapas_rows = parse_corte_etapas_txt(etapas_path.read_bytes())
    if not saldos_rows or not pedidos_rows or not etapas_rows:
        raise ValueError(
            f"Lectura vacia: saldos={len(saldos_rows)}, pedidos={len(pedidos_rows)}, etapas={len(etapas_rows)}"
        )

    stats_saldos = import_rows(DB_PATH, saldos_rows, replace_all=True, accumulate_on_conflict=True)
    stats_pedidos = import_pedidos_talla_rows(DB_PATH, pedidos_rows)
    stats_etapas = import_corte_etapas_rows(DB_PATH, etapas_rows)
    stats_comparativo = {"read": 0, "inserted": 0, "updated": 0}
    stats_deudas = {"read": 0, "inserted": 0, "updated": 0}

    comparativo_path = sources.get("comparativo")
    if comparativo_path:
        comparativo_rows = parse_comparativo_clientes_txt(comparativo_path.read_bytes())
        if comparativo_rows:
            stats_comparativo = import_comparativo_clientes_rows(DB_PATH, comparativo_rows)

    deudas_path = sources.get("deudas")
    if deudas_path:
        deuda_rows = parse_deudas_vencidas_csv(deudas_path.read_bytes())
        if deuda_rows:
            stats_deudas = import_deuda_clientes_rows(DB_PATH, deuda_rows)

    return {
        "saldos": stats_saldos,
        "pedidos": stats_pedidos,
        "etapas": stats_etapas,
        "comparativo": stats_comparativo,
        "deudas": stats_deudas,
    }


def _sources_configured() -> bool:
    return all(
        str(src or "").strip()
        for src in (AUTOLOAD_SALDOS_SOURCE, AUTOLOAD_PEDIDOS_SOURCE, AUTOLOAD_ETAPAS_SOURCE)
    )


def _directory_sources_configured() -> bool:
    sources = _directory_autoload_sources()
    return all(sources.get(key) for key in ("saldos", "pedidos", "etapas"))


def _current_refresh_mode() -> str:
    if _sources_configured():
        return "explicit"
    if _directory_sources_configured():
        return "directory"
    return ""


def _sources_signature() -> str:
    parts = []
    sources = [
        ("saldos", AUTOLOAD_SALDOS_SOURCE),
        ("pedidos", AUTOLOAD_PEDIDOS_SOURCE),
        ("etapas", AUTOLOAD_ETAPAS_SOURCE),
    ]
    if AUTOLOAD_COMPARATIVO_SOURCE:
        sources.append(("comparativo", AUTOLOAD_COMPARATIVO_SOURCE))
    for label, source in sources:
        payload = _read_source_bytes(source)
        digest = hashlib.sha256(payload).hexdigest()
        parts.append(f"{label}:{digest}")
    raw = "|".join(parts).encode("utf-8")
    return hashlib.sha256(raw).hexdigest()


def _directory_sources_signature() -> str:
    sources = _directory_autoload_sources()
    parts = []
    for path in _find_all_autoload_files("*SALDOS-SECCI*.TXT", "*SALDOS-SECCI*.txt"):
        payload = path.read_bytes()
        digest = hashlib.sha256(payload).hexdigest()
        parts.append(f"saldos:{path.name}:{digest}")
    for label in ("pedidos", "etapas", "comparativo", "deudas"):
        path = sources.get(label)
        if not path:
            continue
        payload = path.read_bytes()
        digest = hashlib.sha256(payload).hexdigest()
        parts.append(f"{label}:{path.name}:{digest}")
    raw = "|".join(parts).encode("utf-8")
    return hashlib.sha256(raw).hexdigest()


def _refresh_if_sources_changed(force: bool = False) -> bool:
    global _last_sources_signature, _last_refresh_mode
    mode = _current_refresh_mode()
    if not mode:
        return False
    with _refresh_lock:
        signature = _sources_signature() if mode == "explicit" else _directory_sources_signature()
        if not force and signature == _last_sources_signature and mode == _last_refresh_mode:
            return False
        stats = _refresh_web_data() if mode == "explicit" else _refresh_directory_data()
        _last_sources_signature = signature
        _last_refresh_mode = mode
    app.logger.info(
        "Auto refresh %s aplicado. SALDOS I%s/A%s | PEDIDOS I%s/A%s | ETAPAS I%s/A%s | COMPARATIVO I%s/A%s | DEUDAS I%s/A%s",
        mode,
        stats["saldos"].get("inserted", 0),
        stats["saldos"].get("updated", 0),
        stats["pedidos"].get("inserted", 0),
        stats["pedidos"].get("updated", 0),
        stats["etapas"].get("inserted", 0),
        stats["etapas"].get("updated", 0),
        stats["comparativo"].get("inserted", 0),
        stats["comparativo"].get("updated", 0),
        stats.get("deudas", {}).get("inserted", 0),
        stats.get("deudas", {}).get("updated", 0),
    )
    return True


def _auto_refresh_web_on_startup() -> None:
    if not AUTO_REFRESH_WEB_ON_START:
        app.logger.info("Auto refresh web al iniciar deshabilitado (ADECOM_AUTO_REFRESH_WEB_ON_START=0).")
        return
    if not _current_refresh_mode():
        app.logger.warning(
            "Auto refresh omitido: sin fuentes explicitas ni archivos validos en carpetas vigiladas (%s, %s).",
            AUTOLOAD_DIR,
            AUTOLOAD_DIR_FALLBACK,
        )
        return
    try:
        _refresh_if_sources_changed(force=True)
        app.logger.info("Auto refresh OK al iniciar.")
    except Exception as exc:
        app.logger.exception("Auto refresh fallo al iniciar: %s", exc, exc_info=exc)


def _parse_daily_time(value: str) -> tuple[int, int] | None:
    text = str(value or "").strip()
    if not text:
        return None
    m = re.fullmatch(r"([01]?\d|2[0-3]):([0-5]\d)", text)
    if not m:
        return None
    return int(m.group(1)), int(m.group(2))


def _auto_refresh_web_loop() -> None:
    daily_time = _parse_daily_time(AUTO_REFRESH_WEB_DAILY_TIME)
    if not daily_time and AUTO_REFRESH_WEB_DAILY_TIME:
        app.logger.warning(
            "Hora diaria invalida en ADECOM_AUTO_REFRESH_WEB_DAILY_TIME=%s (usar HH:MM).",
            AUTO_REFRESH_WEB_DAILY_TIME,
        )
    if AUTO_REFRESH_WEB_POLL_SECONDS <= 0 and not daily_time:
        app.logger.info(
            "Auto refresh web en loop deshabilitado (ADECOM_AUTO_REFRESH_WEB_POLL_SECONDS=0 sin hora diaria)."
        )
        return
    sleep_seconds = AUTO_REFRESH_WEB_POLL_SECONDS if AUTO_REFRESH_WEB_POLL_SECONDS > 0 else 30
    app.logger.info(
        "Auto refresh web loop activo cada %ss. Hora diaria=%s. Solo diario=%s.",
        sleep_seconds,
        AUTO_REFRESH_WEB_DAILY_TIME or "-",
        "1" if AUTO_REFRESH_WEB_ONLY_DAILY else "0",
    )
    global _last_daily_refresh_date
    while True:
        time.sleep(sleep_seconds)
        try:
            if daily_time:
                now = datetime.now()
                hh, mm = daily_time
                today_key = now.date().isoformat()
                if now.hour == hh and now.minute == mm and _last_daily_refresh_date != today_key:
                    _refresh_if_sources_changed(force=True)
                    _last_daily_refresh_date = today_key
                    app.logger.info("Actualizacion diaria ejecutada (%s).", AUTO_REFRESH_WEB_DAILY_TIME)
            if AUTO_REFRESH_WEB_POLL_SECONDS > 0 and not AUTO_REFRESH_WEB_ONLY_DAILY:
                changed = _refresh_if_sources_changed(force=False)
                if changed:
                    app.logger.info("Cambio detectado en fuentes. Data web actualizada automaticamente.")
        except Exception as exc:
            app.logger.warning("Auto refresh web loop: %s", exc)


def _start_auto_refresh_web_loop() -> None:
    global _refresh_thread_started
    if _refresh_thread_started:
        return
    if not AUTO_REFRESH_WEB_BACKGROUND:
        app.logger.info("Auto refresh web en segundo plano deshabilitado (ADECOM_AUTO_REFRESH_WEB_BACKGROUND=0).")
        return
    thread = threading.Thread(target=_auto_refresh_web_loop, name="adecom-auto-refresh", daemon=True)
    thread.start()
    _refresh_thread_started = True


def _to_float(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    text = text.replace("%", "").replace(" ", "")
    text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _canonical_area(area: object) -> str:
    probe = _norm_text(str(area or ""))
    if "corte" in probe:
        return "CORTE"
    if "taller externo" in probe or "t externo" in probe or "ext" in probe:
        return "TALLER EXTERNO"
    if "taller" in probe:
        return "TALLER"
    if "limpi" in probe:
        return "LIMPIADO"
    if "lavander" in probe:
        return "LAVANDERIA"
    if "termina" in probe:
        return "TERMINACION"
    if "bodega" in probe:
        return "BODEGA"
    return str(area or "").strip().upper() or "SIN AREA"


def _status_from_ratio(ratio: float) -> str:
    if ratio >= 1:
        return "green"
    if ratio >= 0.85:
        return "yellow"
    return "red"


def _production_goal_status(ratio: float) -> str:
    if ratio >= 1:
        return "green"
    if ratio >= 0.5:
        return "yellow"
    return "red"


def _load_programas_mhc_snapshot() -> dict[str, object] | None:
    def _load_programas_snapshot_json() -> dict[str, object] | None:
        if not PROGRAMAS_MHC_SNAPSHOT_PATH.exists():
            return None
        try:
            raw = json.loads(PROGRAMAS_MHC_SNAPSHOT_PATH.read_text(encoding="utf-8"))
        except Exception:
            return None
        if not isinstance(raw, dict):
            return None
        weeks = raw.get("weeks") or []
        sections = raw.get("sections") or {}
        if not isinstance(weeks, list) or not isinstance(sections, dict):
            return None
        return raw

    if not PROGRAMAS_MHC_PATH.exists():
        return _load_programas_snapshot_json()
    try:
        from openpyxl import load_workbook
    except Exception:
        return None
    try:
        wb = load_workbook(PROGRAMAS_MHC_PATH, data_only=True, read_only=True)
    except Exception:
        return _load_programas_snapshot_json()

    month_aliases = {
        1: "enero",
        2: "febrero",
        3: "marzo",
        4: "abril",
        5: "mayo",
        6: "junio",
        7: "julio",
        8: "agosto",
        9: "septiembre",
        10: "octubre",
        11: "noviembre",
        12: "diciembre",
    }
    current_label = f"{month_aliases.get(date.today().month, '')} {date.today().year}".strip()
    sheet_name = None
    for name in wb.sheetnames:
        if _norm_text(name) == _norm_text(current_label):
            sheet_name = name
            break
    if not sheet_name:
        candidates = []
        for name in wb.sheetnames:
            norm = _norm_text(name)
            if any(m in norm for m in month_aliases.values()):
                y = re.search(r"(20\d{2})", norm)
                year = int(y.group(1)) if y else 0
                m_idx = next((k for k, v in month_aliases.items() if v in norm), 0)
                candidates.append((year, m_idx, name))
        if candidates:
            candidates.sort()
            sheet_name = candidates[-1][2]
    if not sheet_name:
        return _load_programas_snapshot_json()

    ws = wb[sheet_name]
    row2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), ())
    days_month = int(_to_float(row2[13] if len(row2) > 13 else 0) or 0)
    days_elapsed = int(_to_float(row2[16] if len(row2) > 16 else 0) or 0)
    days_remaining = max(days_month - days_elapsed, 0)

    section_key_map = {
        "corte": "corte",
        "urrutia": "urrutia",
        "tierra del fuego": "tierra_fuego",
        "lavanderia": "lavanderia",
        "terminacion": "terminacion",
    }
    sections: dict[str, dict[str, float]] = {}
    for r in range(5, 10):
        row = next(ws.iter_rows(min_row=r, max_row=r, values_only=True), ())
        raw_name = str(row[15] or "").strip() if len(row) > 15 else ""
        key = section_key_map.get(_norm_text(raw_name), "")
        if not key:
            continue
        sections[key] = {
            "name": raw_name,
            "meta_day": float(_to_float(row[13] if len(row) > 13 else 0)),
            "meta_month": float(_to_float(row[14] if len(row) > 14 else 0)),
            "projected": float(_to_float(row[16] if len(row) > 16 else 0)),
            "actual": float(_to_float(row[17] if len(row) > 17 else 0)),
            "daily_avg": float(_to_float(row[20] if len(row) > 20 else 0)),
        }

    week_starts = [4, 22, 40, 58, 76]
    week_area_rows = [
        ("corte", 0, "Corte"),
        ("urrutia", 3, "Urrutia"),
        ("sur", 6, "Sur"),
        ("lavanderia", 9, "Lavanderia"),
        ("terminacion", 12, "Terminacion"),
    ]
    weeks: list[dict[str, object]] = []
    for idx, start in enumerate(week_starts, start=1):
        hab_row = next(ws.iter_rows(min_row=max(start - 3, 1), max_row=max(start - 3, 1), values_only=True), ())
        day_row = next(ws.iter_rows(min_row=max(start - 2, 1), max_row=max(start - 2, 1), values_only=True), ())
        day_candidates = [
            (i, int(_to_float(v) or 0))
            for i, v in enumerate(day_row)
            if int(_to_float(v) or 0) > 0 and int(_to_float(v) or 0) <= 31
        ]
        best_run: list[tuple[int, int]] = []
        for i in range(len(day_candidates)):
            run = [day_candidates[i]]
            for j in range(i + 1, len(day_candidates)):
                prev_col, prev_val = run[-1]
                col, val = day_candidates[j]
                if col > prev_col and val == prev_val + 1:
                    run.append((col, val))
                elif val <= prev_val:
                    break
            if len(run) > len(best_run):
                best_run = run
        selected = best_run if len(best_run) >= 2 else day_candidates
        day_col_indices = [i for i, _ in selected]
        habiles = []
        fechas = []
        for i_col in day_col_indices:
            h_raw = hab_row[i_col] if i_col < len(hab_row) else None
            d_raw = day_row[i_col] if i_col < len(day_row) else None
            h_val = int(_to_float(h_raw) or 0)
            d_val = int(_to_float(d_raw) or 0)
            habiles.append(h_val if h_val > 0 else None)
            fechas.append(d_val if d_val > 0 else None)
        week_rows = []
        for area_key, offset, label in week_area_rows:
            row = next(ws.iter_rows(min_row=start + offset, max_row=start + offset, values_only=True), ())
            values = []
            for i_col in day_col_indices:
                raw = row[i_col] if i_col < len(row) else None
                num = int(_to_float(raw) or 0)
                values.append(num if num > 0 else None)
            total_cell = int(_to_float(row[10] if len(row) > 10 else 0) or 0)
            total = total_cell if total_cell > 0 else sum(int(v or 0) for v in values)
            week_rows.append({"key": area_key, "name": label, "values": values, "total": total})
        weeks.append(
            {
                "label": f"Semana {idx}",
                "habiles": habiles,
                "fechas": fechas,
                "rows": week_rows,
            }
        )

    snapshot = {
        "sheet_name": sheet_name,
        "days_month": days_month,
        "days_elapsed": days_elapsed,
        "days_remaining": days_remaining,
        "sections": sections,
        "weeks": weeks,
    }
    try:
        PROGRAMAS_MHC_SNAPSHOT_PATH.write_text(
            json.dumps(snapshot, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception:
        pass
    return snapshot


def _build_production_goals_summary() -> dict[str, object]:
    snapshot = _load_programas_mhc_snapshot()
    if snapshot and snapshot.get("sections"):
        sections_seed = []
        ordered = ["corte", "urrutia", "tierra_fuego", "lavanderia", "terminacion"]
        for key in ordered:
            src = (snapshot.get("sections") or {}).get(key) or {}
            if not src:
                continue
            sections_seed.append(
                {
                    "name": str(src.get("name") or key.title()),
                    "goal": int(src.get("meta_month") or 0),
                    "actual": int(src.get("actual") or 0),
                    "projected": float(src.get("projected") or 0),
                    "daily_avg": float(src.get("daily_avg") or 0),
                    "comment": "Actualizado desde planilla PROGRAMAS MHC.",
                }
            )
        if sections_seed:
            sections: list[dict[str, object]] = []
            total_goal = 0
            total_actual = 0
            total_projected = 0.0
            green_count = 0
            delayed_count = 0
            for item in sections_seed:
                goal = int(item["goal"])
                actual = int(item["actual"])
                projected = float(item["projected"])
                ratio = (actual / goal) if goal > 0 else 0.0
                projected_ratio = (actual / projected) if projected > 0 else 0.0
                status = _production_goal_status(ratio)
                if status == "green":
                    green_count += 1
                if status == "red":
                    delayed_count += 1
                total_goal += goal
                total_actual += actual
                total_projected += projected
                sections.append(
                    {
                        "name": item["name"],
                        "goal": goal,
                        "actual": actual,
                        "projected": int(round(projected)),
                        "ratio_pct": round(ratio * 100, 1),
                        "projected_ratio_pct": round(projected_ratio * 100, 1),
                        "status": status,
                        "status_label": "Cumplida" if status == "green" else ("En marcha" if status == "yellow" else "Baja"),
                        "daily_avg": float(item["daily_avg"]),
                        "comment": item["comment"],
                    }
                )
            total_ratio = (total_actual / total_goal) if total_goal > 0 else 0.0
            projected_total_ratio = (total_actual / total_projected) if total_projected > 0 else 0.0
            overall_status = _production_goal_status(total_ratio)
            return {
                "month_label": str(snapshot.get("sheet_name") or "Metas produccion"),
                "work_days": int(snapshot.get("days_month") or 0),
                "projected_day": int(snapshot.get("days_elapsed") or 0),
                "total_goal": total_goal,
                "total_actual": total_actual,
                "total_ratio_pct": round(total_ratio * 100, 1),
                "projected_total": int(round(total_projected)),
                "projected_total_ratio_pct": round(projected_total_ratio * 100, 1),
                "status": overall_status,
                "status_label": "Cumplido" if overall_status == "green" else ("A media meta" if overall_status == "yellow" else "Bajo media meta"),
                "green_count": green_count,
                "delayed_count": delayed_count,
                "sections": sections,
                "comments": [
                    "Datos cargados automaticamente desde 1_PROGRAMAS DE PRODUCCION MHC .xlsx",
                ],
            }

    sections_seed = [
        {
            "name": "Corte",
            "goal": 10000,
            "actual": 2918,
            "projected": 1904.76,
            "daily_avg": 729.5,
            "comment": "Supera con holgura el ritmo proyectado del corte revisado.",
        },
        {
            "name": "Urrutia",
            "goal": 6000,
            "actual": 936,
            "projected": 1142.86,
            "daily_avg": 234,
            "comment": "Subio respecto a la carga anterior, pero sigue bajo el ritmo esperado.",
        },
        {
            "name": "Tierra del Fuego",
            "goal": 4000,
            "actual": 0,
            "projected": 761.9,
            "daily_avg": 0,
            "comment": "No hay avance registrado en la hoja revisada.",
        },
        {
            "name": "Lavandería",
            "goal": 10000,
            "actual": 2453,
            "projected": 1904.76,
            "daily_avg": 613.25,
            "comment": "Va por encima de lo proyectado y acelero su acumulado.",
        },
        {
            "name": "Terminación",
            "goal": 10000,
            "actual": 1803,
            "projected": 1904.76,
            "daily_avg": 450.75,
            "comment": "Muy cerca del proyectado; le falta un pequeño empuje para quedar en verde.",
        },
    ]

    sections: list[dict[str, object]] = []
    total_goal = 0
    total_actual = 0
    total_projected = 0.0
    green_count = 0
    delayed_count = 0
    for item in sections_seed:
        goal = int(item["goal"])
        actual = int(item["actual"])
        projected = float(item["projected"])
        ratio = (actual / goal) if goal > 0 else 0.0
        projected_ratio = (actual / projected) if projected > 0 else 0.0
        status = _production_goal_status(ratio)
        if status == "green":
            green_count += 1
        if status == "red":
            delayed_count += 1
        total_goal += goal
        total_actual += actual
        total_projected += projected
        sections.append(
            {
                "name": item["name"],
                "goal": goal,
                "actual": actual,
                "projected": int(round(projected)),
                "ratio_pct": round(ratio * 100, 1),
                "projected_ratio_pct": round(projected_ratio * 100, 1),
                "status": status,
                "status_label": "Cumplida" if status == "green" else ("En marcha" if status == "yellow" else "Baja"),
                "daily_avg": float(item["daily_avg"]),
                "comment": item["comment"],
            }
        )

    total_ratio = (total_actual / total_goal) if total_goal > 0 else 0.0
    projected_total_ratio = (total_actual / total_projected) if total_projected > 0 else 0.0
    overall_status = _production_goal_status(total_ratio)
    return {
        "month_label": "Abril 2026",
        "work_days": 21,
        "projected_day": 4,
        "total_goal": total_goal,
        "total_actual": total_actual,
        "total_ratio_pct": round(total_ratio * 100, 1),
        "projected_total": int(round(total_projected)),
        "projected_total_ratio_pct": round(projected_total_ratio * 100, 1),
        "status": overall_status,
        "status_label": "Cumplido" if overall_status == "green" else ("A media meta" if overall_status == "yellow" else "Bajo media meta"),
        "green_count": green_count,
        "delayed_count": delayed_count,
        "sections": sections,
        "comments": [
            "La carga visible se concentra en los primeros días del mes.",
            "Corte y Lavandería lideran el avance del periodo revisado.",
            "Urrutia y Tierra del Fuego necesitan seguimiento más cercano.",
        ],
    }


def _build_new_section_dashboard() -> dict[str, object]:
    snapshot = _load_programas_mhc_snapshot()
    if snapshot and snapshot.get("sections") and snapshot.get("weeks"):
        sheet_name = str(snapshot.get("sheet_name") or "Metas produccion")
        m = re.search(r"(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s+(20\d{2})", _norm_text(sheet_name))
        month_num = 4
        year_num = date.today().year
        if m:
            month_word = m.group(1)
            year_num = int(m.group(2))
            month_to_num = {
                "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
                "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
            }
            month_num = month_to_num.get(month_word, month_num)
        weekday_names = {0: "Lun", 1: "Mar", 2: "Mie", 3: "Jue", 4: "Vie", 5: "Sab", 6: "Dom"}
        def _day_label(day: int | None) -> str | None:
            if not day:
                return None
            try:
                dt = date(year_num, month_num, int(day))
            except Exception:
                return str(day)
            return f"{weekday_names.get(dt.weekday(), '')} {day}"

        sections_seed = []
        section_order = [("corte", "Corte"), ("urrutia", "Urrutia"), ("tierra_fuego", "Tierra del Fuego"), ("lavanderia", "Lavanderia"), ("terminacion", "Terminacion")]
        for key, label in section_order:
            src = (snapshot.get("sections") or {}).get(key) or {}
            if not src:
                continue
            actual = int(src.get("actual") or 0)
            goal = int(src.get("meta_month") or 0)
            proj = int(round(float(src.get("projected") or 0)))
            sections_seed.append(
                {
                    "name": str(src.get("name") or label),
                    "meta_day": int(round(float(src.get("meta_day") or 0))),
                    "meta_month": goal,
                    "actual": actual,
                    "projection": proj,
                    "remaining": max(goal - actual, 0),
                    "focus": "Actualizado desde planilla PROGRAMAS MHC.",
                }
            )

        weeks = snapshot.get("weeks") or []
        weekly_rows = []
        daily_weeks = []
        for i, week in enumerate(weeks, start=1):
            week_map = {str(r.get("key") or ""): int(r.get("total") or 0) for r in (week.get("rows") or [])}
            weekly_rows.append(
                {
                    "label": f"Sem {i}",
                    "corte": week_map.get("corte", 0),
                    "urrutia": week_map.get("urrutia", 0),
                    "sur": week_map.get("sur", 0),
                    "lavanderia": week_map.get("lavanderia", 0),
                    "terminacion": week_map.get("terminacion", 0),
                }
            )
            fechas = week.get("fechas") or []
            daily_weeks.append(
                {
                    "label": str(week.get("label") or f"Semana {i}"),
                    "habiles": week.get("habiles") or [],
                    "fechas": fechas,
                    "fecha_labels": [_day_label(int(d)) if d else None for d in fechas],
                    "rows": [{"name": r.get("name"), "values": r.get("values") or [], "total": int(r.get("total") or 0)} for r in (week.get("rows") or [])],
                }
            )

        days_month = int(snapshot.get("days_month") or 0)
        days_elapsed = int(snapshot.get("days_elapsed") or 0)
        days_remaining = int(snapshot.get("days_remaining") or max(days_month - days_elapsed, 0))
        expected_pct = round((days_elapsed / days_month) * 100, 1) if days_month else 0.0

        sections: list[dict[str, object]] = []
        for item in sections_seed:
            ratio = (int(item["actual"]) / int(item["meta_month"])) if int(item["meta_month"]) > 0 else 0.0
            avg_day = round(int(item["actual"]) / max(days_elapsed, 1), 2) if days_elapsed > 0 else 0.0
            avg_ratio = (avg_day / int(item["meta_day"])) if int(item["meta_day"]) > 0 else 0.0
            avg_status = _production_goal_status(avg_ratio)
            status = _production_goal_status(ratio)
            sections.append(
                {
                    **item,
                    "ratio_pct": round(ratio * 100, 1),
                    "status": status,
                    "status_label": "En meta" if status == "green" else ("A media meta" if status == "yellow" else "Bajo objetivo"),
                    "avg_day": avg_day,
                    "avg_ratio_pct": round(avg_ratio * 100, 1),
                    "avg_status": avg_status,
                    "avg_status_label": "En meta diaria" if avg_status == "green" else ("Media meta diaria" if avg_status == "yellow" else "Bajo meta diaria"),
                }
            )

        return {
            "month_title": sheet_name,
            "sheet_reference": "Formato base: 1_PROGRAMAS DE PRODUCCION MHC",
            "days_month": days_month,
            "days_remaining": days_remaining,
            "days_elapsed": days_elapsed,
            "expected_pct": expected_pct,
            "areas": sections,
            "weekly_rows": weekly_rows,
            "daily_weeks": daily_weeks,
            "comments": [
                "Datos diarios y acumulados cargados desde planilla PROGRAMAS MHC.",
            ],
        }

    weekday_names = {
        0: "Lun",
        1: "Mar",
        2: "Mie",
        3: "Jue",
        4: "Vie",
        5: "Sab",
        6: "Dom",
    }

    def _day_label(day: int | None) -> str | None:
        if not day:
            return None
        dt = date(2026, 4, int(day))
        return f"{weekday_names.get(dt.weekday(), '')} {day}"

    days_month = 21
    days_remaining = 17
    days_elapsed = 4
    expected_pct = round((days_elapsed / days_month) * 100, 1) if days_month else 0.0
    weekly_rows = [
        {"label": "Sem 1", "corte": 850, "urrutia": 338, "sur": 0, "lavanderia": 1108, "terminacion": 524},
        {"label": "Sem 2", "corte": 2068, "urrutia": 1000, "sur": 0, "lavanderia": 2141, "terminacion": 1693},
        {"label": "Sem 3", "corte": 0, "urrutia": 0, "sur": 0, "lavanderia": 0, "terminacion": 0},
        {"label": "Sem 4", "corte": 0, "urrutia": 0, "sur": 0, "lavanderia": 0, "terminacion": 0},
        {"label": "Sem 5", "corte": 0, "urrutia": 0, "sur": 0, "lavanderia": 0, "terminacion": 0},
    ]
    daily_weeks = [
        {
            "label": "Semana 1",
            "habiles": [1, 2, None, None, None],
            "fechas": [1, 2, 3, 4, 5],
            "fecha_labels": [_day_label(1), _day_label(2), _day_label(3), _day_label(4), _day_label(5)],
            "rows": [
                {"name": "Corte", "values": [652, 198, None, None, None], "total": 850},
                {"name": "Urrutia", "values": [176, 162, None, None, None], "total": 338},
                {"name": "Sur", "values": [None, None, None, None, None], "total": 0},
                {"name": "Lavanderia", "values": [456, 652, None, None, None], "total": 1108},
                {"name": "Terminacion", "values": [152, 372, None, None, None], "total": 524},
            ],
        },
        {
            "label": "Semana 2",
            "habiles": [3, 4, 5, 6, 7, None, None],
            "fechas": [6, 7, 8, 9, 10, 11, 12],
            "fecha_labels": [_day_label(6), _day_label(7), _day_label(8), _day_label(9), _day_label(10), _day_label(11), _day_label(12)],
            "rows": [
                {"name": "Corte", "values": [652, 652, 164, 600, None, None, None], "total": 2068},
                {"name": "Urrutia", "values": [150, 198, 250, 200, 202, None, None], "total": 1000},
                {"name": "Sur", "values": [None, None, None, None, None, None, None], "total": 0},
                {"name": "Lavanderia", "values": [341, 535, 469, 274, 522, None, None], "total": 2141},
                {"name": "Terminacion", "values": [468, 392, 419, 414, None, None, None], "total": 1693},
            ],
        },
        {
            "label": "Semana 3",
            "habiles": [8, 9, 10, 11, 12, None, None],
            "fechas": [13, 14, 15, 16, 17, 18, 19],
            "fecha_labels": [_day_label(13), _day_label(14), _day_label(15), _day_label(16), _day_label(17), _day_label(18), _day_label(19)],
            "rows": [
                {"name": "Corte", "values": [None, None, None, None, None, None, None], "total": 0},
                {"name": "Urrutia", "values": [None, None, None, None, None, None, None], "total": 0},
                {"name": "Sur", "values": [None, None, None, None, None, None, None], "total": 0},
                {"name": "Lavanderia", "values": [None, None, None, None, None, None, None], "total": 0},
                {"name": "Terminacion", "values": [None, None, None, None, None, None, None], "total": 0},
            ],
        },
        {
            "label": "Semana 4",
            "habiles": [13, 14, 15, 16, 17, None, None],
            "fechas": [20, 21, 22, 23, 24, 25, 26],
            "fecha_labels": [_day_label(20), _day_label(21), _day_label(22), _day_label(23), _day_label(24), _day_label(25), _day_label(26)],
            "rows": [
                {"name": "Corte", "values": [None, None, None, None, None, None, None], "total": 0},
                {"name": "Urrutia", "values": [None, None, None, None, None, None, None], "total": 0},
                {"name": "Sur", "values": [None, None, None, None, None, None, None], "total": 0},
                {"name": "Lavanderia", "values": [None, None, None, None, None, None, None], "total": 0},
                {"name": "Terminacion", "values": [None, None, None, None, None, None, None], "total": 0},
            ],
        },
        {
            "label": "Semana 5",
            "habiles": [18, 19, 20, 21, None],
            "fechas": [27, 28, 29, 30, None],
            "fecha_labels": [_day_label(27), _day_label(28), _day_label(29), _day_label(30), None],
            "rows": [
                {"name": "Corte", "values": [None, None, None, None, None], "total": 0},
                {"name": "Urrutia", "values": [None, None, None, None, None], "total": 0},
                {"name": "Sur", "values": [None, None, None, None, None], "total": 0},
                {"name": "Lavanderia", "values": [None, None, None, None, None], "total": 0},
                {"name": "Terminacion", "values": [None, None, None, None, None], "total": 0},
            ],
        },
    ]
    sections_seed = [
        {"name": "Corte", "meta_day": 476, "meta_month": 10000, "actual": 2918, "projection": 1905, "remaining": 7082, "focus": "Arranca claramente sobre el ritmo esperado y lidera el avance actual."},
        {"name": "Urrutia", "meta_day": 286, "meta_month": 6000, "actual": 686, "projection": 1143, "remaining": 5314, "focus": "Va bajo el ritmo proyectado y necesita recuperación temprana."},
        {"name": "Tierra del Fuego", "meta_day": 190, "meta_month": 4000, "actual": 0, "projection": 762, "remaining": 4000, "focus": "No muestra carga en la foto revisada del mes."},
        {"name": "Lavanderia", "meta_day": 476, "meta_month": 10000, "actual": 2453, "projection": 1905, "remaining": 7547, "focus": "Aumenta el acumulado y queda por encima del esperado del periodo."},
        {"name": "Terminacion", "meta_day": 476, "meta_month": 10000, "actual": 1803, "projection": 1905, "remaining": 8197, "focus": "Está cerca del objetivo proyectado y con margen de ajuste corto."},
    ]

    sections: list[dict[str, object]] = []
    for item in sections_seed:
        if item["name"] == "Urrutia":
            item["actual"] = 936
            item["remaining"] = 5064
            item["focus"] = "Subio respecto a la carga anterior, pero sigue bajo el proyectado."
        ratio = (int(item["actual"]) / int(item["meta_month"])) if int(item["meta_month"]) > 0 else 0.0
        avg_day = round(int(item["actual"]) / days_elapsed, 2) if days_elapsed > 0 else 0.0
        avg_ratio = (avg_day / int(item["meta_day"])) if int(item["meta_day"]) > 0 else 0.0
        avg_status = _production_goal_status(avg_ratio)
        status = _production_goal_status(ratio)
        sections.append(
            {
                **item,
                "ratio_pct": round(ratio * 100, 1),
                "status": status,
                "status_label": "En meta" if status == "green" else ("A media meta" if status == "yellow" else "Bajo objetivo"),
                "avg_day": avg_day,
                "avg_ratio_pct": round(avg_ratio * 100, 1),
                "avg_status": avg_status,
                "avg_status_label": "En meta diaria" if avg_status == "green" else ("Media meta diaria" if avg_status == "yellow" else "Bajo meta diaria"),
            }
        )

    return {
        "month_title": "Abril 2026",
        "sheet_reference": "Formato base: 1_PROGRAMAS DE PRODUCCION MHC",
        "days_month": days_month,
        "days_remaining": days_remaining,
        "days_elapsed": days_elapsed,
        "expected_pct": expected_pct,
        "areas": sections,
        "weekly_rows": weekly_rows,
        "daily_weeks": daily_weeks,
        "comments": [
            "La vista diaria replica los bloques de la hoja de abril: días hábiles, fecha, carga por sección y total del bloque.",
            "El panel lateral resume meta mensual, acumulado proyectado, real y porcentaje acumulado por sección.",
        ],
    }


def _build_excel_preview_dashboard() -> dict[str, object]:
    snapshot = _load_programas_mhc_snapshot()
    if snapshot and snapshot.get("sections") and snapshot.get("weeks"):
        sheet_name = str(snapshot.get("sheet_name") or "Metas produccion")
        m = re.search(r"(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s+(20\d{2})", _norm_text(sheet_name))
        month_num = 4
        year_num = date.today().year
        if m:
            month_word = m.group(1)
            year_num = int(m.group(2))
            month_to_num = {
                "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
                "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
            }
            month_num = month_to_num.get(month_word, month_num)

        weekday_names = {0: "L", 1: "M", 2: "M", 3: "J", 4: "V", 5: "S", 6: "D"}
        def _wd(day: int) -> str:
            try:
                return weekday_names[date(year_num, month_num, day).weekday()]
            except Exception:
                return "-"

        area_meta = [
            {"name": "CORTE", "meta": int(round(float(((snapshot.get("sections") or {}).get("corte") or {}).get("meta_day") or 0))), "month": int(round(float(((snapshot.get("sections") or {}).get("corte") or {}).get("meta_month") or 0)))},
            {"name": "URRUTIA", "meta": int(round(float(((snapshot.get("sections") or {}).get("urrutia") or {}).get("meta_day") or 0))), "month": int(round(float(((snapshot.get("sections") or {}).get("urrutia") or {}).get("meta_month") or 0)))},
            {"name": "TIERRA FUEGO", "meta": int(round(float(((snapshot.get("sections") or {}).get("tierra_fuego") or {}).get("meta_day") or 0))), "month": int(round(float(((snapshot.get("sections") or {}).get("tierra_fuego") or {}).get("meta_month") or 0)))},
            {"name": "SUR", "meta": 0, "month": 0},
            {"name": "LAVANDERIA", "meta": int(round(float(((snapshot.get("sections") or {}).get("lavanderia") or {}).get("meta_day") or 0))), "month": int(round(float(((snapshot.get("sections") or {}).get("lavanderia") or {}).get("meta_month") or 0)))},
            {"name": "TERMINACION", "meta": int(round(float(((snapshot.get("sections") or {}).get("terminacion") or {}).get("meta_day") or 0))), "month": int(round(float(((snapshot.get("sections") or {}).get("terminacion") or {}).get("meta_month") or 0)))},
        ]

        week_defs = []
        for i, week in enumerate((snapshot.get("weeks") or []), start=1):
            day_list = [int(d) for d in (week.get("fechas") or []) if d]
            row_map = {str(r.get("key") or ""): (r.get("values") or []) for r in (week.get("rows") or [])}
            week_data = {
                "label": f"TOTAL SEM {i}",
                "days": day_list,
                "areas": {
                    "CORTE": row_map.get("corte", []),
                    "URRUTIA": row_map.get("urrutia", []),
                    "TIERRA FUEGO": [None for _ in day_list],
                    "SUR": row_map.get("sur", []),
                    "LAVANDERIA": row_map.get("lavanderia", []),
                    "TERMINACION": row_map.get("terminacion", []),
                },
            }
            week_data["rows"] = []
            for idx_day, day in enumerate(day_list):
                row = {"day": day, "weekday": _wd(day), "cells": []}
                for area in area_meta:
                    vals = week_data["areas"].get(area["name"], [])
                    prod = vals[idx_day] if idx_day < len(vals) else None
                    row["cells"].append({"prod": prod, "mues": None})
                week_data["rows"].append(row)
            totals = []
            for area in area_meta:
                vals = week_data["areas"].get(area["name"], [])
                totals.append(sum(int(v or 0) for v in vals))
            week_data["totals"] = totals
            week_defs.append(week_data)

        accum_lookup = {
            "CORTE": (snapshot.get("sections") or {}).get("corte") or {},
            "URRUTIA": (snapshot.get("sections") or {}).get("urrutia") or {},
            "TIERRA FUEGO": (snapshot.get("sections") or {}).get("tierra_fuego") or {},
            "SUR": {},
            "LAVANDERIA": (snapshot.get("sections") or {}).get("lavanderia") or {},
            "TERMINACION": (snapshot.get("sections") or {}).get("terminacion") or {},
        }
        summary_rows = []
        for area in area_meta:
            extra = accum_lookup.get(area["name"], {})
            actual = int(round(float(extra.get("actual") or 0)))
            proj = int(round(float(extra.get("projected") or 0)))
            avg = float(extra.get("daily_avg") or 0)
            pct = round((actual / proj) * 100, 2) if proj else 0.0
            summary_rows.append(
                {
                    "name": area["name"],
                    "meta_day": area["meta"],
                    "meta_month": area["month"],
                    "actual": actual,
                    "proj": proj,
                    "avg": avg,
                    "pct": pct,
                    "advance_vs_projection_pct": pct,
                    "diff": area["month"] - actual if area["month"] else 0,
                }
            )
        return {
            "title": f"MES DE PROCESO {str(sheet_name).upper()}",
            "days_note": f"Dias habiles del mes: {int(snapshot.get('days_month') or 0)}",
            "columns": area_meta,
            "weeks": week_defs,
            "summary_rows": summary_rows,
        }

    weekday_names = {
        0: "L",
        1: "M",
        2: "M",
        3: "J",
        4: "V",
        5: "S",
        6: "D",
    }

    def _wd(day: int) -> str:
        return weekday_names[date(2026, 4, day).weekday()]

    week_defs = [
        {
            "label": "TOTAL SEM 1",
            "days": [1, 2, 3, 4, 5],
            "areas": {
                "CORTE": [652, 198, None, None, None],
                "URRUTIA": [176, 162, None, None, None],
                "TIERRA FUEGO": [None, None, None, None, None],
                "SUR": [None, None, None, None, None],
                "LAVANDERIA": [456, 652, None, None, None],
                "TERMINACION": [152, 372, None, None, None],
            },
        },
        {
            "label": "TOTAL SEM 2",
            "days": [6, 7, 8, 9, 10, 11, 12],
            "areas": {
                "CORTE": [652, 652, 164, 600, None, None, None],
                "URRUTIA": [150, 198, 250, 200, 202, None, None],
                "TIERRA FUEGO": [None, None, None, None, None, None, None],
                "SUR": [None, None, None, None, None, None, None],
                "LAVANDERIA": [341, 535, 469, 274, 522, None, None],
                "TERMINACION": [468, 392, 419, 414, None, None, None],
            },
        },
        {
            "label": "TOTAL SEM 3",
            "days": [13, 14, 15, 16, 17, 18, 19],
            "areas": {
                "CORTE": [None, None, None, None, None, None, None],
                "URRUTIA": [None, None, None, None, None, None, None],
                "TIERRA FUEGO": [None, None, None, None, None, None, None],
                "SUR": [None, None, None, None, None, None, None],
                "LAVANDERIA": [None, None, None, None, None, None, None],
                "TERMINACION": [None, None, None, None, None, None, None],
            },
        },
        {
            "label": "TOTAL SEM 4",
            "days": [20, 21, 22, 23, 24, 25, 26],
            "areas": {
                "CORTE": [None, None, None, None, None, None, None],
                "URRUTIA": [None, None, None, None, None, None, None],
                "TIERRA FUEGO": [None, None, None, None, None, None, None],
                "SUR": [None, None, None, None, None, None, None],
                "LAVANDERIA": [None, None, None, None, None, None, None],
                "TERMINACION": [None, None, None, None, None, None, None],
            },
        },
        {
            "label": "TOTAL SEM 5",
            "days": [27, 28, 29, 30],
            "areas": {
                "CORTE": [None, None, None, None],
                "URRUTIA": [None, None, None, None],
                "TIERRA FUEGO": [None, None, None, None],
                "SUR": [None, None, None, None],
                "LAVANDERIA": [None, None, None, None],
                "TERMINACION": [None, None, None, None],
            },
        },
    ]

    area_meta = [
        {"name": "CORTE", "meta": 476, "month": 10000},
        {"name": "URRUTIA", "meta": 286, "month": 6000},
        {"name": "TIERRA FUEGO", "meta": 190, "month": 4000},
        {"name": "SUR", "meta": 0, "month": 0},
        {"name": "LAVANDERIA", "meta": 476, "month": 10000},
        {"name": "TERMINACION", "meta": 476, "month": 10000},
    ]

    for week in week_defs:
        week["rows"] = []
        for day in week["days"]:
            row = {"day": day, "weekday": _wd(day), "cells": []}
            for area in area_meta:
                prod = week["areas"][area["name"]][week["days"].index(day)]
                row["cells"].append({"prod": prod, "mues": None})
            week["rows"].append(row)
        totals = []
        for area in area_meta:
            vals = week["areas"][area["name"]]
            totals.append(sum(int(v or 0) for v in vals))
        week["totals"] = totals

    accum_lookup = {
        "CORTE": {"actual": 2918, "proj": 1905, "avg": 729.5, "pct": 153.3},
        "URRUTIA": {"actual": 936, "proj": 1143, "avg": 234.0, "pct": 81.8},
        "TIERRA FUEGO": {"actual": 0, "proj": 762, "avg": 0.0, "pct": 0.0},
        "SUR": {"actual": 0, "proj": 0, "avg": 0.0, "pct": 0.0},
        "LAVANDERIA": {"actual": 2453, "proj": 1905, "avg": 613.25, "pct": 128.8},
        "TERMINACION": {"actual": 1803, "proj": 1905, "avg": 450.75, "pct": 94.7},
    }

    summary_rows = []
    for area in area_meta:
        extra = accum_lookup[area["name"]]
        summary_rows.append(
            {
                "name": area["name"],
                "meta_day": area["meta"],
                "meta_month": area["month"],
                "actual": extra["actual"],
                "proj": extra["proj"],
                "avg": extra["avg"],
                "pct": extra["pct"],
                "advance_vs_projection_pct": round((extra["actual"] / extra["proj"]) * 100, 2) if extra["proj"] else 0.0,
                "diff": area["month"] - extra["actual"] if area["month"] else 0,
            }
        )

    return {
        "title": "MES DE PROCESO ABRIL 2026",
        "days_note": "Días hábiles restantes de 21",
        "columns": area_meta,
        "weeks": week_defs,
        "summary_rows": summary_rows,
    }


def _is_local_request() -> bool:
    host = str(request.host or "").strip().lower()
    server = str(request.environ.get("SERVER_NAME") or "").strip().lower()
    if host.startswith(("127.0.0.1", "localhost")):
        return True
    if server.startswith(("127.0.0.1", "localhost")):
        return True
    return os.environ.get("ADECOM_LOCAL_PREVIEW", "0").strip() == "1"


def _month_from_text(value: object) -> tuple[str, str]:
    raw = str(value or "").strip()
    if not raw:
        return "", ""
    probe = _norm_text(raw)
    months = {
        "enero": 1,
        "ene": 1,
        "febrero": 2,
        "feb": 2,
        "marzo": 3,
        "mar": 3,
        "abril": 4,
        "abr": 4,
        "mayo": 5,
        "may": 5,
        "junio": 6,
        "jun": 6,
        "julio": 7,
        "jul": 7,
        "agosto": 8,
        "ago": 8,
        "septiembre": 9,
        "setiembre": 9,
        "sep": 9,
        "octubre": 10,
        "oct": 10,
        "noviembre": 11,
        "nov": 11,
        "diciembre": 12,
        "dic": 12,
    }
    # YYYY-MM
    m = re.search(r"(\d{4})[-/](\d{1,2})", raw)
    if m:
        y = int(m.group(1))
        mo = int(m.group(2))
        if 1 <= mo <= 12:
            key = f"{y:04d}-{mo:02d}"
            return key, f"{mo:02d}/{y}"
    # MM-YYYY or MM/YY
    m2 = re.search(r"(\d{1,2})[-/](\d{2,4})", raw)
    if m2:
        mo = int(m2.group(1))
        y = int(m2.group(2))
        if y < 100:
            y += 2000
        if 1 <= mo <= 12:
            key = f"{y:04d}-{mo:02d}"
            return key, f"{mo:02d}/{y}"
    # Marzo 2026
    y3 = re.search(r"(20\d{2})", probe)
    for token, mo in months.items():
        if token in probe:
            y = int(y3.group(1)) if y3 else date.today().year
            key = f"{y:04d}-{mo:02d}"
            return key, f"{mo:02d}/{y}"
    return "", ""


def _parse_day(value: object) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        d = int(value)
        return d if 1 <= d <= 31 else 0
    text = str(value).strip()
    if text.isdigit():
        d = int(text)
        return d if 1 <= d <= 31 else 0
    m0 = re.match(r"^\s*(\d{1,2})\b", text)
    if m0:
        d = int(m0.group(1))
        return d if 1 <= d <= 31 else 0
    m = re.search(r"(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})", text)
    if m:
        d = int(m.group(1))
        if 1 <= d <= 31:
            return d
    m2 = re.search(r"(20\d{2})[-/](\d{1,2})[-/](\d{1,2})", text)
    if m2:
        d = int(m2.group(3))
        if 1 <= d <= 31:
            return d
    return 0


def _parse_proyeccion_rows_from_bytes(content: bytes, filename: str) -> list[dict[str, object]]:
    def parse_tabular_rows(values: list[tuple]) -> list[dict[str, object]]:
        if not values:
            return []
        header = [str(c or "").strip() for c in values[0]]
        rows_dict = [dict(zip(header, row)) for row in values[1:]]
        out: list[dict[str, object]] = []
        for row in rows_dict:
            keys = {_norm_text(k): k for k in row.keys() if str(k or "").strip()}
            area_key = next((keys[k] for k in keys if k in {"area", "seccion", "proceso"}), None)
            actual_key = next(
                (
                    keys[k]
                    for k in keys
                    if k in {"actual", "producido", "avance", "unidades", "cantidad", "real", "total"}
                ),
                None,
            )
            fecha_key = next((keys[k] for k in keys if k in {"fecha", "date"}), None)
            mes_key = next((keys[k] for k in keys if k in {"mes", "month", "periodo", "periodo mes"}), None)
            dia_key = next((keys[k] for k in keys if k in {"dia", "day"}), None)
            meta_key = next((keys[k] for k in keys if k in {"meta", "meta dia", "meta_diaria", "objetivo"}), None)
            persona_key = next((keys[k] for k in keys if k in {"persona", "operario", "nombre", "trabajador"}), None)
            meta_personal_key = next((keys[k] for k in keys if k in {"meta personal", "meta_personal", "meta_persona"}), None)
            if not area_key or not actual_key:
                continue
            month_key = ""
            month_label = ""
            day = 0
            if fecha_key:
                mk, ml = _month_from_text(row.get(fecha_key))
                month_key, month_label = mk, ml
                day = _parse_day(row.get(fecha_key))
            if not month_key and mes_key:
                mk, ml = _month_from_text(row.get(mes_key))
                month_key, month_label = mk, ml
            if day == 0 and dia_key:
                day = _parse_day(row.get(dia_key))
            if not month_key:
                continue
            out.append(
                {
                    "month_key": month_key,
                    "month_label": month_label or month_key,
                    "area": str(row.get(area_key) or "").strip(),
                    "persona": str(row.get(persona_key) or "").strip() if persona_key else "",
                    "day": day,
                    "actual": int(round(_to_float(row.get(actual_key)))),
                    "meta_day": int(round(_to_float(row.get(meta_key)))) if meta_key else 0,
                    "meta_personal": int(round(_to_float(row.get(meta_personal_key)))) if meta_personal_key else 0,
                }
            )
        return out

    def parse_matrix_rows(sheet_name: str, values: list[tuple]) -> list[dict[str, object]]:
        if len(values) < 6:
            return []
        mk, ml = _month_from_text(sheet_name)
        if not mk:
            for i in range(min(6, len(values))):
                for cell in values[i]:
                    mk, ml = _month_from_text(cell)
                    if mk:
                        break
                if mk:
                    break
        if not mk:
            return []
        head0 = list(values[0]) if len(values) > 0 else []
        head1 = list(values[1]) if len(values) > 1 else []
        head2 = list(values[2]) if len(values) > 2 else []
        head3 = list(values[3]) if len(values) > 3 else []
        head4 = list(values[4]) if len(values) > 4 else []

        days_month = 0
        line0 = [str(c or "").strip() for c in head0]
        for i, cell in enumerate(line0[:-1]):
            n = _norm_text(cell)
            if "dias habiles mes" in n:
                days_month = int(round(_to_float(line0[i + 1])))
                break
        if days_month <= 0:
            for row in values[5:]:
                day = _parse_day(row[0] if len(row) > 0 else "")
                if not day:
                    continue
                marker = _norm_text(str(row[2] if len(row) > 2 else ""))
                if marker in {"f", "feriado"}:
                    continue
                days_month += 1

        max_cols = max(len(head2), len(head3), len(head4))
        day_col = 0
        for idx in range(len(head2)):
            if _norm_text(head2[idx]) in {"dia", "day"}:
                day_col = idx
                break

        current_area = ""
        out: list[dict[str, object]] = []
        for col in range(day_col + 1, max_cols):
            area_raw = str(head2[col] if col < len(head2) else "").strip()
            if area_raw:
                current_area = area_raw
            area_name = current_area.strip()
            tipo_raw = str(head3[col] if col < len(head3) else "").strip()
            meta_day = int(round(_to_float(head4[col] if col < len(head4) else 0)))
            if not area_name and not tipo_raw:
                continue
            area_label = _canonical_area(area_name or tipo_raw)
            tipo_norm = _norm_text(tipo_raw)
            if tipo_norm and tipo_norm not in {"dia"}:
                area_label = f"{area_label} / {tipo_raw.upper()}"

            has_any = False
            for row in values[5:]:
                day = _parse_day(row[day_col] if len(row) > day_col else "")
                if not day:
                    continue
                raw_val = row[col] if col < len(row) else 0
                val = int(round(_to_float(raw_val)))
                if val == 0 and not str(raw_val or "").strip():
                    continue
                has_any = True
                out.append(
                    {
                        "month_key": mk,
                        "month_label": ml or mk,
                        "area": area_label,
                        "persona": "",
                        "day": day,
                        "actual": val,
                        "meta_day": max(meta_day, 0),
                        "meta_personal": 0,
                        "days_month": max(days_month, 0),
                    }
                )
            if not has_any and meta_day > 0:
                out.append(
                    {
                        "month_key": mk,
                        "month_label": ml or mk,
                        "area": area_label,
                        "persona": "",
                        "day": 0,
                        "actual": 0,
                        "meta_day": max(meta_day, 0),
                        "meta_personal": 0,
                        "days_month": max(days_month, 0),
                    }
                )
        return out

    ext = Path(filename or "").suffix.lower()
    if ext in {".csv", ".txt"}:
        text = content.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(io.StringIO(text), delimiter=";" if ";" in text.splitlines()[0] else ",")
        raw_rows = list(reader)
    elif ext == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        parsed: list[dict[str, object]] = []
        for ws in wb.worksheets:
            values = list(ws.iter_rows(values_only=True, min_row=1, max_row=220))
            if not values:
                continue
            sheet_rows = parse_matrix_rows(ws.title, values)
            if not sheet_rows:
                sheet_rows = parse_tabular_rows(values)
            parsed.extend(sheet_rows)
        return parsed
    elif ext == ".xls":
        try:
            import xlrd  # type: ignore
        except Exception as exc:
            raise ValueError("Para leer .xls falta dependencia xlrd. Guarda el archivo como .xlsx.") from exc
        wb = xlrd.open_workbook(file_contents=content)
        parsed: list[dict[str, object]] = []
        for sheet in wb.sheets():
            values: list[tuple] = []
            for r in range(min(sheet.nrows, 220)):
                row = tuple(sheet.cell_value(r, c) for c in range(sheet.ncols))
                values.append(row)
            if not values:
                continue
            sheet_rows = parse_matrix_rows(sheet.name, values)
            if not sheet_rows:
                sheet_rows = parse_tabular_rows(values)
            parsed.extend(sheet_rows)
        return parsed
    else:
        raise ValueError("Formato no soportado para proyeccion. Usa CSV, XLS o XLSX.")
    # CSV/TXT tabular
    parsed: list[dict[str, object]] = []
    for row in raw_rows:
        keys = {_norm_text(k): k for k in row.keys() if str(k or "").strip()}
        area_key = next((keys[k] for k in keys if k in {"area", "seccion", "proceso"}), None)
        actual_key = next((keys[k] for k in keys if k in {"actual", "real", "total", "cantidad", "producido"}), None)
        fecha_key = next((keys[k] for k in keys if k in {"fecha", "date"}), None)
        mes_key = next((keys[k] for k in keys if k in {"mes", "month", "periodo"}), None)
        dia_key = next((keys[k] for k in keys if k in {"dia", "day"}), None)
        meta_key = next((keys[k] for k in keys if k in {"meta", "meta dia", "meta_diaria"}), None)
        persona_key = next((keys[k] for k in keys if k in {"persona", "operario", "nombre", "trabajador"}), None)
        meta_personal_key = next((keys[k] for k in keys if k in {"meta personal", "meta_personal", "meta_persona"}), None)
        if not area_key or not actual_key:
            continue
        month_key = ""
        month_label = ""
        day = 0
        if fecha_key:
            mk, ml = _month_from_text(row.get(fecha_key))
            month_key, month_label = mk, ml
            day = _parse_day(row.get(fecha_key))
        if not month_key and mes_key:
            mk, ml = _month_from_text(row.get(mes_key))
            month_key, month_label = mk, ml
        if day == 0 and dia_key:
            day = _parse_day(row.get(dia_key))
        if not month_key:
            continue
        parsed.append(
            {
                "month_key": month_key,
                "month_label": month_label or month_key,
                "area": str(row.get(area_key) or "").strip(),
                "persona": str(row.get(persona_key) or "").strip() if persona_key else "",
                "day": day,
                "actual": int(round(_to_float(row.get(actual_key)))),
                "meta_day": int(round(_to_float(row.get(meta_key)))) if meta_key else 0,
                "meta_personal": int(round(_to_float(row.get(meta_personal_key)))) if meta_personal_key else 0,
            }
        )
    return parsed


def _build_proyeccion_view(monthly_goal: int, rows: list[dict[str, object]]) -> dict[str, object]:
    goal = max(int(monthly_goal or 0), 0)
    total_weight = sum(v for v in AREA_WEIGHTS.values() if v > 0) or 1

    by_month_area_daily: dict[str, dict[str, dict[int, int]]] = {}
    by_month_area_meta_day: dict[str, dict[str, int]] = {}
    by_month_days_count: dict[str, int] = {}
    by_month_label: dict[str, str] = {}
    by_month_people: dict[str, dict[tuple[str, str], dict[str, int]]] = {}
    for row in rows:
        month_key = str(row.get("month_key") or "").strip()
        if not month_key:
            continue
        by_month_label[month_key] = str(row.get("month_label") or month_key)
        area = str(row.get("area") or "").strip().upper()
        area = _canonical_area(area) if area in AREA_WEIGHTS else area
        persona = str(row.get("persona") or "").strip()
        day = int(row.get("day") or 0)
        day = day if 1 <= day <= 31 else 0
        by_month_area_daily.setdefault(month_key, {}).setdefault(area, {})
        by_month_area_daily[month_key][area][day] = by_month_area_daily[month_key][area].get(day, 0) + int(row.get("actual") or 0)
        if persona:
            p_key = (area, persona)
            by_month_people.setdefault(month_key, {})
            if p_key not in by_month_people[month_key]:
                by_month_people[month_key][p_key] = {"actual": 0, "meta_personal": 0}
            by_month_people[month_key][p_key]["actual"] += int(row.get("actual") or 0)
            by_month_people[month_key][p_key]["meta_personal"] = max(
                by_month_people[month_key][p_key]["meta_personal"],
                int(row.get("meta_personal") or 0),
            )
        meta_day = int(row.get("meta_day") or 0)
        if meta_day > 0:
            by_month_area_meta_day.setdefault(month_key, {})
            by_month_area_meta_day[month_key][area] = max(by_month_area_meta_day[month_key].get(area, 0), meta_day)
        dm = int(row.get("days_month") or 0)
        if dm > 0:
            by_month_days_count[month_key] = max(by_month_days_count.get(month_key, 0), dm)

    months: list[dict[str, object]] = []
    for month_key in sorted(by_month_area_daily.keys()):
        area_rows: list[dict[str, object]] = []
        month_total = 0
        month_areas = sorted(by_month_area_daily[month_key].keys())
        if not month_areas:
            continue
        days_month = by_month_days_count.get(month_key, 0)
        if days_month <= 0:
            day_set = set()
            for a in month_areas:
                for d in by_month_area_daily[month_key].get(a, {}):
                    if d > 0:
                        day_set.add(d)
            days_month = len(day_set)

        use_sheet_meta = bool(by_month_area_meta_day.get(month_key))
        area_target_map: dict[str, int] = {}
        for area in month_areas:
            area_weight = AREA_WEIGHTS.get(area, 0)
            fallback_target = int(round(goal * (area_weight / total_weight))) if area_weight > 0 else 0
            meta_day = by_month_area_meta_day.get(month_key, {}).get(area, 0)
            area_target = (meta_day * days_month) if use_sheet_meta and meta_day > 0 else fallback_target
            if area_target <= 0 and not use_sheet_meta:
                # para areas extras (ej. ventas) sin peso fijo, repartir una fraccion minima
                area_target = int(round(goal / max(len(month_areas), 1)))
            area_target_map[area] = area_target
            area_daily_map = by_month_area_daily.get(month_key, {}).get(area, {})
            area_actual = sum(int(v) for v in area_daily_map.values())
            month_total += area_actual
            ratio = (area_actual / area_target) if area_target > 0 else 0.0
            daily = [
                {"day": int(d), "actual": int(v)}
                for d, v in sorted(area_daily_map.items(), key=lambda x: x[0])
                if int(d) > 0
            ]
            area_rows.append(
                {
                    "area": area,
                    "target": area_target,
                    "actual": area_actual,
                    "ratio_pct": round(ratio * 100, 1),
                    "status": _status_from_ratio(ratio),
                    "daily_rows": daily,
                }
            )
        month_ratio = (month_total / goal) if goal > 0 else 0.0
        months.append(
            {
                "key": month_key,
                "label": by_month_label.get(month_key, month_key),
                "areas": area_rows,
                "area_target_map": area_target_map,
                "total_actual": month_total,
                "total_ratio_pct": round(month_ratio * 100, 1),
                "status": _status_from_ratio(month_ratio),
            }
        )

    if not months:
        today = date.today()
        month_key = f"{today.year:04d}-{today.month:02d}"
        month_label = f"{today.month:02d}/{today.year}"
        area_rows: list[dict[str, object]] = []
        for area, weight in AREA_WEIGHTS.items():
            target = int(round(goal * (weight / total_weight))) if total_weight > 0 else 0
            area_rows.append(
                {
                    "area": area,
                    "target": target,
                    "actual": 0,
                    "ratio_pct": 0.0,
                    "status": "red",
                    "daily_rows": [],
                }
            )
        months.append(
            {
                "key": month_key,
                "label": month_label,
                "areas": area_rows,
                "area_target_map": {a["area"]: int(a["target"]) for a in area_rows},
                "total_actual": 0,
                "total_ratio_pct": 0.0,
                "status": "red",
            }
        )

    default_month_key = months[-1]["key"] if months else ""
    active_month = next((m for m in months if m.get("key") == default_month_key), months[-1] if months else None)
    active_areas = list((active_month or {}).get("areas") or [])
    active_target_map = dict((active_month or {}).get("area_target_map") or {})
    people_rows: list[dict[str, object]] = []
    if active_month:
        month_people = by_month_people.get(default_month_key, {})
        by_area_people: dict[str, list[tuple[str, dict[str, int]]]] = {}
        for (area, persona), payload in month_people.items():
            by_area_people.setdefault(area, []).append((persona, payload))
        for area, people in sorted(by_area_people.items(), key=lambda x: x[0]):
            area_target = int(active_target_map.get(area, 0))
            count = len(people) or 1
            default_target = int(round(area_target / count)) if count else 0
            for persona, payload in sorted(people, key=lambda x: x[0].lower()):
                person_actual = int(payload.get("actual", 0))
                person_target = int(payload.get("meta_personal", 0)) or default_target
                ratio = (person_actual / person_target) if person_target > 0 else 0.0
                people_rows.append(
                    {
                        "area": area,
                        "persona": persona,
                        "target": person_target,
                        "actual": person_actual,
                        "ratio_pct": round(ratio * 100, 1),
                        "status": _status_from_ratio(ratio),
                    }
                )

    return {
        "monthly_goal": goal,
        "weekly_goal": goal,
        "total_actual": int((active_month or {}).get("total_actual") or 0),
        "total_ratio_pct": float((active_month or {}).get("total_ratio_pct") or 0.0),
        "status": str((active_month or {}).get("status") or "red"),
        "areas": active_areas,
        "people": people_rows,
        "months": months,
        "default_month_key": default_month_key,
    }


def _load_proyeccion_state() -> dict[str, object]:
    if not PROYECCION_STATE_PATH.exists():
        auto_rows = _autoload_proyeccion_rows()
        return {"monthly_goal": 12000, "rows": auto_rows, "view": _build_proyeccion_view(12000, auto_rows)}
    try:
        payload = json.loads(PROYECCION_STATE_PATH.read_text(encoding="utf-8"))
        goal = int(payload.get("monthly_goal") or payload.get("weekly_goal") or 12000)
        rows = payload.get("rows") or []
        if not isinstance(rows, list):
            rows = []
        if not rows:
            rows = _autoload_proyeccion_rows()
        view = _build_proyeccion_view(goal, rows)
        return {"monthly_goal": goal, "rows": rows, "view": view}
    except Exception:
        auto_rows = _autoload_proyeccion_rows()
        return {"monthly_goal": 12000, "rows": auto_rows, "view": _build_proyeccion_view(12000, auto_rows)}


def _save_proyeccion_state(monthly_goal: int, rows: list[dict[str, object]]) -> None:
    payload = {"monthly_goal": int(monthly_goal), "rows": rows}
    PROYECCION_STATE_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _autoload_proyeccion_rows() -> list[dict[str, object]]:
    candidates = []
    if AUTOLOAD_DIR.exists():
        for pattern in ("*MOVTOS*SECCIONES*.xlsx", "*MOVTOS*SECCIONES*.xls", "*MOVTOS*.xlsx", "*MOVTOS*.xls"):
            candidates.extend(sorted(AUTOLOAD_DIR.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True))
    seen: set[str] = set()
    for path in candidates:
        key = str(path.resolve()).lower()
        if key in seen:
            continue
        seen.add(key)
        try:
            rows = _parse_proyeccion_rows_from_bytes(path.read_bytes(), path.name)
            if rows:
                return rows
        except Exception:
            continue
    return []


def _format_fecha_display(fecha_iso: object) -> str:
    raw = str(fecha_iso or "").strip()
    parts = raw.split("-")
    if len(parts) == 3 and all(parts):
        return f"{parts[2]}/{parts[1]}/{parts[0]}"
    return raw


def _full_table_stage_label(row: dict[str, object]) -> str:
    stages = [
        ("bodega", "BODEGA"),
        ("corte_1", "CORTE"),
        ("taller", "TALLER"),
        ("t_externo", "T.EXTERNO"),
        ("limpiado", "LIMPIADO"),
        ("lavanderia", "LAVANDERIA"),
        ("terminacion", "TERMINACION"),
        ("muestra", "MUESTRA"),
        ("segunda", "SEGUNDA"),
    ]
    for key, label in stages:
        if int(row.get(key) or 0) > 0:
            return label
    return "Sin movimiento"


def _full_table_restante_detalle(row: dict[str, object]) -> str:
    labels = [
        ("corte_1", "Corte"),
        ("taller", "Taller"),
        ("t_externo", "T. Externo"),
        ("limpiado", "Limpiado"),
        ("lavanderia", "Lavanderia"),
        ("terminacion", "Terminacion"),
        ("muestra", "Muestra"),
        ("segunda", "Segunda"),
    ]
    parts: list[str] = []
    for key, label in labels:
        value = int(row.get(key) or 0)
        if value > 0:
            parts.append(f"{label}: {value}")
    return " | ".join(parts) if parts else "Sin restante"


def _temporada_from_seed_saldos(path: Path) -> str:
    name = str(path.name or "").upper()
    m = re.search(r"SALDOS-SECCI\s*(\d{2})", name)
    if m:
        return m.group(1)
    m2 = re.search(r"(\d{2})", name)
    return m2.group(1) if m2 else ""


def _full_table_tipo_label(row: dict[str, object]) -> str:
    corte = str(row.get("corte") or "").strip()
    normalized = corte.lstrip("0")
    if normalized.startswith("96"):
        return "muestras"
    return "produccion"


def _load_full_table_rows_from_seed() -> tuple[list[dict[str, object]], dict[str, int], list[str]]:
    numeric_fields = [
        "programa",
        "proceso",
        "bodega",
        "saldo",
        "corte_1",
        "taller",
        "t_externo",
        "limpiado",
        "lavanderia",
        "terminacion",
        "muestra",
        "segunda",
    ]
    totals = {key: 0 for key in numeric_fields}
    rows: list[dict[str, object]] = []
    temporadas_seen: set[str] = set()
    seen_rows: set[tuple] = set()

    saldos_seed_paths = sorted(
        {p.resolve(): p for p in [*SEED_DIR.glob("SALDOS-SECCI*.TXT"), *SEED_DIR.glob("SALDOS-SECCI*.txt")]}.values(),
        key=lambda p: p.name.lower(),
    )
    for path in saldos_seed_paths:
        temporada = _temporada_from_seed_saldos(path)
        if temporada not in {"42", "43"}:
            continue
        try:
            parsed_rows = parse_saldos_txt(path.read_bytes())
        except Exception:
            continue
        temporadas_seen.add(temporada)
        for parsed in parsed_rows:
            row = dict(parsed)
            row_key = (
                temporada,
                str(row.get("articulo") or "").strip(),
                str(row.get("corte") or "").strip(),
                str(row.get("fecha_iso") or "").strip(),
                *[int(row.get(key) or 0) for key in numeric_fields],
            )
            if row_key in seen_rows:
                continue
            seen_rows.add(row_key)
            for key in numeric_fields:
                row[key] = int(row.get(key) or 0)
                totals[key] += int(row[key] or 0)
            row["temporada"] = temporada
            row["tipo"] = _full_table_tipo_label(row)
            row["fecha_display"] = _format_fecha_display(row.get("fecha_iso"))
            row["proceso_actual"] = _full_table_stage_label(row)
            row["restante_detalle"] = _full_table_restante_detalle(row)
            rows.append(row)

    rows.sort(
        key=lambda r: (
            str(r.get("fecha_iso") or ""),
            str(r.get("corte") or ""),
            str(r.get("articulo") or ""),
        ),
        reverse=True,
    )
    temporadas = sorted(temporadas_seen, key=lambda x: int(x) if str(x).isdigit() else 999)
    return rows, totals, temporadas


def ensure_seed_data() -> None:
    if not ENABLE_SEED:
        init_db(DB_PATH)
        return
    init_db(DB_PATH)
    seed_saldos_files = sorted(
        {p.resolve(): p for p in [*SEED_DIR.glob("SALDOS-SECCI*.TXT"), *SEED_DIR.glob("SALDOS-SECCI*.txt")]}.values(),
        key=lambda p: p.name.lower(),
    )
    if not seed_saldos_files and SEED_SALDOS.exists():
        seed_saldos_files = [SEED_SALDOS]
    if seed_saldos_files and _table_count("saldos_seccion") == 0:
        saldos_rows: list[dict] = []
        for seed_file in seed_saldos_files:
            saldos_rows.extend(parse_saldos_txt(seed_file.read_bytes()))
        if saldos_rows:
            import_rows(DB_PATH, saldos_rows, replace_all=True, accumulate_on_conflict=True)
    if SEED_PEDIDOS.exists() and _table_count("pedidos_talla") == 0:
        pedidos_rows = parse_pedidos_talla_txt(SEED_PEDIDOS.read_bytes())
        if pedidos_rows:
            import_pedidos_talla_rows(DB_PATH, pedidos_rows)
    if SEED_COMPARATIVO.exists() and _table_count("comparativo_clientes") == 0:
        comparativo_rows = parse_comparativo_clientes_txt(SEED_COMPARATIVO.read_bytes())
        if comparativo_rows:
            import_comparativo_clientes_rows(DB_PATH, comparativo_rows)
    if SEED_DEUDAS.exists() and _table_count("deuda_clientes") == 0:
        deuda_rows = parse_deudas_vencidas_csv(SEED_DEUDAS.read_bytes())
        if deuda_rows:
            import_deuda_clientes_rows(DB_PATH, deuda_rows)


ensure_seed_data()

_auto_refresh_web_on_startup()
_start_auto_refresh_web_loop()


@app.template_filter("miles")
def miles(value):
    try:
        number = int(value or 0)
    except (TypeError, ValueError):
        return value
    return f"{number:,}".replace(",", ".")


def _ventas_docs_file() -> Path | None:
    latest = _find_latest_autoload_file(
        "*VENTAS-TOD-*.CSV",
        "*VENTAS-TOD-*.csv",
        "*VENTAS*.CSV",
        "*VENTAS*.csv",
    )
    if latest and latest.exists():
        return latest
    if SEED_VENTAS_DOCS.exists():
        return SEED_VENTAS_DOCS
    return None


def _to_int(value: object) -> int:
    raw = str(value or "").strip().replace(".", "").replace(" ", "")
    if not raw:
        return 0
    try:
        return int(raw)
    except ValueError:
        return 0


def _inventory_norm(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-zA-Z0-9]+", " ", text).strip().lower()
    return text


def _inventory_to_float(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value or "").strip()
    if not text:
        return 0.0
    text = text.replace(" ", "").replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _extract_collection_from_articulo(articulo: str) -> str:
    digits = "".join(ch for ch in str(articulo or "") if ch.isdigit())
    if len(digits) >= 6:
        return digits[2:6]
    if len(digits) >= 4:
        return digits[:4]
    return (str(articulo or "").strip() or "Sin coleccion")[:12]


def _extract_collection_from_sheet(sheet_name: str) -> str:
    text = _inventory_norm(sheet_name)
    m = re.search(r"(\d{2})", text)
    if m:
        return m.group(1)
    return (str(sheet_name or "").strip() or "Sin coleccion")[:20]


def _load_inventory_rows_from_excel() -> tuple[list[dict[str, object]], str, str]:
    candidate_paths = [INVENTORY_BOOK_PATH, SEED_DIR / "INVENTARIO 01-04 COMPLETO.xlsx"]
    existing_paths = [p for p in candidate_paths if p and p.exists()]
    if not existing_paths:
        return [], "", "No se encontro el archivo de inventario configurado."

    try:
        from openpyxl import load_workbook
    except Exception as exc:
        return [], "", f"No se pudo cargar openpyxl: {exc}"

    wb = None
    active_path = None
    last_error = ""
    for p in existing_paths:
        try:
            wb = load_workbook(p, data_only=True, read_only=True)
            active_path = p
            break
        except Exception as exc:
            last_error = str(exc)
            continue
    if wb is None or active_path is None:
        return [], "", f"No se pudo abrir el inventario: {last_error}"

    rows_out: list[dict[str, object]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_idx = None
        header_map: dict[str, int] = {}
        size_cols: dict[int, int] = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
            normalized = [_inventory_norm(v) for v in row]
            if not any(normalized):
                continue
            size_cols = {}
            for col_idx, cell in enumerate(normalized):
                if not cell:
                    continue
                if "articulo" in cell or cell in {"codigo", "cod articulo", "cod"}:
                    header_map["articulo"] = col_idx
                if "tiro" in cell:
                    header_map["tiro"] = col_idx
                if "bota" in cell:
                    header_map["bota"] = col_idx
                if "color" in cell:
                    header_map["color"] = col_idx
                if cell in {"coleccion", "temporada", "linea"}:
                    header_map["coleccion"] = col_idx
                if (
                    "stock" in cell
                    or "saldo" in cell
                    or "existencia" in cell
                    or "disponible" in cell
                    or cell in {"total", "cantidad", "cant"}
                ):
                    header_map.setdefault("stock", col_idx)
            for col_idx, raw in enumerate(row):
                size_num = None
                if isinstance(raw, (int, float)):
                    size_num = int(raw)
                else:
                    digits = re.sub(r"\D+", "", str(raw or ""))
                    if digits:
                        size_num = int(digits)
                if size_num and 30 <= size_num <= 60:
                    size_cols[size_num] = col_idx
            if "articulo" in header_map and "stock" in header_map:
                header_idx = row_idx
                break

        if not header_idx or "articulo" not in header_map or "stock" not in header_map:
            continue

        for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
            articulo = str(row[header_map["articulo"]] or "").strip() if header_map["articulo"] < len(row) else ""
            if not articulo:
                continue
            stock_value = _inventory_to_float(row[header_map["stock"]] if header_map["stock"] < len(row) else 0)
            sizes_map: dict[int, int] = {}
            for size_num, col_idx in sorted(size_cols.items()):
                value = int(round(_inventory_to_float(row[col_idx] if col_idx < len(row) else 0)))
                sizes_map[size_num] = value
            sizes_total = sum(v for v in sizes_map.values() if v > 0)
            if stock_value <= 0 and sizes_total > 0:
                stock_value = float(sizes_total)
            if stock_value <= 0:
                continue
            tiro = str(row[header_map["tiro"]] or "").strip() if "tiro" in header_map and header_map["tiro"] < len(row) else ""
            bota = str(row[header_map["bota"]] or "").strip() if "bota" in header_map and header_map["bota"] < len(row) else ""
            color = str(row[header_map["color"]] or "").strip() if "color" in header_map and header_map["color"] < len(row) else ""
            coleccion = ""
            if "coleccion" in header_map and header_map["coleccion"] < len(row):
                coleccion = str(row[header_map["coleccion"]] or "").strip()
            if not coleccion:
                coleccion = _extract_collection_from_sheet(sheet_name)
            if not coleccion:
                coleccion = _extract_collection_from_articulo(articulo)
            rows_out.append(
                {
                    "articulo": articulo,
                    "tiro": tiro,
                    "bota": bota,
                    "color": color,
                    "coleccion": coleccion,
                    "stock": int(round(stock_value)),
                    "sizes": sizes_map,
                    "fuente": "excel",
                }
            )
    if not rows_out:
        return [], str(active_path), "No se detectaron columnas compatibles (articulo + stock) con stock positivo."
    return rows_out, str(active_path), ""


def _build_inventory_book_dashboard_from_db_rows(rows: list[dict[str, object]], source_label: str = "Base de datos") -> dict[str, object]:
    grouped: dict[str, dict[str, object]] = {}
    for item in rows:
        key = str(item.get("coleccion") or "").strip() or "Sin coleccion"
        sizes = {
            36: int(item.get("talla_36") or 0),
            38: int(item.get("talla_38") or 0),
            40: int(item.get("talla_40") or 0),
            42: int(item.get("talla_42") or 0),
            44: int(item.get("talla_44") or 0),
            46: int(item.get("talla_46") or 0),
        }
        article_item = {
            "id": int(item.get("id") or 0),
            "coleccion": key,
            "articulo": str(item.get("articulo") or "").strip(),
            "tiro": str(item.get("tiro") or "").strip() or "-",
            "bota": str(item.get("bota") or "").strip() or "-",
            "color": str(item.get("color") or "").strip() or "-",
            "stock": int(item.get("stock") or 0),
            "sizes": sizes,
        }
        bucket = grouped.setdefault(
            key,
            {
                "name": key,
                "total_stock": 0,
                "items_count": 0,
                "articles": [],
                "size_headers": {36, 38, 40, 42, 44, 46},
            },
        )
        bucket["total_stock"] = int(bucket["total_stock"]) + int(article_item["stock"] or 0)
        bucket["items_count"] = int(bucket["items_count"]) + 1
        bucket["articles"].append(article_item)

    collections = []
    for bucket in grouped.values():
        articles = sorted(
            bucket["articles"],
            key=lambda x: (-int(x.get("stock") or 0), str(x.get("articulo") or "")),
        )
        collections.append(
            {
                "name": bucket["name"],
                "total_stock": int(bucket["total_stock"]),
                "items_count": int(bucket["items_count"]),
                "articles": articles,
                "size_headers": sorted(int(x) for x in (bucket.get("size_headers") or set())),
            }
        )

    def _collection_sort_key(item: dict[str, object]):
        name = str(item.get("name") or "").strip()
        if name == "42":
            return (0, 0, name)
        if name.isdigit():
            return (1, int(name), name)
        return (2, 9999, name)

    collections.sort(key=_collection_sort_key)
    default_collection = "42"
    if not any(str(c.get("name") or "") == "42" for c in collections):
        default_collection = str((collections[0] if collections else {}).get("name") or "")

    return {
        "available": bool(collections),
        "error": "" if collections else "No hay datos de stock cargados.",
        "file_name": source_label,
        "path": source_label,
        "collections": collections,
        "default_collection": default_collection,
        "total_stock": sum(int(c["total_stock"]) for c in collections),
        "total_items": sum(int(c["items_count"]) for c in collections),
    }


def _load_inventory_book_dashboard() -> dict[str, object]:
    rows_db = query_inventory_stock_rows(DB_PATH)
    if rows_db:
        return _build_inventory_book_dashboard_from_db_rows(rows_db, "Base de datos")

    parsed_rows, excel_path, err = _load_inventory_rows_from_excel()
    if parsed_rows:
        replace_inventory_stock_rows(DB_PATH, parsed_rows)
        rows_db = query_inventory_stock_rows(DB_PATH)
        if rows_db:
            source = Path(excel_path).name if excel_path else "Excel"
            return _build_inventory_book_dashboard_from_db_rows(rows_db, source)
    return {
        "available": False,
        "error": err or "No hay datos de stock cargados.",
        "file_name": "",
        "path": excel_path or str(INVENTORY_BOOK_PATH),
        "collections": [],
        "default_collection": "42",
        "total_stock": 0,
        "total_items": 0,
    }


def _inventory_row_from_form(form: object) -> dict[str, object]:
    get = form.get
    return {
        "coleccion": str(get("coleccion") or "").strip(),
        "articulo": str(get("articulo") or "").strip(),
        "tiro": str(get("tiro") or "").strip(),
        "bota": str(get("bota") or "").strip(),
        "color": str(get("color") or "").strip(),
        "talla_36": _to_int(get("talla_36")),
        "talla_38": _to_int(get("talla_38")),
        "talla_40": _to_int(get("talla_40")),
        "talla_42": _to_int(get("talla_42")),
        "talla_44": _to_int(get("talla_44")),
        "talla_46": _to_int(get("talla_46")),
        "stock": _to_int(get("stock")),
        "fuente": "web",
    }


def _load_ventas_docs_summary() -> dict:
    base = {
        "available": False,
        "file_name": "",
        "latest_date": "",
        "latest_date_label": "",
        "dates": [],
        "by_date": {},
    }
    path = _ventas_docs_file()
    if not path or not path.exists():
        return base

    rows: list[dict[str, str]] | None = None
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            with path.open("r", encoding=encoding, newline="") as fh:
                rows = [
                    {str(k or "").strip(): str(v or "").strip() for k, v in row.items()}
                    for row in csv.DictReader(fh, delimiter=";")
                ]
            break
        except Exception:
            rows = None
            continue
    if not rows:
        return base

    grouped: dict[str, dict] = {}
    for idx, row in enumerate(rows, start=1):
        raw_date = str(row.get("fecha") or "").strip()
        if not raw_date:
            continue
        try:
            date_obj = datetime.strptime(raw_date, "%d-%m-%Y").date()
        except ValueError:
            continue
        date_iso = date_obj.isoformat()
        bucket = grouped.setdefault(
            date_iso,
            {
                "date": date_iso,
                "label": date_obj.strftime("%d/%m/%Y"),
                "total_docs": set(),
                "f_elec_docs": set(),
                "boletas_docs": set(),
                "n_credito_docs": set(),
                "n_debito_docs": set(),
                "sin_tipo_docs": set(),
                "sales_net": 0,
                "units_net": 0,
                "f_elec_amount": 0,
                "boletas_amount": 0,
                "n_credito_amount": 0,
            },
        )
        doc_number = str(row.get("Numero") or "").strip()
        doc_key = doc_number or f"row-{idx}"
        doc_type = str(row.get("Tipo") or "").strip().lower()
        if doc_type == "f/elec":
            bucket["f_elec_docs"].add(doc_key)
        elif doc_type == "bole":
            bucket["boletas_docs"].add(doc_key)
        elif doc_type == "n/cre":
            bucket["n_credito_docs"].add(doc_key)
        elif doc_type == "ndeb":
            bucket["n_debito_docs"].add(doc_key)
        else:
            bucket["sin_tipo_docs"].add(doc_key)
        bucket["total_docs"].add(doc_key)
        total_value = _to_int(row.get("Total"))
        units_value = _to_int(row.get("Cant"))
        bucket["sales_net"] += total_value
        bucket["units_net"] += units_value
        if doc_type == "f/elec":
            bucket["f_elec_amount"] += total_value
        elif doc_type == "bole":
            bucket["boletas_amount"] += total_value
        elif doc_type == "n/cre":
            bucket["n_credito_amount"] += total_value

    if not grouped:
        return base

    by_date: dict[str, dict] = {}
    for date_iso, item in grouped.items():
        by_date[date_iso] = {
            "date": date_iso,
            "label": item["label"],
            "total_docs": len(item["total_docs"]),
            "f_elec_docs": len(item["f_elec_docs"]),
            "boletas_docs": len(item["boletas_docs"]),
            "n_credito_docs": len(item["n_credito_docs"]),
            "n_debito_docs": len(item["n_debito_docs"]),
            "sin_tipo_docs": len(item["sin_tipo_docs"]),
            "sales_net": item["sales_net"],
            "units_net": item["units_net"],
            "f_elec_amount": item["f_elec_amount"],
            "boletas_amount": item["boletas_amount"],
            "n_credito_amount": item["n_credito_amount"],
        }

    dates = sorted(by_date.keys())
    latest = dates[-1]
    return {
        "available": True,
        "file_name": path.name,
        "latest_date": latest,
        "latest_date_label": by_date[latest]["label"],
        "dates": [{"value": key, "label": by_date[key]["label"]} for key in reversed(dates)],
        "by_date": by_date,
    }


def _load_disponibles_ranking_4200(ventas_top_articulos: list[dict]) -> dict:
    base = {
        "available": False,
        "file_name": "",
        "rows": [],
        "familias_con_stock": 0,
        "total_disponible": 0,
    }
    candidates = sorted(
        list(AUTOLOAD_DIR.glob("Cortes 4200*.xlsx")) + list(AUTOLOAD_DIR.glob("Cortes 4200*.xls")),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    path = candidates[0] if candidates else (SEED_CORTES_4200_XLSX if SEED_CORTES_4200_XLSX.exists() else None)
    if not path or not path.exists():
        return base
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True, read_only=True)
        if "Ranking 42" not in wb.sheetnames:
            return base
        ws = wb["Ranking 42"]
    except Exception:
        return base

    def _family_key(raw: object) -> str:
        text = str(raw or "").strip().upper()
        if not text:
            return ""
        digits = "".join(ch for ch in text if ch.isdigit())
        if len(digits) >= 6 and digits.startswith("42"):
            return digits[:4]
        if len(digits) >= 4 and digits.startswith("42"):
            return digits[:4]
        match = re.search(r"(42\d{2})", text)
        return match.group(1) if match else ""

    def _value_to_int(raw: object) -> int:
        if raw is None:
            return 0
        if isinstance(raw, (int, float)):
            return int(raw)
        text = str(raw).strip().replace(".", "").replace(",", "")
        if not text:
            return 0
        try:
            return int(float(text))
        except ValueError:
            return 0

    def _variant_suffix(raw: object) -> str:
        text = str(raw or "").strip().upper()
        digits = "".join(ch for ch in text if ch.isdigit())
        if len(digits) >= 6:
            return digits[-2:]
        if "-" in text:
            tail = text.split("-")[-1].strip()
            tail_digits = "".join(ch for ch in tail if ch.isdigit())
            if tail_digits:
                return tail_digits[-2:].zfill(2)
        return "00"

    ranking_by_family: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        row_type = str(row[18] or "").strip().lower() if len(row) > 18 else ""
        articulo_raw = str(row[19] or "").strip() if len(row) > 19 else ""
        if row_type != "stock" or not articulo_raw:
            continue
        family = _family_key(articulo_raw)
        if not family:
            continue
        descripcion = str(row[24] or "").strip() if len(row) > 24 else ""
        total_raw = _value_to_int(row[33] if len(row) > 33 else 0)
        disponible = abs(total_raw) if total_raw < 0 else 0
        bucket = ranking_by_family.setdefault(
            family,
            {
                "familia": family,
                "disponible_total": 0,
                "variantes": [],
            },
        )
        bucket["disponible_total"] += disponible
        bucket["variantes"].append(
            {
                "codigo": articulo_raw,
                "sufijo": _variant_suffix(articulo_raw),
                "descripcion": descripcion or "-",
                "disponible": disponible,
            }
        )

    ventas_por_familia: dict[str, int] = {}
    articulo_por_familia: dict[str, str] = {}
    for item in ventas_top_articulos:
        articulo = str(item.get("articulo") or "").strip()
        family = _family_key(articulo)
        if not family:
            continue
        ventas_por_familia[family] = ventas_por_familia.get(family, 0) + int(item.get("total") or 0)
        articulo_por_familia.setdefault(family, articulo)

    rows = []
    for family, total_pedidos in ventas_por_familia.items():
        ranking = ranking_by_family.get(family) or {"disponible_total": 0, "variantes": []}
        vendidos_por_sufijo: dict[str, int] = {}
        for articulo_item in ventas_top_articulos:
            articulo_codigo = str(articulo_item.get("articulo") or "").strip()
            if _family_key(articulo_codigo) != family:
                continue
            sufijo = _variant_suffix(articulo_codigo)
            vendidos_por_sufijo[sufijo] = vendidos_por_sufijo.get(sufijo, 0) + int(articulo_item.get("total") or 0)
        variantes = sorted(
            ranking.get("variantes") or [],
            key=lambda item: (-int(item.get("disponible") or 0), str(item.get("codigo") or "")),
        )
        top_variantes = [item for item in variantes if int(item.get("disponible") or 0) > 0][:3]
        rows.append(
            {
                "familia": family,
                "articulo": articulo_por_familia.get(family, family),
                "pedidos_total": int(total_pedidos or 0),
                "disponible_total": int(ranking.get("disponible_total") or 0),
                "status": "Disponible" if int(ranking.get("disponible_total") or 0) > 0 else "No disponible",
                "status_class": "estado-pill estado-pill-green" if int(ranking.get("disponible_total") or 0) > 0 else "estado-pill estado-pill-red",
                "variantes": [
                    {
                        "codigo": str(item.get("codigo") or ""),
                        "sufijo": str(item.get("sufijo") or "00"),
                        "descripcion": str(item.get("descripcion") or "-"),
                        "vendido": int(vendidos_por_sufijo.get(str(item.get("sufijo") or "00"), 0)),
                        "disponible": int(item.get("disponible") or 0),
                    }
                    for item in variantes
                ],
                "variantes_label": " | ".join(
                    f"{item['codigo']}: {item['disponible']}" for item in top_variantes
                )
                or "Sin disponible",
            }
        )

    rows.sort(key=lambda item: (-int(item["pedidos_total"]), -int(item["disponible_total"]), item["familia"]))
    return {
        "available": True,
        "file_name": path.name,
        "rows": rows,
        "familias_con_stock": sum(1 for item in rows if int(item.get("disponible_total") or 0) > 0),
        "total_disponible": sum(int(item.get("disponible_total") or 0) for item in rows),
    }


@app.get("/")
def index():
    assistant_provider = os.environ.get("ADECOM_ASSISTANT_PROVIDER", "local").strip().lower()
    if assistant_provider in {"google", "gemini"}:
        assistant_provider = "gemini"
    else:
        assistant_provider = "local"
    if not ASSISTANT_ENABLED:
        assistant_provider = "off"
    filters = {
        "q": request.args.get("q", "").strip(),
        "fecha": request.args.get("fecha", "").strip(),
        "articulo_exact": request.args.get("articulo_exact", "").strip(),
    }
    rows, totals, summary = query_rows(DB_PATH, filters)
    pedidos_q = filters["q"] or filters["articulo_exact"]
    pedidos_sections = query_pedidos_talla_sections(DB_PATH, pedidos_q)
    exs_summary = query_exs_balance_summary(DB_PATH, filters["q"])
    comparativo_summary = query_comparativo_clientes(DB_PATH, "")
    pedidos_count = sum(len(section_rows) for section_rows in pedidos_sections.values())
    search_error = ""
    search_success = ""
    if filters["articulo_exact"] and not rows and pedidos_count == 0:
        search_error = "Articulo no encontrado"
    elif filters["q"] and not rows and pedidos_count == 0:
        search_error = "No se encontraron resultados. Escriba el articulo completo o familia. Ej: 01420100 o 4201."
    elif (filters["articulo_exact"] or filters["q"]) and (rows or pedidos_count > 0):
        search_success = "Encontrado"
    ventas_rows = pedidos_sections.get("ventas", [])
    saldo_rows = pedidos_sections.get("saldo", [])
    corte_rows = pedidos_sections.get("corte", [])
    ventas_total = sum(int(r.get("total") or 0) for r in ventas_rows)
    ventas_por_articulo: dict[str, int] = {}
    ventas_por_familia: dict[str, dict] = {}
    ventas_por_talla: dict[int, int] = {}
    for r in ventas_rows:
        articulo = str(r.get("articulo") or "").strip()
        total = int(r.get("total") or 0)
        if not articulo:
            continue
        ventas_por_articulo[articulo] = ventas_por_articulo.get(articulo, 0) + total
        familia = articulo[2:6] if len(articulo) >= 6 else articulo
        if familia not in ventas_por_familia:
            ventas_por_familia[familia] = {
                "familia": familia,
                "total": 0,
                "saldo_total": 0,
                "articulos": {},
                "sufijos": {},
                "sufijos_saldo": {},
                "tallas": {},
                "tallas_corte": {},
                "articulos_tallas": {},
            }
        if articulo not in ventas_por_familia[familia]["articulos_tallas"]:
            ventas_por_familia[familia]["articulos_tallas"][articulo] = {
                "articulo": articulo,
                "vendida": {},
                "cortada": {},
            }
        for item in r.get("tallas_items") or []:
            talla = int(item.get("talla") or 0)
            cantidad = int(item.get("cantidad") or 0)
            if talla > 0:
                ventas_por_talla[talla] = ventas_por_talla.get(talla, 0) + cantidad
                ventas_por_familia[familia]["tallas"][talla] = (
                    int(ventas_por_familia[familia]["tallas"].get(talla) or 0) + cantidad
                )
                ventas_por_familia[familia]["articulos_tallas"][articulo]["vendida"][talla] = (
                    int(ventas_por_familia[familia]["articulos_tallas"][articulo]["vendida"].get(talla) or 0)
                    + cantidad
                )
        ventas_por_familia[familia]["total"] += total
        if articulo not in ventas_por_familia[familia]["articulos"]:
            ventas_por_familia[familia]["articulos"][articulo] = {
                "articulo": articulo,
                "total": 0,
            }
        ventas_por_familia[familia]["articulos"][articulo]["total"] += total
        prefijo = articulo[:2] if len(articulo) >= 2 else ""
        sufijo = articulo[-2:] if len(articulo) >= 2 else articulo
        sufijo_key = f"{prefijo}|{sufijo}"
        if sufijo_key not in ventas_por_familia[familia]["sufijos"]:
            ventas_por_familia[familia]["sufijos"][sufijo_key] = {
                "prefijo": prefijo,
                "sufijo": sufijo,
                "total": 0,
            }
        ventas_por_familia[familia]["sufijos"][sufijo_key]["total"] += total

    for r in saldo_rows:
        articulo = str(r.get("articulo") or "").strip()
        total = int(r.get("total") or 0)
        if not articulo:
            continue
        familia = articulo[2:6] if len(articulo) >= 6 else articulo
        if familia not in ventas_por_familia:
            ventas_por_familia[familia] = {
                "familia": familia,
                "total": 0,
                "saldo_total": 0,
                "articulos": {},
                "sufijos": {},
                "sufijos_saldo": {},
                "tallas": {},
                "tallas_corte": {},
                "articulos_tallas": {},
            }
        ventas_por_familia[familia]["saldo_total"] += total
        prefijo = articulo[:2] if len(articulo) >= 2 else ""
        sufijo = articulo[-2:] if len(articulo) >= 2 else articulo
        sufijo_key = f"{prefijo}|{sufijo}"
        if sufijo_key not in ventas_por_familia[familia]["sufijos_saldo"]:
            ventas_por_familia[familia]["sufijos_saldo"][sufijo_key] = {
                "prefijo": prefijo,
                "sufijo": sufijo,
                "total": 0,
            }
        ventas_por_familia[familia]["sufijos_saldo"][sufijo_key]["total"] += total

    for r in corte_rows:
        articulo = str(r.get("articulo") or "").strip()
        if not articulo:
            continue
        familia = articulo[2:6] if len(articulo) >= 6 else articulo
        if familia not in ventas_por_familia:
            ventas_por_familia[familia] = {
                "familia": familia,
                "total": 0,
                "saldo_total": 0,
                "articulos": {},
                "sufijos": {},
                "sufijos_saldo": {},
                "tallas": {},
                "tallas_corte": {},
                "articulos_tallas": {},
            }
        if articulo not in ventas_por_familia[familia]["articulos_tallas"]:
            ventas_por_familia[familia]["articulos_tallas"][articulo] = {
                "articulo": articulo,
                "vendida": {},
                "cortada": {},
            }
        for item in r.get("tallas_items") or []:
            talla = int(item.get("talla") or 0)
            cantidad = int(item.get("cantidad") or 0)
            if talla > 0:
                ventas_por_familia[familia]["tallas_corte"][talla] = (
                    int(ventas_por_familia[familia]["tallas_corte"].get(talla) or 0) + cantidad
                )
                ventas_por_familia[familia]["articulos_tallas"][articulo]["cortada"][talla] = (
                    int(ventas_por_familia[familia]["articulos_tallas"][articulo]["cortada"].get(talla) or 0)
                    + cantidad
                )

    def _sort_sufijo_keys(keys):
        prioridad = {"00": 0, "01": 1, "60": 2}

        def _key(suf):
            s = str(suf or "")
            if s in prioridad:
                return (0, prioridad[s], 0, s)
            if s.isdigit():
                return (1, 0, int(s), s)
            return (2, 0, 0, s)

        return sorted(keys, key=_key)

    def _sort_sufijos_saldo(items):
        prioridad = {"00": 0, "01": 1, "60": 2}

        def _item_key(s):
            suf = str(s.get("sufijo") or "")
            if suf in prioridad:
                return (0, prioridad[suf], 0, suf)
            if suf.isdigit():
                return (1, 0, int(suf), suf)
            return (2, 0, 0, suf)

        return sorted(
            items,
            key=_item_key,
        )

    def _build_sufijos_comp(g):
        pedidos_by_sufijo = {}
        saldo_by_sufijo = {}
        for s in g["sufijos"].values():
            suf = str(s.get("sufijo") or "")
            pedidos_by_sufijo[suf] = pedidos_by_sufijo.get(suf, 0) + int(s.get("total") or 0)
        for s in g["sufijos_saldo"].values():
            suf = str(s.get("sufijo") or "")
            saldo_by_sufijo[suf] = saldo_by_sufijo.get(suf, 0) + int(s.get("total") or 0)
        sufijos_ordenados = _sort_sufijo_keys(
            set(pedidos_by_sufijo.keys()) | set(saldo_by_sufijo.keys())
        )
        return [
            {
                "sufijo": suf,
                "pedidos_total": pedidos_by_sufijo.get(suf, 0),
                "saldo_total": saldo_by_sufijo.get(suf, 0),
            }
            for suf in sufijos_ordenados
        ]

    ventas_grouped = sorted(
        [
            {
                "familia": g["familia"],
                "total": g["total"],
                "saldo_total": g["saldo_total"],
                "cortado_total": sum(int(v) for v in (g.get("tallas_corte") or {}).values()),
                "articulos": sorted(
                    g["articulos"].values(),
                    key=lambda v: v["total"],
                    reverse=True,
                ),
                "sufijos": sorted(
                    g["sufijos"].values(),
                    key=lambda s: s["total"],
                    reverse=True,
                ),
                "sufijos_saldo": _sort_sufijos_saldo(
                    list(g["sufijos_saldo"].values())
                ),
                "sufijos_comp": _build_sufijos_comp(g),
                "tallas": sorted(
                    [
                        {
                            "talla": talla,
                            "vendida": int((g.get("tallas") or {}).get(talla) or 0),
                            "cortada": int((g.get("tallas_corte") or {}).get(talla) or 0),
                        }
                        for talla in sorted(
                            set((g.get("tallas") or {}).keys()) | set((g.get("tallas_corte") or {}).keys())
                        )
                    ],
                    key=lambda x: x["talla"],
                ),
                "articulos_tallas": sorted(
                    [
                        {
                            "articulo": a.get("articulo"),
                            "total": int(
                                (
                                    (
                                        g.get("articulos") or {}
                                    ).get(a.get("articulo")) or {}
                                ).get("total")
                                or 0
                            ),
                            "tallas": [
                                {
                                    "talla": talla,
                                    "vendida": int((a.get("vendida") or {}).get(talla) or 0),
                                    "cortada": int((a.get("cortada") or {}).get(talla) or 0),
                                }
                                for talla in sorted(
                                    set((a.get("vendida") or {}).keys()) | set((a.get("cortada") or {}).keys())
                                )
                            ],
                        }
                        for a in (g.get("articulos_tallas") or {}).values()
                    ],
                    key=lambda x: str(x.get("articulo") or ""),
                ),
            }
            for g in ventas_por_familia.values()
        ],
        key=lambda g: g["total"],
        reverse=True,
    )
    ventas_top_familia = ventas_grouped[0] if ventas_grouped else None
    ventas_tallas = sorted(
        [{"talla": talla, "total": total} for talla, total in ventas_por_talla.items()],
        key=lambda x: x["talla"],
    )
    ventas_top_talla = max(ventas_tallas, key=lambda x: x["total"]) if ventas_tallas else None
    ventas_top_articulo = None
    if ventas_por_articulo:
        top_articulo, top_total = max(ventas_por_articulo.items(), key=lambda x: x[1])
        ventas_top_articulo = {"articulo": top_articulo, "total": int(top_total)}
    ventas_top_articulos = sorted(
        [
            {"articulo": articulo, "total": int(total)}
            for articulo, total in ventas_por_articulo.items()
        ],
        key=lambda x: x["total"],
        reverse=True,
    )
    corte_por_articulo = {
        str(row.get("articulo") or "").strip(): {
            "cortado_total": int(row.get("total") or 0),
            "cortado_tallas_detalle": str(row.get("tallas_detalle") or "-"),
        }
        for row in corte_rows
        if str(row.get("articulo") or "").strip()
    }
    bodega_rows = []
    for row in rows:
        if int(row.get("bodega") or 0) <= 0:
            continue
        articulo = str(row.get("articulo") or "").strip()
        corte_info = corte_por_articulo.get(articulo) or {}
        row["cortado_total"] = int(corte_info.get("cortado_total") or 0)
        row["cortado_tallas_detalle"] = str(corte_info.get("cortado_tallas_detalle") or "-")
        bodega_rows.append(row)
    trazabilidad_labels = [
        ("bodega", "Bodega"),
        ("corte_1", "Corte"),
        ("taller", "Taller"),
        ("t_externo", "T. Externo"),
        ("limpiado", "Limpiado"),
        ("lavanderia", "Lavanderia"),
        ("terminacion", "Terminacion"),
        ("segunda", "Segunda"),
    ]
    trazabilidad_por_articulo: dict[str, list[dict]] = {}
    for row in rows:
        articulo = str(row.get("articulo") or "").strip()
        if not articulo:
            continue
        partidas = [
            {
                "etapa": label,
                "cantidad": int(row.get(key) or 0),
            }
            for key, label in trazabilidad_labels
            if int(row.get(key) or 0) > 0
        ]
        trazabilidad_por_articulo.setdefault(articulo, []).append(
            {
                "articulo": articulo,
                "corte": str(row.get("corte") or "").strip(),
                "fecha": str(row.get("fecha_display") or "-"),
                "proceso_actual": str(row.get("proceso_actual") or "Sin movimiento"),
                "cantidad_total": int(row.get("proceso") or 0),
                "cantidad_bodega": int(row.get("bodega") or 0),
                "cantidad_pendiente": int(row.get("pendiente_en_trazabilidad") or 0),
                "restante_detalle": str(row.get("restante_detalle") or "Sin restante"),
                "partidas": partidas,
            }
        )
    for articulo, partidas in trazabilidad_por_articulo.items():
        trazabilidad_por_articulo[articulo] = sorted(
            partidas,
            key=lambda item: (
                0 if item["proceso_actual"] == "BODEGA" else 1,
                -int(item["cantidad_total"] or 0),
                str(item["corte"] or ""),
            ),
        )
    bodega_total = sum(int(row.get("proceso") or 0) for row in bodega_rows)
    bodega_en_bodega = sum(int(row.get("bodega") or 0) for row in bodega_rows)
    bodega_restante = sum(int(row.get("pendiente_en_trazabilidad") or 0) for row in bodega_rows)
    bodega_con_etapas = sum(1 for row in bodega_rows if str(row.get("etapas_fechas_detalle") or "-") != "-")
    muestras_rows = [
        row
        for row in rows
        if str(row.get("corte", "")).lstrip("0").startswith("96")
    ]
    muestras_total = sum(int(row.get("proceso") or 0) for row in muestras_rows)
    muestras_bodega = sum(int(row.get("bodega") or 0) for row in muestras_rows)
    muestras_restante = max(muestras_total - muestras_bodega, 0)
    muestras_con_etapas = sum(1 for row in muestras_rows if str(row.get("etapas_fechas_detalle") or "-") != "-")
    upload_debug = session.pop("upload_debug", "")
    proyeccion_state = _load_proyeccion_state()
    production_goals = _build_production_goals_summary()
    production_daily_dashboard = _build_new_section_dashboard()
    excel_preview_dashboard = _build_excel_preview_dashboard()
    local_preview_enabled = _is_local_request()
    ventas_docs_summary = _load_ventas_docs_summary()
    disponibles_summary = _load_disponibles_ranking_4200(ventas_top_articulos)
    inventory_book = _load_inventory_book_dashboard()
    full_table_rows, full_table_totals, full_table_temporadas = _load_full_table_rows_from_seed()
    inventory_manage_enabled = _can_upload() and _portal_section() == "web"
    return render_template(
        "index.html",
        rows=rows,
        totals=totals,
        summary=summary,
        pedidos_sections=pedidos_sections,
        ventas_total=ventas_total,
        ventas_grouped=ventas_grouped,
        ventas_top_familia=ventas_top_familia,
        ventas_tallas=ventas_tallas,
        ventas_top_talla=ventas_top_talla,
        ventas_top_articulo=ventas_top_articulo,
        ventas_top_articulos=ventas_top_articulos,
        disponibles_summary=disponibles_summary,
        ventas_trazabilidad_por_articulo=trazabilidad_por_articulo,
        exs_summary=exs_summary,
        comparativo_summary=comparativo_summary,
        search_error=search_error,
        search_success=search_success,
        filters=filters,
        bodega_rows=bodega_rows,
        bodega_total=bodega_total,
        bodega_en_bodega=bodega_en_bodega,
        bodega_restante=bodega_restante,
        bodega_con_etapas=bodega_con_etapas,
        muestras_rows=muestras_rows,
        muestras_total=muestras_total,
        muestras_bodega=muestras_bodega,
        muestras_restante=muestras_restante,
        muestras_con_etapas=muestras_con_etapas,
        upload_debug=upload_debug,
        proyeccion_state=proyeccion_state,
        production_goals=production_goals,
        production_daily_dashboard=production_daily_dashboard,
        excel_preview_dashboard=excel_preview_dashboard,
        local_preview_enabled=local_preview_enabled,
        can_upload=_can_upload(),
        inventory_manage_enabled=inventory_manage_enabled,
        admin_key_enabled=bool(_admin_key()),
        assistant_enabled=ASSISTANT_ENABLED,
        assistant_provider=assistant_provider,
        ventas_docs_summary=ventas_docs_summary,
        inventory_book=inventory_book,
        full_table_rows=full_table_rows,
        full_table_totals=full_table_totals,
        full_table_temporadas=full_table_temporadas,
    )


@app.post("/upload-proyeccion")
def upload_proyeccion():
    if not _can_upload():
        flash("Acceso denegado para actualizar proyeccion.", "error")
        return redirect(url_for("index"))
    try:
        monthly_goal = int(
            str(request.form.get("weekly_goal") or request.form.get("monthly_goal") or "12000").strip() or "12000"
        )
        if monthly_goal <= 0:
            raise ValueError("La meta semanal debe ser mayor que 0.")
        file = request.files.get("proyeccion_file")
        if not file or not file.filename:
            raise ValueError("Debes seleccionar una hoja CSV o XLSX.")
        rows = _parse_proyeccion_rows_from_bytes(file.read(), file.filename or "")
        if not rows:
            raise ValueError("No se detectaron filas validas. Usa columnas: area, real y fecha/mes.")
        _save_proyeccion_state(monthly_goal, rows)
        flash(
            f"Proyeccion cargada. Meta semanal: {monthly_goal}. Filas validas: {len(rows)}.",
            "success",
        )
    except Exception as exc:
        flash(f"No se pudo cargar la proyeccion: {exc}", "error")
    return redirect(url_for("index"))


@app.post("/clear-proyeccion")
def clear_proyeccion():
    if not _can_upload():
        flash("Acceso denegado para limpiar proyeccion.", "error")
        return redirect(url_for("index"))
    try:
        if PROYECCION_STATE_PATH.exists():
            PROYECCION_STATE_PATH.unlink()
        flash("Proyeccion reiniciada.", "success")
    except Exception as exc:
        flash(f"No se pudo limpiar la proyeccion: {exc}", "error")
    return redirect(url_for("index"))


@app.post("/upload")
def upload():
    if not _can_upload():
        flash("Acceso denegado para cargar archivos.", "error")
        return redirect(url_for("index"))

    try:
        file = request.files.get("file")
        if not file or not file.filename:
            flash("No se pudo cargar la data. Intentelo nuevamente.", "error")
            return redirect(url_for("index"))

        parsed = parse_uploaded_file(file)
        kind = parsed["kind"]
        rows = parsed["rows"]
        if not rows:
            flash("No se encontraron filas validas en el archivo. Verifique formato e intentelo nuevamente.", "error")
            return redirect(url_for("index"))
        if kind == "pedidos_talla":
            stats = import_pedidos_talla_rows(DB_PATH, rows)
        elif kind == "pedidos_talla_todas":
            stats = import_pedidos_talla_todas_rows(DB_PATH, rows)
        elif kind == "corte_etapas":
            stats = import_corte_etapas_rows(DB_PATH, rows)
        elif kind == "exs_map":
            stats = import_exs_map_rows(DB_PATH, rows)
        elif kind == "comparativo_clientes":
            stats = import_comparativo_clientes_rows(DB_PATH, rows)
        elif kind == "deudas_vencidas":
            stats = import_deuda_clientes_rows(DB_PATH, rows)
        elif kind == "saldos":
            stats = import_rows(DB_PATH, rows, replace_all=False, accumulate_on_conflict=True)
        else:
            stats = import_rows(DB_PATH, rows, replace_all=True)
    except RequestEntityTooLarge:
        session["upload_debug"] = "RequestEntityTooLarge: archivo supera limite ADECOM_MAX_UPLOAD_MB."
        flash("El archivo supera el tamano permitido. Intentelo con un archivo mas liviano.", "error")
        return redirect(url_for("index"))
    except Exception as exc:
        app.logger.exception("Fallo en carga de archivo", exc_info=exc)
        session["upload_debug"] = f"{exc.__class__.__name__}: {exc}"
        flash("No se pudo cargar la data. Intentelo nuevamente.", "error")
        return redirect(url_for("index"))

    session.pop("upload_debug", None)
    flash(
        f"Data cargada con exito. Tipo: {kind}. Leidos: {stats.get('read', 0)} | Insertados: {stats.get('inserted', 0)} | Actualizados: {stats.get('updated', 0)}",
        "success",
    )
    return redirect(url_for("index"))


@app.get("/upload")
def upload_get_redirect():
    return redirect(url_for("index"))


@app.post("/upload/refresh-web")
def upload_refresh_web():
    if not _can_upload():
        flash("Acceso denegado para cargar archivos.", "error")
        return redirect(url_for("index"))

    sources = {
        "SALDOS-SECCI": AUTOLOAD_SALDOS_SOURCE,
        "PEDIDOSXTALLA": AUTOLOAD_PEDIDOS_SOURCE,
        "Grande-Adecom": AUTOLOAD_ETAPAS_SOURCE,
    }
    missing_cfg = [label for label, src in sources.items() if not str(src or "").strip()]
    if missing_cfg:
        session["upload_debug"] = (
            "Faltan variables de entorno para actualizacion web: "
            "ADECOM_AUTOLOAD_SALDOS_SOURCE, ADECOM_AUTOLOAD_PEDIDOS_SOURCE, "
            "ADECOM_AUTOLOAD_ETAPAS_SOURCE."
        )
        flash(
            f"Configuracion incompleta para actualizacion web. Falta: {', '.join(missing_cfg)}.",
            "error",
        )
        return redirect(url_for("index"))

    try:
        global _last_sources_signature
        stats = _refresh_web_data()
        _last_sources_signature = _sources_signature()
        stats_saldos = stats["saldos"]
        stats_pedidos = stats["pedidos"]
        stats_etapas = stats["etapas"]
        stats_comparativo = stats["comparativo"]

        session.pop("upload_debug", None)
        success_msg = (
            "Actualizacion web completa. "
            f"SALDOS-SECCI: I {stats_saldos.get('inserted', 0)} / A {stats_saldos.get('updated', 0)}. "
            f"PEDIDOSXTALLA: I {stats_pedidos.get('inserted', 0)} / A {stats_pedidos.get('updated', 0)}. "
            f"Grande-Adecom: I {stats_etapas.get('inserted', 0)} / A {stats_etapas.get('updated', 0)}."
        )
        if AUTOLOAD_COMPARATIVO_SOURCE:
            success_msg += (
                f" COMPARATIVO: I {stats_comparativo.get('inserted', 0)} / "
                f"A {stats_comparativo.get('updated', 0)}."
            )
        flash(success_msg, "success")
    except url_error.URLError as exc:
        app.logger.exception("Error de red en actualizacion web", exc_info=exc)
        session["upload_debug"] = f"URLError: {exc}"
        flash("No se pudo descargar una o mas fuentes web.", "error")
    except Exception as exc:
        app.logger.exception("Fallo en actualizacion web", exc_info=exc)
        session["upload_debug"] = f"{exc.__class__.__name__}: {exc}"
        flash("No se pudo actualizar la data web. Intentelo nuevamente.", "error")
    return redirect(url_for("index"))


@app.post("/upload/refresh-local")
def upload_refresh_local():
    if not _can_upload():
        flash("Acceso denegado para cargar archivos.", "error")
        return redirect(url_for("index"))
    try:
        changed = _refresh_if_sources_changed(force=True)
        if not changed:
            flash("No se encontraron fuentes locales validas para actualizar.", "error")
            return redirect(url_for("index"))
        session.pop("upload_debug", None)
        flash("Actualizacion local completa desde carpeta vigilada.", "success")
    except Exception as exc:
        app.logger.exception("Fallo en actualizacion local", exc_info=exc)
        session["upload_debug"] = f"{exc.__class__.__name__}: {exc}"
        flash("No se pudo actualizar la data local.", "error")
    return redirect(url_for("index"))


@app.post("/inventory/manage/save")
def inventory_manage_save():
    if not _can_upload():
        flash("Acceso denegado para gestionar stock.", "error")
        return redirect(url_for("index"))
    try:
        row_id_raw = str(request.form.get("item_id") or "").strip()
        row_id = int(row_id_raw) if row_id_raw.isdigit() else None
        payload = _inventory_row_from_form(request.form)
        if not str(payload.get("articulo") or "").strip():
            raise ValueError("El articulo es obligatorio.")
        stats = save_inventory_stock_row(DB_PATH, payload, row_id=row_id)
        if int(stats.get("updated") or 0) > 0:
            flash("Stock actualizado correctamente.", "success")
        else:
            flash("Stock agregado correctamente.", "success")
    except Exception as exc:
        flash(f"No se pudo guardar el stock: {exc}", "error")
    return redirect(url_for("index"))


@app.post("/inventory/manage/delete")
def inventory_manage_delete():
    if not _can_upload():
        flash("Acceso denegado para gestionar stock.", "error")
        return redirect(url_for("index"))
    try:
        row_id_raw = str(request.form.get("item_id") or "").strip()
        row_id = int(row_id_raw) if row_id_raw.isdigit() else 0
        if row_id <= 0:
            raise ValueError("ID de registro invalido.")
        deleted = delete_inventory_stock_row(DB_PATH, row_id)
        if deleted:
            flash("Registro de stock eliminado.", "success")
        else:
            flash("No se encontro el registro de stock.", "error")
    except Exception as exc:
        flash(f"No se pudo eliminar el stock: {exc}", "error")
    return redirect(url_for("index"))


@app.post("/inventory/manage/sync-excel")
def inventory_manage_sync_excel():
    if not _can_upload():
        flash("Acceso denegado para sincronizar stock.", "error")
        return redirect(url_for("index"))
    try:
        parsed_rows, _, err = _load_inventory_rows_from_excel()
        if not parsed_rows:
            raise ValueError(err or "No se detectaron filas validas en el inventario.")
        stats = replace_inventory_stock_rows(DB_PATH, parsed_rows)
        flash(
            f"Stock sincronizado desde Excel. Filas: {stats.get('inserted', 0)}.",
            "success",
        )
    except Exception as exc:
        flash(f"No se pudo sincronizar desde Excel: {exc}", "error")
    return redirect(url_for("index"))


@app.post("/admin/login")
def admin_login():
    key = _admin_key()
    if not key:
        flash("La clave de administrador no esta configurada.", "error")
        return redirect(url_for("index"))
    entered = str(request.form.get("admin_key") or "").strip()
    if entered and entered == key:
        session["can_upload"] = True
        flash("Modo carga activado.", "success")
    else:
        flash("Clave incorrecta.", "error")
    return redirect(url_for("index"))


@app.post("/admin/logout")
def admin_logout():
    session.pop("can_upload", None)
    flash("Modo carga desactivado.", "success")
    return redirect(url_for("index"))


@app.post("/assistant/query")
def assistant_query():
    if not ASSISTANT_ENABLED:
        return jsonify(
            {
                "answer": "Asistente virtual deshabilitado.",
                "provider": "off",
                "fallback": False,
                "detail": "ADECOM_ASSISTANT_ENABLED=0",
            }
        ), 410
    provider = os.environ.get("ADECOM_ASSISTANT_PROVIDER", "local").strip().lower()
    try:
        payload = request.get_json(silent=True) or {}
        question = str(payload.get("question") or "").strip()
        result = _answer_assistant_router(question)
        return jsonify(result)
    except Exception as exc:
        app.logger.exception("Error en /assistant/query", exc_info=exc)
        if provider in {"gemini", "google"}:
            return jsonify(
                {
                    "answer": f"Gemini no disponible. Detalle: {exc}",
                    "provider": "gemini",
                    "fallback": False,
                    "detail": str(exc),
                }
            ), 503
        return jsonify(
            {
                "answer": "No fue posible responder en este momento. Intenta nuevamente.",
                "provider": "local",
                "fallback": True,
                "detail": str(exc),
            }
        ), 200


@app.get("/export.csv")
def export_csv():
    filters = {
        "q": request.args.get("q", "").strip(),
        "fecha": request.args.get("fecha", "").strip(),
    }
    rows, _, _ = query_rows(DB_PATH, filters)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "articulo",
            "corte",
            "fecha",
            "programa",
            "proceso",
            "bodega",
            "saldo",
            "corte_1",
            "taller",
            "t_externo",
            "limpiado",
            "lavanderia",
            "terminacion",
            "muestra",
            "segunda",
            "taller_nombre",
            "proceso_actual",
            "faltante",
            "restante_fuera_bodega",
            "restante",
        ]
    )
    for row in rows:
        writer.writerow(
            [
                row["articulo"],
                row["corte"],
                row["fecha_display"],
                row["programa"],
                row["proceso"],
                row["bodega"],
                row["saldo"],
                row["corte_1"],
                row["taller"],
                row["t_externo"],
                row["limpiado"],
                row["lavanderia"],
                row["terminacion"],
                row["muestra"],
                row["segunda"],
                row["taller_nombre"],
                row["proceso_actual"],
                row["faltante"],
                row["restante_fuera_bodega"],
                row["restante_detalle"],
            ]
        )

    return Response(
        output.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=saldos_seccion.csv"},
    )


if __name__ == "__main__":
    init_db(DB_PATH)
    debug_mode = os.environ.get("ADECOM_DEBUG", "0").strip() == "1"
    app.run(
        host="127.0.0.1",
        port=int(os.environ.get("PORT", "5000")),
        debug=debug_mode,
        use_reloader=False,
    )

